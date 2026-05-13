"""
Run this script to generate sample_expense_report.xlsx for testing.
Usage: python generate_sample.py
"""
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    print("Install openpyxl first: pip install openpyxl")
    raise

ROWS = [
    ("Rahul Sharma",   "EMP001", "15/04/2024", "Travel",       "Cab to Delhi Airport",        1250.00, "bill_001.pdf"),
    ("Priya Mehta",    "EMP002", "16/04/2024", "Food",         "Team lunch at Olive",          3200.00, "bill_002.pdf"),
    ("Arjun Singh",    "EMP003", "17/04/2024", "Accommodation","Hotel stay Noida",            4500.00, "bill_003.pdf"),
    ("Sneha Patel",    "EMP004", "18/04/2024", "Travel",       "Train ticket Mumbai",         2100.00, "bill_004.pdf"),
    ("Vikram Nair",    "EMP005", "19/04/2024", "Office Supply","Printer cartridges",           850.00, "bill_005.pdf"),
    ("Rahul Sharma",   "EMP001", "20/04/2024", "Food",         "Client dinner",               5600.00, "bill_006.pdf"),
    ("Ananya Roy",     "EMP006", "21/04/2024", "Travel",       "Flight to Bangalore",         8400.00, "bill_007.pdf"),
    ("Priya Mehta",    "EMP002", "22/04/2024", "Office Supply","Whiteboard markers",           350.00, "bill_002.pdf"),  # duplicate bill
    ("Kabir Hassan",   "EMP007", "23/04/2024", "Accommodation","Hotel ITC Bangalore",         7200.00, ""),             # missing bill
    ("Deepa Krishnan", "EMP008", "24/04/2024", "Travel",       "Taxi reimbursement",           600.00, "bill_009.pdf"), # bill not in ZIP
]

HEADERS = [
    "Employee Name", "Employee ID", "Expense Date", "Expense Category",
    "Description", "Claimed Amount", "Attachment Name",
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Expense Report"

header_fill = PatternFill("solid", fgColor="1F4E79")
header_font = Font(bold=True, color="FFFFFF", size=11)

for col_i, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=col_i, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

for row_i, row_data in enumerate(ROWS, start=2):
    for col_i, value in enumerate(row_data, start=1):
        ws.cell(row=row_i, column=col_i, value=value)

for col in ws.columns:
    max_len = max(len(str(c.value or "")) for c in col)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

out = os.path.join(os.path.dirname(__file__), "sample_expense_report.xlsx")
wb.save(out)
print(f"Saved: {out}")
print("\nNOTE: You also need a sample_bills.zip with these files:")
for row in ROWS:
    if row[6]:
        print(f"  {row[6]}")
print("\nFor testing, create dummy PDFs or use any PDF/image files renamed to match the filenames above.")
print("Place them in a ZIP at: samples/sample_bills.zip")
