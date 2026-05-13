import os
from typing import List
from models.schemas import ExpenseRow, ExpenseStatus

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


STATUS_COLORS = {
    ExpenseStatus.APPROVED: "C6EFCE",   # green
    ExpenseStatus.REJECTED: "FFC7CE",   # red
    ExpenseStatus.FLAGGED:  "FFEB9C",   # yellow
}

HEADER_FILL = "1F4E79"
HEADER_FONT_COLOR = "FFFFFF"

COLUMNS = [
    ("Row #",            "row_index"),
    ("Employee Name",    "employee_name"),
    ("Employee ID",      "employee_id"),
    ("Expense Date",     "expense_date"),
    ("Expense Nature",   "expense_nature"),       # tagged category
    ("Category (Raw)",   "expense_category"),
    ("Description",      "description"),
    ("Claimed Amount ₹", "claimed_amount"),
    ("Bill Amount ₹",    "bill_amount"),
    ("Difference ₹",     "amount_diff"),          # claimed - bill
    ("Matched Bill",     "matched_file"),
    ("Status",           "status"),
    ("Remarks",          "remarks"),
]


def export_to_excel(rows: List[ExpenseRow], output_path: str) -> str:
    if not OPENPYXL_AVAILABLE:
        raise RuntimeError("openpyxl not installed")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header row
    for col_idx, (header, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 22

    # Data rows
    for r_idx, row in enumerate(rows, start=2):
        status_color = STATUS_COLORS.get(row.status, "FFFFFF")
        fill = PatternFill("solid", fgColor=status_color)

        diff = row.amount_diff
        diff_display = (
            f"+₹{diff:,.2f} (over)" if diff is not None and diff > 0
            else f"-₹{abs(diff):,.2f} (under)" if diff is not None and diff < 0
            else "₹0.00" if diff == 0
            else "N/A"
        )

        values = {
            "row_index":        row.row_index,
            "employee_name":    row.employee_name,
            "employee_id":      row.employee_id,
            "expense_date":     row.expense_date,
            "expense_nature":   row.expense_nature or row.expense_category,
            "expense_category": row.expense_category,
            "description":      row.description,
            "claimed_amount":   row.claimed_amount,
            "bill_amount":      row.bill_amount if row.bill_amount is not None else "N/A",
            "amount_diff":      diff_display,
            "matched_file":     row.matched_file or "—",
            "status":           row.status.value,
            "remarks":          "; ".join(row.remarks) if row.remarks else "—",
        }

        for col_idx, (_, field) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=r_idx, column=col_idx, value=values[field])
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto-size columns
    for col_idx in range(1, len(COLUMNS) + 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(rows) + 2)
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 50)

    # Freeze header
    ws.freeze_panes = "A2"

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_summary = wb.create_sheet("Summary")
    approved = sum(1 for r in rows if r.status == ExpenseStatus.APPROVED)
    rejected = sum(1 for r in rows if r.status == ExpenseStatus.REJECTED)
    flagged  = sum(1 for r in rows if r.status == ExpenseStatus.FLAGGED)
    total_claimed = sum(r.claimed_amount for r in rows)

    summary_data = [
        ("Metric", "Value"),
        ("Total Claims", len(rows)),
        ("Approved", approved),
        ("Rejected", rejected),
        ("Flagged", flagged),
        ("", ""),
        ("Total Claimed Amount (₹)", f"{total_claimed:,.2f}"),
        ("Approved Amount (₹)",      f"{sum(r.claimed_amount for r in rows if r.status == ExpenseStatus.APPROVED):,.2f}"),
        ("Rejected Amount (₹)",      f"{sum(r.claimed_amount for r in rows if r.status == ExpenseStatus.REJECTED):,.2f}"),
    ]

    for r_idx, (label, value) in enumerate(summary_data, start=1):
        ws_summary.cell(row=r_idx, column=1, value=label).font = Font(bold=r_idx == 1)
        ws_summary.cell(row=r_idx, column=2, value=value)

    ws_summary.column_dimensions["A"].width = 34
    ws_summary.column_dimensions["B"].width = 22

    # ── Expense Nature breakdown sheet ────────────────────────────────────────
    ws_nature = wb.create_sheet("By Expense Nature")
    nature_header = ["Expense Nature", "Total Claims", "Approved", "Rejected", "Flagged",
                     "Total Claimed (₹)", "Approved Amount (₹)", "Rejected Amount (₹)"]
    for c, h in enumerate(nature_header, 1):
        cell = ws_nature.cell(row=1, column=c, value=h)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR)
        cell.alignment = Alignment(horizontal="center")

    # Group by expense_nature
    from collections import defaultdict
    nature_groups: dict = defaultdict(list)
    for row in rows:
        nature_groups[row.expense_nature or row.expense_category or "Miscellaneous"].append(row)

    for row_i, (nature, group_rows) in enumerate(sorted(nature_groups.items()), start=2):
        g_approved = sum(1 for r in group_rows if r.status == ExpenseStatus.APPROVED)
        g_rejected = sum(1 for r in group_rows if r.status == ExpenseStatus.REJECTED)
        g_flagged  = sum(1 for r in group_rows if r.status == ExpenseStatus.FLAGGED)
        g_total    = sum(r.claimed_amount for r in group_rows)
        g_app_amt  = sum(r.claimed_amount for r in group_rows if r.status == ExpenseStatus.APPROVED)
        g_rej_amt  = sum(r.claimed_amount for r in group_rows if r.status == ExpenseStatus.REJECTED)

        row_data = [nature, len(group_rows), g_approved, g_rejected, g_flagged,
                    round(g_total, 2), round(g_app_amt, 2), round(g_rej_amt, 2)]
        for c, val in enumerate(row_data, 1):
            ws_nature.cell(row=row_i, column=c, value=val)

    for c in range(1, len(nature_header) + 1):
        ws_nature.column_dimensions[get_column_letter(c)].width = 22

    wb.save(output_path)
    return output_path
