import os
import sys
import json
import shutil
import asyncio
import logging
import tempfile
import uuid
import zipfile
from datetime import datetime
from typing import Optional

log = logging.getLogger(__name__)

# Load .env from project root
try:
    from dotenv import load_dotenv
    _env = os.path.join(os.path.dirname(__file__), "..", ".env")
    load_dotenv(dotenv_path=os.path.abspath(_env))
except Exception:
    pass

from fastapi import FastAPI, UploadFile, File, Form as _FastAPIForm, HTTPException, BackgroundTasks, Query, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

from models.schemas import SessionResult, ExpenseStatus

from services.zip_extractor import extract_zip
from services.ocr import process_bill_ocr
from services.matcher import read_excel_rows, match_bills_to_rows, match_by_amount_fallback
from services.validator import validate_expenses
from services.exporter import export_to_excel
from services.sheets import push_to_google_sheets
from services.zoho import push_approved_to_zoho
from services.zoho_verifier import verify_zoho_entries

app = FastAPI(title="Expense Validator API", version="1.0.0", docs_url="/docs")


@app.on_event("startup")
async def _warmup():
    """Init DB, pre-load OCR models."""
    from services.db import init_db
    init_db()
    try:
        from services.ocr import RAPIDOCR_AVAILABLE, _rapid_ocr
        if RAPIDOCR_AVAILABLE:
            import numpy as np
            _rapid_ocr(np.zeros((64, 256, 3), dtype=np.uint8))
    except Exception:
        pass

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

UPLOAD_BASE = os.environ.get("UPLOAD_DIR", os.path.join(tempfile.gettempdir(), "expense_validator"))
os.makedirs(UPLOAD_BASE, exist_ok=True)

_BASE_TEMP = tempfile.gettempdir()

# In-memory session store (replace with Redis for multi-worker prod)
_sessions: dict = {}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _run_in_proactor_thread(fn, *args, **kwargs):
    """
    Run *fn* in a brand-new thread that owns an explicit asyncio.ProactorEventLoop.

    Why this is needed on Windows:
      FastAPI sync-endpoint threads (threadpool) already have a SelectorEventLoop
      assigned.  SelectorEventLoop cannot create subprocesses, so Playwright's
      sync_playwright() raises NotImplementedError when it tries to launch the
      browser.  Spawning a *fresh* thread and calling
          asyncio.set_event_loop(asyncio.ProactorEventLoop())
      before any Playwright call gives it a loop that supports subprocess creation.
    """
    import threading
    import queue as _queue

    q = _queue.Queue()

    def _worker():
        if sys.platform == "win32":
            # Set the policy FIRST so that Playwright's internal asyncio.new_event_loop()
            # also returns a ProactorEventLoop (uvicorn sets WindowsSelectorEventLoopPolicy
            # globally — overriding just the loop is not enough; policy must also be reset
            # in this thread before Playwright's __enter__ calls new_event_loop()).
            asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())
            loop = asyncio.ProactorEventLoop()
            asyncio.set_event_loop(loop)
        try:
            result = fn(*args, **kwargs)
            q.put(("ok", result))
        except Exception as exc:
            import traceback as _tb
            log.error("Proactor-thread error in %s: %s\n%s",
                      getattr(fn, "__name__", fn), exc, _tb.format_exc())
            q.put(("err", exc))

    t = threading.Thread(target=_worker, daemon=True)
    t.start()
    t.join(timeout=300)          # 5-minute hard cap
    if t.is_alive():
        raise RuntimeError("Browser automation timed out (>300 s)")
    try:
        kind, val = q.get_nowait()
    except Exception:
        raise RuntimeError("Browser thread returned no result")
    if kind == "err":
        raise val
    return val


def _session_dir(session_id: str) -> str:
    return os.path.join(UPLOAD_BASE, session_id)


def _save_session(session_id: str, result: SessionResult):
    # Always recompute summary counts from actual row statuses before saving.
    # This ensures the dashboard never shows stale approved/rejected/flagged totals
    # after manual row edits or keka actions.
    result.approved     = sum(1 for r in result.rows if str(r.status) == "Approved")
    result.rejected     = sum(1 for r in result.rows if str(r.status) == "Rejected")
    result.flagged      = sum(1 for r in result.rows if str(r.status) == "Flagged")
    result.total_claims = len(result.rows)
    _sessions[session_id] = result
    # Also persist to disk for restart resilience
    path = os.path.join(_session_dir(session_id), "result.json")
    with open(path, "w", encoding="utf-8") as f:
        f.write(result.model_dump_json())


def _load_session(session_id: str) -> Optional[SessionResult]:
    if session_id in _sessions:
        return _sessions[session_id]
    path = os.path.join(_session_dir(session_id), "result.json")
    if os.path.exists(path):
        # Try UTF-8 first; fall back to Windows-1252 for older sessions
        for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
            try:
                with open(path, encoding=enc) as f:
                    data = json.load(f)
                break
            except (UnicodeDecodeError, json.JSONDecodeError):
                continue
        else:
            return None
        # Always recompute approved/rejected/flagged from actual row statuses
        # so stale top-level counts never mislead the dashboard.
        rows_data = data.get("rows", [])
        data["approved"] = sum(1 for r in rows_data if r.get("status") == "Approved")
        data["rejected"] = sum(1 for r in rows_data if r.get("status") == "Rejected")
        data["flagged"]  = sum(1 for r in rows_data if r.get("status") == "Flagged")
        data["total_claims"] = len(rows_data)
        result = SessionResult(**data)
        _sessions[session_id] = result
        return result
    return None


def _get_actor(authorization: Optional[str]) -> str:
    """Extract username from Bearer token, or 'anonymous'."""
    try:
        if authorization and authorization.lower().startswith("bearer "):
            from services.auth import verify_token
            p = verify_token(authorization[7:].strip())
            if p:
                return p.get("u", "anonymous")
    except Exception:
        pass
    return "anonymous"


def _require_admin(authorization: Optional[str]) -> None:
    """
    Raise HTTP 401/403 unless the caller is an authenticated admin.
    Call this at the top of every write endpoint that should be admin-only.
    """
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "Authentication required")
    payload = verify_token(authorization[7:].strip())
    if not payload:
        raise HTTPException(401, "Invalid or expired token")
    if payload.get("role", "reviewer") != "admin":
        raise HTTPException(403, "Admin access required")


# ---------------------------------------------------------------------------
# Policy rules helpers
# ---------------------------------------------------------------------------

POLICY_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "policy_rules.json")

DEFAULT_POLICY = {
    "enabled": False,
    "rules": [
        {"category": "food", "label": "Food & Meals", "limit": 500, "per": "claim", "enabled": False},
        {"category": "travel", "label": "Travel & Transport", "limit": 5000, "per": "claim", "enabled": False},
        {"category": "accommodation", "label": "Hotel & Accommodation", "limit": 3000, "per": "claim", "enabled": False},
        {"category": "subscription", "label": "Subscriptions", "limit": 2000, "per": "claim", "enabled": False},
        {"category": "office", "label": "Office Supplies", "limit": 1000, "per": "claim", "enabled": False},
        {"category": "other", "label": "Other / Miscellaneous", "limit": 5000, "per": "claim", "enabled": False},
    ],
}


def _load_policy_rules() -> dict:
    try:
        if os.path.exists(POLICY_FILE):
            with open(POLICY_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return DEFAULT_POLICY


# ---------------------------------------------------------------------------
# Background processing
# ---------------------------------------------------------------------------

async def run_validation(session_id: str, excel_path: str, zip_path: str, skip_row_indices: list = None):
    result = SessionResult(session_id=session_id, processing_status="processing", current_step="Reading Excel…")
    _save_session(session_id, result)

    def _step(msg: str):
        result.current_step = msg
        _save_session(session_id, result)

    try:
        bills_dir = os.path.join(_session_dir(session_id), "bills")

        # Step 1: Parse Excel
        _step("Reading Excel…")
        rows, warnings = read_excel_rows(excel_path)

        # Skip rows that were already processed in previous batches
        if skip_row_indices:
            skip_set = set(skip_row_indices)
            original_count = len(rows)
            rows = [r for r in rows if r.row_index not in skip_set]
            skipped_count = original_count - len(rows)
            if skipped_count:
                _step(f"Skipped {skipped_count} duplicate claim(s) from previous batches…")

        # Step 2: Extract ZIP (preserves employee-ID folder structure)
        _step("Extracting ZIP…")
        by_folder, all_files = extract_zip(zip_path, bills_dir)

        # Step 3: First-pass match — employee ID folder → filename match
        _step("Matching bills to rows…")
        rows, unmapped_bills = match_bills_to_rows(rows, by_folder, all_files)

        # Step 4: OCR ALL bills — deduplicate by real path, then run in parallel
        loop = asyncio.get_event_loop()
        ocr_map: dict = {}

        # Deduplicate: map unique filepath → list of display_keys that share it
        path_to_keys: dict = {}
        for display_key, filepath in all_files.items():
            path_to_keys.setdefault(filepath, []).append(display_key)
        path_hints: dict = {}
        for row in rows:
            for key in (row.matched_files or ([row.matched_file] if row.matched_file else [])):
                fp = all_files.get(key)
                if fp and row.claimed_amount:
                    path_hints.setdefault(fp, float(row.claimed_amount))

        unique_paths = list(path_to_keys.keys())
        _step(f"Running OCR on {len(unique_paths)} bills…")

        # Run OCR concurrently — limit workers to avoid OOM on Railway (512MB RAM)
        # 2 workers max: each PDF load ~20-50MB, so 2 workers keeps peak usage low
        from concurrent.futures import ThreadPoolExecutor
        MAX_OCR_WORKERS = int(os.environ.get("MAX_OCR_WORKERS", "1"))
        MAX_OCR_WORKERS = min(MAX_OCR_WORKERS, len(unique_paths)) if unique_paths else 1
        with ThreadPoolExecutor(max_workers=MAX_OCR_WORKERS) as pool:
            tasks = {
                filepath: loop.run_in_executor(pool, process_bill_ocr, filepath, path_hints.get(filepath))
                for filepath in unique_paths
            }
            results = await asyncio.gather(*tasks.values(), return_exceptions=True)
            for filepath, ocr_result in zip(tasks.keys(), results):
                if isinstance(ocr_result, Exception):
                    from services.ocr import OCRResult as _OCRResult
                    ocr_result = _OCRResult()
                for dk in path_to_keys[filepath]:
                    ocr_map[dk] = ocr_result

        # Persist OCR cache to disk (single write after all threads finish)
        from services.ocr import flush_ocr_cache
        flush_ocr_cache()

        # Step 5: Second-pass match — amount-based fallback for still-unmatched rows
        _step("Amount-based fallback matching…")
        rows = match_by_amount_fallback(rows, by_folder, all_files, ocr_map)

        # Recompute unmapped after second pass (use all matched_files, not just primary)
        matched_keys = {k for r in rows for k in (r.matched_files or ([r.matched_file] if r.matched_file else []))}
        unmapped_bills = sorted(set(all_files.keys()) - matched_keys)

        # Attach all OCR results to each row (one per matched bill)
        for row in rows:
            row.ocr_results = [ocr_map[k] for k in row.matched_files if k in ocr_map]
            if row.ocr_results:
                row.ocr_result = row.ocr_results[0]  # primary for compat

        # Step 6: Pre-fetch USD/INR rates in parallel (warms disk cache)
        from services.validator import prefetch_usd_rates
        prefetch_usd_rates(rows)

        # Step 7: Validate (amount is primary gate)
        _step(f"Validating {len(rows)} claims…")
        rows = validate_expenses(rows, ocr_map, file_map=all_files)

        # Step 7b: Vendor category suggestion (populate suggested_category per row)
        _step("Running vendor intelligence…")
        try:
            from services.vendor_master import suggest_category as _suggest_cat
            for row in rows:
                if row.ocr_result and row.ocr_result.vendor_name:
                    suggestion = _suggest_cat(row.ocr_result.vendor_name)
                    if suggestion:
                        row.suggested_category = suggestion["category"]
                        row.suggested_category_confidence = suggestion["confidence"]
                        row.suggested_vendor_type = suggestion.get("type")
        except Exception as _e:
            log.warning("Vendor suggestion step failed: %s", _e)

        # Step 7c: Anomaly detection
        _step("Detecting anomalies…")
        try:
            from services.anomaly import detect_anomalies as _detect
            rows = _detect(rows, session_id, UPLOAD_BASE)
        except Exception as _e:
            log.warning("Anomaly detection failed: %s", _e)

        # Step 6: Aggregate counts
        approved = sum(1 for r in rows if r.status == ExpenseStatus.APPROVED)
        rejected = sum(1 for r in rows if r.status == ExpenseStatus.REJECTED)
        flagged  = sum(1 for r in rows if r.status == ExpenseStatus.FLAGGED)

        result = SessionResult(
            session_id=session_id,
            processing_status="completed",
            total_claims=len(rows),
            approved=approved,
            rejected=rejected,
            flagged=flagged,
            rows=rows,
            unmapped_bills=unmapped_bills,
        )
        _save_session(session_id, result)
        try:
            from services.db import update_session_record
            update_session_record(session_id, "completed", len(rows), approved, rejected, flagged)
        except Exception:
            pass

    except Exception as exc:
        result = SessionResult(
            session_id=session_id,
            processing_status="error",
            error=str(exc),
        )
        _save_session(session_id, result)
        try:
            from services.db import update_session_record
            update_session_record(session_id, "error")
        except Exception:
            pass


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@app.get("/health")
def root():
    return {"status": "ok", "service": "Expense Validator API v1.0"}


# ── Authentication endpoints ──────────────────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str


@app.post("/auth/login")
def auth_login(body: LoginRequest, request: Request):
    """Validate credentials → issue a 24-hour session token."""
    from services.auth import authenticate, TOKEN_TTL_SECONDS, verify_token
    from services.db import log_activity, get_user
    token = authenticate(body.username, body.password)
    if not token:
        log_activity(body.username, "login_failed", ip_address=request.client.host if request.client else "")
        raise HTTPException(401, "Invalid username or password")
    payload = verify_token(token)
    role = payload.get("role", "reviewer") if payload else "reviewer"
    log_activity(body.username, "login", ip_address=request.client.host if request.client else "")
    return {
        "token":      token,
        "username":   body.username,
        "role":       role,
        "expires_in": TOKEN_TTL_SECONDS,
    }


from fastapi import Header as _FastAPIHeader

@app.get("/auth/me")
def auth_me(authorization: Optional[str] = _FastAPIHeader(None)):
    """Validate the current Authorization header. Returns user info + role."""
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "No token")
    payload = verify_token(authorization[7:].strip())
    if not payload:
        raise HTTPException(401, "Invalid or expired token")
    username = payload.get("u")
    role = payload.get("role", "reviewer")
    # Always re-read role from DB so old tokens (without role field) get correct role
    try:
        from services.db import get_user
        user = get_user(username)
        if user:
            role = user.get("role", role)
    except Exception:
        pass
    return {
        "username":   username,
        "role":       role,
        "expires_at": payload.get("exp"),
    }


@app.post("/auth/logout")
def auth_logout(request: Request, authorization: Optional[str] = _FastAPIHeader(None)):
    """Log the logout event. Frontend clears localStorage."""
    from services.auth import verify_token
    from services.db import log_activity
    if authorization and authorization.lower().startswith("bearer "):
        payload = verify_token(authorization[7:].strip())
        if payload:
            log_activity(payload.get("u", "?"), "logout",
                         ip_address=request.client.host if request.client else "")
    return {"status": "ok"}


# ── Admin — User Management ───────────────────────────────────────────────────

class CreateUserRequest(BaseModel):
    username:  str
    password:  str
    role:      str = "reviewer"
    full_name: str = ""
    email:     str = ""

class UpdateUserRequest(BaseModel):
    role:      Optional[str]  = None
    full_name: Optional[str]  = None
    email:     Optional[str]  = None
    is_active: Optional[bool] = None
    password:  Optional[str]  = None


@app.get("/admin/users")
def admin_list_users(authorization: Optional[str] = _FastAPIHeader(None)):
    from services.auth import require_admin
    require_admin(authorization)
    from services.db import get_all_users
    return {"users": get_all_users()}


@app.post("/admin/users")
def admin_create_user(body: CreateUserRequest, request: Request,
                      authorization: Optional[str] = _FastAPIHeader(None)):
    from services.auth import require_admin
    admin = require_admin(authorization)
    from services.db import get_user, create_user, log_activity
    if body.role not in ("admin", "reviewer"):
        raise HTTPException(400, "Role must be 'admin' or 'reviewer'")
    if get_user(body.username):
        raise HTTPException(409, f"User '{body.username}' already exists")
    user = create_user(body.username, body.password, body.role, body.full_name, body.email)
    log_activity(admin["u"], "create_user", "user", body.username,
                 {"role": body.role, "full_name": body.full_name},
                 ip_address=request.client.host if request.client else "")
    return user


@app.put("/admin/users/{user_id}")
def admin_update_user(user_id: int, body: UpdateUserRequest, request: Request,
                      authorization: Optional[str] = _FastAPIHeader(None)):
    from services.auth import require_admin
    admin = require_admin(authorization)
    from services.db import update_user, get_user_by_id, log_activity
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if body.role and body.role not in ("admin", "reviewer"):
        raise HTTPException(400, "Role must be 'admin' or 'reviewer'")
    kwargs = {k: v for k, v in body.model_dump().items() if v is not None}
    updated = update_user(user_id, **kwargs)
    log_activity(admin["u"], "update_user", "user", target["username"], kwargs,
                 ip_address=request.client.host if request.client else "")
    return updated


@app.delete("/admin/users/{user_id}")
def admin_delete_user(user_id: int, request: Request,
                      authorization: Optional[str] = _FastAPIHeader(None)):
    from services.auth import require_admin
    admin = require_admin(authorization)
    from services.db import get_user_by_id, delete_user, log_activity
    target = get_user_by_id(user_id)
    if not target:
        raise HTTPException(404, "User not found")
    if target["username"] == admin["u"]:
        raise HTTPException(400, "Cannot delete your own account")
    delete_user(user_id)
    log_activity(admin["u"], "delete_user", "user", target["username"], {},
                 ip_address=request.client.host if request.client else "")
    return {"status": "deleted", "username": target["username"]}


# ── Admin — Activity Logs ──────────────────────────────────────────────────────

@app.get("/admin/logs")
def admin_get_logs(
    limit:    int           = Query(100, ge=1, le=500),
    offset:   int           = Query(0,   ge=0),
    username: Optional[str] = Query(None),
    action:   Optional[str] = Query(None),
    authorization: Optional[str] = _FastAPIHeader(None),
):
    from services.auth import require_admin
    require_admin(authorization)
    from services.db import get_activity_logs, get_log_stats
    logs  = get_activity_logs(limit, offset, username, action)
    stats = get_log_stats()
    return {"logs": logs, "stats": stats}


@app.get("/admin/sessions")
def admin_get_sessions(
    limit:  int = Query(100, ge=1, le=500),
    offset: int = Query(0,   ge=0),
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """List all sessions stored in the database (admin only)."""
    from services.auth import require_admin
    require_admin(authorization)
    from services.db import get_all_sessions, get_sessions_stats
    sessions = get_all_sessions(limit, offset)
    stats    = get_sessions_stats()
    return {"sessions": sessions, "stats": stats}


@app.post("/upload/check-duplicates")
async def upload_check_duplicates(
    excel_file: UploadFile = File(..., description="Expense report (.xlsx)"),
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """
    Pre-flight check: read the Excel, cross-reference against all completed sessions.
    Returns which claims already exist so the user can choose to skip them.
    Does NOT start any processing — purely a read-only scan.
    """
    import tempfile as _tf_mod

    # Save Excel to a temp file
    tmp_path = None
    try:
        with _tf_mod.NamedTemporaryFile(suffix=".xlsx", delete=False) as tf:
            tf.write(await excel_file.read())
            tmp_path = tf.name

        rows, _ = read_excel_rows(tmp_path)
    except Exception as exc:
        raise HTTPException(400, f"Cannot read Excel: {exc}")
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except Exception:
                pass

    # Build a stable dedup key per row
    # Priority: keka_claim_id (UUID) > claim_number > employee_id+date+amount
    def _row_key(obj) -> str:
        if isinstance(obj, dict):
            uid  = (obj.get("keka_claim_id") or "").strip()
            cn   = (obj.get("claim_number")  or "").strip()
            eid  = (obj.get("employee_id")   or "").strip()
            date = (obj.get("expense_date")  or "").strip()
            amt  = str(obj.get("claimed_amount") or "0")
        else:  # ExpenseRow
            uid  = (obj.keka_claim_id or "").strip()
            cn   = (obj.claim_number  or "").strip()
            eid  = (obj.employee_id   or "").strip()
            date = (obj.expense_date  or "").strip()
            amt  = str(obj.claimed_amount or 0)
        if uid: return f"uid:{uid}"
        if cn:  return f"cn:{cn}"
        return  f"comp:{eid}|{date}|{amt}"

    # incoming map: key → ExpenseRow (consume as we find matches)
    incoming: dict = {_row_key(r): r for r in rows}

    duplicates = []

    if os.path.exists(UPLOAD_BASE):
        for sid in os.listdir(UPLOAD_BASE):
            result_path = os.path.join(UPLOAD_BASE, sid, "result.json")
            if not os.path.exists(result_path):
                continue
            try:
                data = None
                for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
                    try:
                        with open(result_path, encoding=enc) as f:
                            data = json.load(f)
                        break
                    except (UnicodeDecodeError, json.JSONDecodeError):
                        continue
                if not data:
                    continue
                if data.get("processing_status") != "completed":
                    continue

                # Session date from file mtime (good enough for display)
                session_date = datetime.fromtimestamp(
                    os.path.getmtime(result_path)
                ).strftime("%d %b %Y")

                for er in data.get("rows", []):
                    k = _row_key(er)
                    if k in incoming:
                        inc = incoming.pop(k)  # consume so we don't double-count
                        duplicates.append({
                            "row_index":         inc.row_index,
                            "claim_number":      inc.claim_number,
                            "employee_name":     inc.employee_name,
                            "employee_id":       inc.employee_id,
                            "expense_date":      inc.expense_date,
                            "claimed_amount":    inc.claimed_amount,
                            "expense_category":  inc.expense_category,
                            "prev_session_id":   sid,
                            "prev_session_date": session_date,
                            "prev_status":       er.get("status", ""),
                        })
            except Exception as e:
                log.warning("check-duplicates: session %s error: %s", sid, e)

    return {
        "total":            len(rows),
        "duplicate_count":  len(duplicates),
        "new_count":        len(rows) - len(duplicates),
        "duplicates":       duplicates,
        "skip_row_indices": [d["row_index"] for d in duplicates],
    }


@app.post("/upload")
async def upload_files(
    background_tasks: BackgroundTasks,
    request: Request,
    excel_file: UploadFile = File(..., description="Expense report (.xlsx)"),
    zip_file: UploadFile = File(..., description="Bill attachments (.zip)"),
    skip_row_indices: Optional[str] = _FastAPIForm(None),  # JSON array e.g. "[2,5,8]"
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """Upload Excel + ZIP, kick off background validation. Returns session_id."""
    if not excel_file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Excel file must be .xlsx or .xls")
    if not zip_file.filename.endswith(".zip"):
        raise HTTPException(400, "Bills file must be .zip")

    session_id = str(uuid.uuid4())
    s_dir = _session_dir(session_id)
    os.makedirs(s_dir, exist_ok=True)

    excel_path = os.path.join(s_dir, "report.xlsx")
    zip_path   = os.path.join(s_dir, "bills.zip")

    # Stream files to disk in chunks — avoids loading entire ZIP into RAM at once
    _CHUNK = 1024 * 1024  # 1 MB chunks
    with open(excel_path, "wb") as f:
        while True:
            chunk = await excel_file.read(_CHUNK)
            if not chunk:
                break
            f.write(chunk)
    with open(zip_path, "wb") as f:
        while True:
            chunk = await zip_file.read(_CHUNK)
            if not chunk:
                break
            f.write(chunk)

    actor  = _get_actor(authorization)
    client_ip = request.client.host if request.client else ""
    try:
        from services.db import create_session_record, log_activity
        create_session_record(session_id, actor, excel_file.filename,
                              source="upload", ip_address=client_ip)
        log_activity(actor, "upload", "session", session_id,
                     {"filename": excel_file.filename}, ip_address=client_ip)
    except Exception:
        pass

    # Parse skip_row_indices (from duplicate-check pre-flight)
    skip_indices = []
    if skip_row_indices:
        try:
            skip_indices = json.loads(skip_row_indices)
        except Exception:
            pass

    background_tasks.add_task(run_validation, session_id, excel_path, zip_path, skip_indices)

    return {"session_id": session_id, "status": "processing"}


@app.get("/status/{session_id}")
def get_status(session_id: str):
    """Poll this endpoint until processing_status is 'completed' or 'error'."""
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")
    return {
        "session_id": session_id,
        "processing_status": result.processing_status,
        "current_step": result.current_step,
        "error": result.error,
        "total_claims": result.total_claims,
        "approved": result.approved,
        "rejected": result.rejected,
        "flagged": result.flagged,
    }


@app.get("/results/{session_id}", response_model=SessionResult)
def get_results(session_id: str):
    """Get full validation results for a completed session."""
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")
    if result.processing_status == "processing":
        raise HTTPException(202, "Still processing — poll /status/{session_id}")
    return result


@app.get("/export/excel/{session_id}")
def download_excel(session_id: str, request: Request,
                   authorization: Optional[str] = _FastAPIHeader(None)):
    """Download color-coded Excel validation report."""
    result = _load_session(session_id)
    if not result or result.processing_status != "completed":
        raise HTTPException(404, "Results not ready")

    output_path = os.path.join(_session_dir(session_id), "validation_report.xlsx")
    export_to_excel(result.rows, output_path)

    try:
        from services.db import log_activity
        log_activity(_get_actor(authorization), "export_excel", "session", session_id, {},
                     ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"expense_validation_{session_id[:8]}.xlsx",
    )


class SheetsRequest(BaseModel):
    sheet_title: Optional[str] = "Expense Validation Report"


@app.post("/export/sheets/{session_id}")
def export_sheets(session_id: str, request: Request, body: SheetsRequest = SheetsRequest(),
                  authorization: Optional[str] = _FastAPIHeader(None)):
    """Push results to Google Sheets. Returns shareable link."""
    _require_admin(authorization)
    result = _load_session(session_id)
    if not result or result.processing_status != "completed":
        raise HTTPException(404, "Results not ready")

    try:
        from services.db import log_activity
        log_activity(_get_actor(authorization), "export_sheets", "session", session_id, {},
                     ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    try:
        api_base = os.environ.get("API_BASE_URL", "http://localhost:8002")
        sheets_result = push_to_google_sheets(
            result.rows,
            sheet_title=body.sheet_title,
            session_id=session_id,
            api_base=api_base,
        )
        return sheets_result
    except FileNotFoundError as e:
        raise HTTPException(400, str(e))
    except RuntimeError as e:
        raise HTTPException(501, str(e))


class RowEditRequest(BaseModel):
    status: Optional[str] = None          # "Approved" | "Rejected" | "Flagged"
    note: Optional[str] = None            # manual remark to append
    bill_amount: Optional[float] = None   # manual bill amount override


@app.post("/session/{session_id}/lock")
def lock_session_endpoint(session_id: str, request: Request,
                          authorization: Optional[str] = _FastAPIHeader(None)):
    """Lock a session — admin only."""
    from services.auth import require_admin
    admin = require_admin(authorization)
    actor = admin.get("u") or admin.get("username", "admin")
    from services.db import lock_session, log_activity
    lock_session(session_id, actor)
    log_activity(actor, "session_lock", "session", session_id, {},
                 ip_address=request.client.host if request.client else "")
    return {"locked": True, "session_id": session_id}


@app.post("/session/{session_id}/unlock")
def unlock_session_endpoint(session_id: str, request: Request,
                            authorization: Optional[str] = _FastAPIHeader(None)):
    """Unlock a session — admin only."""
    from services.auth import require_admin
    admin = require_admin(authorization)
    actor = admin.get("u") or admin.get("username", "admin")
    from services.db import unlock_session, log_activity
    unlock_session(session_id)
    log_activity(actor, "session_unlock", "session", session_id, {},
                 ip_address=request.client.host if request.client else "")
    return {"locked": False, "session_id": session_id}


@app.patch("/session/{session_id}/row/{row_index}")
def edit_row(session_id: str, row_index: int, body: RowEditRequest,
             request: Request, authorization: Optional[str] = _FastAPIHeader(None)):
    """Manually override status, add a note, or correct bill amount for a row."""
    _require_admin(authorization)
    from models.schemas import ExpenseStatus
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")

    row = next((r for r in result.rows if r.row_index == row_index), None)
    if not row:
        raise HTTPException(404, f"Row {row_index} not found")

    if body.status:
        try:
            row.status = ExpenseStatus(body.status)
            row.remarks.append(f"[Manual Override] Status set to {body.status}")
        except ValueError:
            raise HTTPException(400, f"Invalid status: {body.status}")

    if body.note and body.note.strip():
        row.remarks.append(f"[Manual Note] {body.note.strip()}")

    if body.bill_amount is not None:
        row.bill_amount = body.bill_amount
        row.amount_diff = round(row.claimed_amount - body.bill_amount, 2)
        row.remarks.append(f"[Manual Override] Bill amount set to {body.bill_amount:.2f}")

    _save_session(session_id, result)

    try:
        from services.db import log_activity
        log_activity(_get_actor(authorization), "row_edit", "session", session_id,
                     {"row": row_index, "status": body.status, "note": body.note},
                     ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    return row


@app.get("/bill/{session_id}/{file_key:path}")
def view_bill(session_id: str, file_key: str):
    """Serve a bill file for in-browser preview."""
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")
    bills_dir = os.path.join(_session_dir(session_id), "bills")
    file_path = os.path.join(bills_dir, file_key)
    if not os.path.exists(file_path):
        raise HTTPException(404, "Bill file not found")
    ext = os.path.splitext(file_path)[1].lower()
    media_types = {
        ".pdf":  "application/pdf",
        ".jpg":  "image/jpeg",
        ".jpeg": "image/jpeg",
        ".png":  "image/png",
        ".webp": "image/webp",
    }
    return FileResponse(file_path, media_type=media_types.get(ext, "application/octet-stream"))


class ZohoPushRequest(BaseModel):
    # row_index → { selected: bool, account: str }
    row_overrides: Optional[dict] = None


@app.get("/zoho/accounts/debug")
def zoho_accounts_debug():
    """Debug: fetch ALL Zoho accounts across pages and group by account_type."""
    from services.zoho import _get_access_token, _headers, ZOHO_API_BASE, ORG_ID, CLIENT_ID, REFRESH_TOKEN
    if not CLIENT_ID or not REFRESH_TOKEN:
        raise HTTPException(501, "Zoho credentials not configured")
    import requests as _req
    token = _get_access_token()
    all_accts = []
    for page in range(1, 8):
        r = _req.get(f"{ZOHO_API_BASE}/chartofaccounts",
                     headers=_headers(token),
                     params={"organization_id": ORG_ID, "page": page, "per_page": 200})
        data = r.json()
        batch = data.get("chartofaccounts", [])
        all_accts.extend(batch)
        if not data.get("page_context", {}).get("has_more_page", False):
            break
    types_seen = {}
    for a in all_accts:
        t = a.get("account_type", "")
        types_seen.setdefault(t, []).append(a.get("account_name", ""))
    return {
        "total_accounts": len(all_accts),
        "account_types": {t: {"count": len(v), "samples": v[:6]} for t, v in types_seen.items()},
    }


@app.get("/zoho/ap-accounts")
def zoho_ap_accounts():
    """Debug: search Zoho for AP / payable accounts by name and all filter combinations."""
    from services.zoho import _get_access_token, _headers, ZOHO_API_BASE, ORG_ID, CLIENT_ID, REFRESH_TOKEN
    if not CLIENT_ID or not REFRESH_TOKEN:
        raise HTTPException(501, "Zoho credentials not configured")
    import requests as _req
    token = _get_access_token()
    results = {}

    search_terms = ["Employe Reimbursement", "Employee Reimbursement", "Reimbursement",
                    "Accounts Payable", "Payable", "EM_PAYABLE"]
    for term in search_terms:
        r = _req.get(f"{ZOHO_API_BASE}/chartofaccounts",
                     headers=_headers(token),
                     params={"organization_id": ORG_ID, "search_text": term, "per_page": 50})
        found = [{"account_id": a["account_id"], "account_name": a.get("account_name"),
                  "account_type": a.get("account_type"), "account_code": a.get("account_code")}
                 for a in r.json().get("chartofaccounts", [])]
        if found:
            results[term] = found

    # Also try filter_by variants
    for fb in ["AccountType.AccountsPayable", "AccountType.AccountPayable", "AccountType.Payable"]:
        r = _req.get(f"{ZOHO_API_BASE}/chartofaccounts",
                     headers=_headers(token),
                     params={"organization_id": ORG_ID, "filter_by": fb, "per_page": 200})
        found = [{"account_id": a["account_id"], "account_name": a.get("account_name"),
                  "account_type": a.get("account_type"), "account_code": a.get("account_code")}
                 for a in r.json().get("chartofaccounts", [])]
        if found:
            results[f"filter:{fb}"] = found

    return results


@app.get("/zoho/accounts/{session_id}")
def zoho_accounts(session_id: str):
    """Return expense account names + auto-detected account per approved row."""
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")
    from services.zoho import get_expense_accounts, _get_access_token, _map_to_account, CLIENT_ID, REFRESH_TOKEN
    if not CLIENT_ID or not REFRESH_TOKEN:
        raise HTTPException(501, "Zoho credentials not configured")
    try:
        token    = _get_access_token()
        accounts = get_expense_accounts(token)
        names    = [a["account_name"] for a in accounts]
        # Pre-detect account for each approved row
        from models.schemas import ExpenseStatus
        detected = {}
        for row in result.rows:
            if row.status == ExpenseStatus.APPROVED:
                detected[row.row_index] = _map_to_account(
                    row.expense_nature or row.expense_category or "",
                    row.description or "",
                    accounts,
                )
        return {"accounts": names, "detected": detected}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/export/zoho/{session_id}")
def export_zoho(session_id: str, request: Request,
                body: ZohoPushRequest = ZohoPushRequest(),
                authorization: Optional[str] = _FastAPIHeader(None)):
    """Push selected Approved claims to Zoho Books as drafts with bill attachments, then AI-verify."""
    _require_admin(authorization)
    result = _load_session(session_id)
    if not result or result.processing_status != "completed":
        raise HTTPException(404, "Results not ready")
    bills_dir = os.path.join(_session_dir(session_id), "bills")

    try:
        from services.db import log_activity
        log_activity(_get_actor(authorization), "export_zoho", "session", session_id, {},
                     ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    # Convert row_overrides keys from string (JSON) back to int
    row_overrides = None
    if body.row_overrides:
        row_overrides = {int(k): v for k, v in body.row_overrides.items()}

    try:
        from datetime import datetime as _dt
        zoho_result = push_approved_to_zoho(result.rows, bills_dir, row_overrides)
        rows_by_index = {r.row_index: r for r in result.rows}
        zoho_result["verification"] = verify_zoho_entries(
            zoho_result["pushed"],
            rows_by_index,
            all_pushed=zoho_result.get("all_pushed"),
        )
        zoho_result["pushed_at"] = _dt.now().isoformat()
        # Persist to disk — use all_pushed (full history) so next push detects duplicates
        push_path = os.path.join(_session_dir(session_id), "zoho_push.json")
        disk_payload = {
            "pushed":     zoho_result.get("all_pushed", zoho_result["pushed"]),
            "errors":     zoho_result["errors"],
            "verification": zoho_result.get("verification", []),
            "pushed_at":  zoho_result["pushed_at"],
        }
        with open(push_path, "w", encoding="utf-8") as _f:
            json.dump(disk_payload, _f, ensure_ascii=False, indent=2)
        return zoho_result
    except RuntimeError as e:
        raise HTTPException(501, str(e))
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/zoho/fix-vendor-numbers/{session_id}")
def fix_zoho_vendor_numbers(session_id: str, authorization: Optional[str] = _FastAPIHeader(None)):
    """
    For vendors already in Zoho with auto-generated VND-XXXXX numbers,
    update their contact_number to the employee code from this session's expense file.
    """
    _require_admin(authorization)
    result = _load_session(session_id)
    if not result or result.processing_status != "completed":
        raise HTTPException(404, "Results not ready")
    try:
        from services.zoho import fix_vendor_numbers
        fix_result = fix_vendor_numbers(result.rows)
        return fix_result
    except RuntimeError as e:
        raise HTTPException(501, str(e))
    except Exception as e:
        raise HTTPException(500, str(e))


# ---------------------------------------------------------------------------
# Session History endpoint
# ---------------------------------------------------------------------------

@app.get("/sessions/history")
def sessions_history(authorization: Optional[str] = _FastAPIHeader(None)):
    """Return all past sessions with metadata for the history page."""
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "No token")
    if not verify_token(authorization[7:].strip()):
        raise HTTPException(401, "Invalid or expired token")

    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        entries = []

    sessions = []
    for session_id in entries:
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            mtime = os.path.getmtime(result_path)
            created_at = datetime.fromtimestamp(mtime).isoformat()
            with open(result_path, "r", encoding="utf-8", errors="replace") as f:
                data = json.load(f)
            rows_list = data.get("rows", [])
            total_amount = sum(r.get("claimed_amount", 0) for r in rows_list)
            # Always recompute from actual rows (never trust stale stored counts)
            approved_count  = sum(1 for r in rows_list if (r.get("status") or "").strip().lower() == "approved")
            rejected_count  = sum(1 for r in rows_list if (r.get("status") or "").strip().lower() == "rejected")
            flagged_count   = len(rows_list) - approved_count - rejected_count
            approved_amount = sum(
                r.get("bill_amount") or r.get("claimed_amount", 0)
                for r in rows_list
                if (r.get("status") or "").strip().lower() == "approved"
            )
            # Detect source: keka sessions have postman_status.json OR bills_downloaded > 0
            is_keka_batch = os.path.isfile(
                os.path.join(UPLOAD_BASE, session_id, "postman_status.json")
            )
            source = "keka" if (is_keka_batch or data.get("bills_downloaded", 0) > 0) else "upload"
            # Count rows where keka_actioned is approve/approved/mark_paid
            # OR all approved rows if this is a Keka batch (they came pre-approved from Keka)
            keka_approved_count = sum(
                1 for r in rows_list
                if (
                    (r.get("keka_actioned") or "").strip().lower() in ("approve", "approved", "mark_paid")
                    or bool(r.get("keka_claim_id"))
                    or (is_keka_batch and (r.get("status") or "").strip().lower() == "approved")
                )
            )
            # Count rows where keka_actioned is not None/empty
            keka_actioned_count = sum(
                1 for r in rows_list
                if r.get("keka_actioned") not in (None, "", "none")
            )
            # Count zoho pushed rows — read zoho_push.json if it exists
            zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            zoho_pushed_count = 0
            if os.path.isfile(zoho_push_path):
                try:
                    with open(zoho_push_path, "r", encoding="utf-8") as _zf:
                        _zpd = json.load(_zf)
                    zoho_pushed_count = len(_zpd.get("pushed", []))
                except Exception:
                    zoho_pushed_count = 0
            # ── Lock info from DB ──────────────────────────────────────────
            from services.db import get_session_lock_info
            lock_info = get_session_lock_info(session_id) or {"locked": False, "locked_by": "", "locked_at": None}

            sessions.append({
                "session_id": session_id,
                "created_at": created_at,
                "processing_status": data.get("processing_status", "unknown"),
                "total_claims": len(rows_list) or data.get("total_claims", 0),
                "approved": approved_count,
                "rejected": rejected_count,
                "flagged":  flagged_count,
                "total_amount": round(total_amount, 2),
                "approved_amount": round(approved_amount, 2),
                "source": source,
                "error": data.get("error") or "",
                "keka_approved_count": keka_approved_count,
                "keka_actioned_count": keka_actioned_count,
                "zoho_pushed_count": zoho_pushed_count,
                "locked": lock_info["locked"],
                "locked_by": lock_info["locked_by"],
                "locked_at": lock_info["locked_at"],
            })
        except Exception:
            continue

    # Sort newest first
    sessions.sort(key=lambda s: s["created_at"], reverse=True)
    return {"sessions": sessions}


# ---------------------------------------------------------------------------
# Search claims endpoint
# ---------------------------------------------------------------------------

@app.get("/search/claims")
def search_claims(
    q: str = Query(""),
    status: str = Query(""),       # "Approved" | "Rejected" | "Flagged" | ""
    category: str = Query(""),     # keyword match on expense_category/expense_nature
    min_amount: float = Query(0.0),
    max_amount: float = Query(0.0),  # 0 = no upper limit
    from_date: str = Query(""),    # YYYY-MM-DD
    to_date: str = Query(""),      # YYYY-MM-DD
    keka_only: bool = Query(False),  # only rows with keka_actioned set
    zoho_only: bool = Query(False),  # only rows from sessions with zoho_push.json
    limit: int = Query(200),
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """Search claims across all sessions."""
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "No token")
    if not verify_token(authorization[7:].strip()):
        raise HTTPException(401, "Invalid or expired token")

    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        entries = []

    results = []
    for session_id in entries:
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            mtime = os.path.getmtime(result_path)
            created_at = datetime.fromtimestamp(mtime).isoformat()
            with open(result_path, "r", encoding="utf-8", errors="replace") as f:
                data = json.load(f)

            # Zoho push info for this session
            zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            zoho_pushed_indices = set()
            if os.path.isfile(zoho_push_path):
                try:
                    with open(zoho_push_path, "r", encoding="utf-8") as _zf:
                        _zpd = json.load(_zf)
                    for p in _zpd.get("pushed", []):
                        if isinstance(p, dict) and "row_index" in p:
                            zoho_pushed_indices.add(p["row_index"])
                except Exception:
                    pass

            source = "keka" if data.get("bills_downloaded", 0) > 0 else "upload"
            if zoho_only and not zoho_pushed_indices:
                continue

            for row in data.get("rows", []):
                # Text search: employee_name, description, expense_category, expense_nature
                if q:
                    haystack = " ".join([
                        row.get("employee_name", ""),
                        row.get("description", ""),
                        row.get("expense_category", ""),
                        row.get("expense_nature", ""),
                        row.get("employee_id", ""),
                    ]).lower()
                    if q.lower() not in haystack:
                        continue

                # Status filter
                if status and row.get("status") != status:
                    continue

                # Category filter (keyword)
                if category:
                    cat_text = (row.get("expense_category", "") + " " + row.get("expense_nature", "")).lower()
                    if category.lower() not in cat_text:
                        continue

                # Amount filters
                amt = row.get("claimed_amount", 0)
                if min_amount > 0 and amt < min_amount:
                    continue
                if max_amount > 0 and amt > max_amount:
                    continue

                # Keka filter
                if keka_only and not row.get("keka_actioned"):
                    continue

                results.append({
                    "session_id": session_id,
                    "created_at": created_at,
                    "source": source,
                    "row_index": row.get("row_index"),
                    "employee_name": row.get("employee_name", ""),
                    "employee_id": row.get("employee_id", ""),
                    "expense_date": row.get("expense_date", ""),
                    "expense_category": row.get("expense_category", ""),
                    "expense_nature": row.get("expense_nature", ""),
                    "description": row.get("description", ""),
                    "claimed_amount": round(amt, 2),
                    "bill_amount": row.get("bill_amount"),
                    "status": row.get("status", ""),
                    "keka_actioned": row.get("keka_actioned"),
                    "zoho_pushed": row.get("row_index") in zoho_pushed_indices,
                    "remarks": row.get("remarks", []),
                })

                if len(results) >= limit:
                    break
        except Exception:
            continue

    # Sort by created_at desc
    results.sort(key=lambda r: r["created_at"], reverse=True)
    return {"results": results, "total": len(results)}


# ---------------------------------------------------------------------------
# Report export endpoint (CSV download)
# ---------------------------------------------------------------------------

@app.get("/report/export")
def report_export(
    month: str = Query(""),    # "2025-05" format — filter to this month only
    status: str = Query(""),   # "Approved" | "Rejected" | "Flagged" | ""
    source: str = Query(""),   # "keka" | "upload" | ""
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """Download a comprehensive CSV report of all claims across sessions."""
    import csv, io as _io
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "No token")
    if not verify_token(authorization[7:].strip()):
        raise HTTPException(401, "Invalid or expired token")

    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        entries = []

    output = _io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Session ID", "Session Date", "Source",
        "Employee Name", "Employee Code", "Expense Date",
        "Category", "Description",
        "Claimed Amount (₹)", "Approved Amount (₹)",
        "Validation Status", "Keka Action", "Zoho Pushed",
        "Remarks"
    ])

    rows_written = 0
    for session_id in sorted(entries, reverse=True):
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            mtime = os.path.getmtime(result_path)
            created_at = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M")
            created_month = datetime.fromtimestamp(mtime).strftime("%Y-%m")
            with open(result_path, "r", encoding="utf-8", errors="replace") as f:
                data = json.load(f)

            sess_source = "keka" if data.get("bills_downloaded", 0) > 0 else "upload"

            # Month filter
            if month and created_month != month:
                continue
            # Source filter
            if source and sess_source != source:
                continue

            # Read zoho push data
            zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            zoho_pushed_indices = set()
            if os.path.isfile(zoho_push_path):
                try:
                    with open(zoho_push_path, "r", encoding="utf-8") as _zf:
                        _zpd = json.load(_zf)
                    for p in _zpd.get("pushed", []):
                        if isinstance(p, dict) and "row_index" in p:
                            zoho_pushed_indices.add(p["row_index"])
                except Exception:
                    pass

            for row in data.get("rows", []):
                row_status = row.get("status", "")
                if status and row_status != status:
                    continue

                amt = row.get("claimed_amount", 0)
                bill_amt = row.get("bill_amount") or amt if row_status == "Approved" else 0
                zoho_pushed = row.get("row_index") in zoho_pushed_indices
                keka_action = row.get("keka_actioned") or "—"
                remarks_str = " | ".join(row.get("remarks", []))

                writer.writerow([
                    session_id[:8], created_at, sess_source,
                    row.get("employee_name", ""), row.get("employee_id", ""),
                    row.get("expense_date", ""),
                    row.get("expense_category", ""), row.get("description", ""),
                    round(amt, 2), round(bill_amt, 2),
                    row_status, keka_action, "Yes" if zoho_pushed else "No",
                    remarks_str,
                ])
                rows_written += 1
        except Exception:
            continue

    csv_bytes = output.getvalue().encode("utf-8-sig")  # utf-8-sig so Excel opens correctly
    filename = f"Wiom_Claims_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.csv"
    return StreamingResponse(
        iter([csv_bytes]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ---------------------------------------------------------------------------
# Full-portfolio export  (Excel multi-sheet  +  PDF/HTML)
# ---------------------------------------------------------------------------

def _collect_all_rows():
    """Scan every completed session, return list of enriched row dicts."""
    all_rows = []
    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        return all_rows

    # Sort sessions by file mtime (oldest first) so batch numbers are chronological
    def _mtime(sid):
        p = os.path.join(UPLOAD_BASE, sid, "result.json")
        try: return os.path.getmtime(p)
        except: return 0

    sorted_entries = sorted(entries, key=_mtime)
    batch_counter = 0

    for session_id in sorted_entries:
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            mtime = os.path.getmtime(result_path)
            session_date = datetime.fromtimestamp(mtime).strftime("%d %b %Y")

            data = None
            for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
                try:
                    with open(result_path, encoding=enc) as f:
                        data = json.load(f)
                    break
                except (UnicodeDecodeError, json.JSONDecodeError):
                    continue
            if not data or data.get("processing_status") != "completed":
                continue

            batch_counter += 1
            batch_no = batch_counter

            # Zoho push index
            zoho_indices: set = set()
            zoho_bill_ids: dict = {}
            zoho_statuses: dict = {}
            zp = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            if os.path.isfile(zp):
                try:
                    with open(zp, encoding="utf-8") as f:
                        zpd = json.load(f)
                    for p in zpd.get("pushed", []):
                        ri = int(p.get("row_index", -1))
                        zoho_indices.add(ri)
                        zoho_bill_ids[ri]   = p.get("bill_id", "")
                        zoho_statuses[ri]   = p.get("bill_status", "")
                except Exception:
                    pass

            for row in data.get("rows", []):
                ri = row.get("row_index", -1)
                is_zoho = ri in zoho_indices or bool(row.get("zoho_bill_id"))
                zoho_st = (row.get("zoho_bill_status") or zoho_statuses.get(ri, "")).strip().lower()
                keka_act = (row.get("keka_actioned") or "").strip().lower()

                all_rows.append({
                    "batch_no":      batch_no,
                    "session_id":    session_id[:8],
                    "session_date":  session_date,
                    "employee_name": row.get("employee_name", ""),
                    "employee_id":   row.get("employee_id",   ""),
                    "department":    row.get("department", "") or "",
                    "category":      row.get("expense_category", "") or row.get("expense_nature", "") or "Other",
                    "expense_date":  row.get("expense_date",  ""),
                    "description":   row.get("description",  "") or "",
                    "amount":        float(row.get("claimed_amount") or 0),
                    "status":        row.get("status", ""),
                    "zoho_booked":   "Yes" if is_zoho else "No",
                    "zoho_paid":     "Yes" if zoho_st == "paid" else ("No" if is_zoho else "—"),
                    "zoho_bill_id":  row.get("zoho_bill_id") or zoho_bill_ids.get(ri, ""),
                    "keka_status":   keka_act.replace("_", " ").title() if keka_act else "—",
                    "claim_number":  row.get("claim_number") or "",
                    "remarks":       " | ".join(row.get("remarks", [])) if isinstance(row.get("remarks"), list) else str(row.get("remarks") or ""),
                })
        except Exception as e:
            log.warning("export: session %s error: %s", session_id, e)

    return all_rows


def _build_excel(rows: list) -> bytes:
    """Build a beautifully formatted multi-sheet Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    import io as _io

    PINK    = "E5007D"
    PINK_LT = "FFF0F8"
    GREEN   = "059669"
    GREEN_LT= "F0FDF4"
    RED     = "DC2626"
    RED_LT  = "FEF2F2"
    AMBER   = "D97706"
    AMBER_LT= "FFFBEB"
    PURPLE  = "7C3AED"
    PURPLE_LT="F5F3FF"
    GRAY    = "64748B"
    GRAY_LT = "F8FAFC"
    WHITE   = "FFFFFF"
    DARK    = "0F1629"
    BORDER_CLR = "E2E8F0"

    def hdr_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
    def thin_border():
        s = Side(style="thin", color=BORDER_CLR)
        return Border(left=s, right=s, top=s, bottom=s)
    def hdr_font(color=WHITE, sz=10): return Font(bold=True, color=color, size=sz, name="Calibri")
    def body_font(bold=False, color=DARK, sz=10): return Font(bold=bold, color=color, size=sz, name="Calibri")

    def style_header_row(ws, row_num, fill_hex, cols):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = hdr_fill(fill_hex)
            cell.font = hdr_font()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border()

    def autofit(ws, min_w=8, max_w=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(w + 2, min_w), max_w)

    wb = Workbook()
    wb.remove(wb.active)   # remove default sheet

    # ── STATUS color helper ────────────────────────────────────────
    def status_fill(st):
        s = (st or "").lower()
        if s == "approved": return hdr_fill(GREEN_LT), body_font(color=GREEN)
        if s == "rejected": return hdr_fill(RED_LT),   body_font(color=RED)
        return hdr_fill(AMBER_LT), body_font(color=AMBER)

    # ── totals helper ──────────────────────────────────────────────
    def write_totals_row(ws, row_num, cols, label_col, amount_col, count, amount):
        for c in range(1, cols + 1):
            cell = ws.cell(row=row_num, column=c)
            cell.fill = hdr_fill("F1F5F9")
            cell.font = body_font(bold=True, color=DARK)
            cell.border = thin_border()
        ws.cell(row=row_num, column=label_col).value = f"TOTAL  ({count} claims)"
        ws.cell(row=row_num, column=amount_col).value = amount
        ws.cell(row=row_num, column=amount_col).number_format = '₹#,##0.00'

    # ══════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════════════════════════
    ws1 = wb.create_sheet("Summary")
    ws1.sheet_view.showGridLines = False

    total       = len(rows)
    approved    = [r for r in rows if (r["status"] or "").lower() == "approved"]
    rejected    = [r for r in rows if (r["status"] or "").lower() == "rejected"]
    flagged     = [r for r in rows if (r["status"] or "").lower() not in ("approved","rejected")]
    zoho_booked = [r for r in approved if r["zoho_booked"] == "Yes"]
    zoho_paid   = [r for r in approved if r["zoho_paid"]   == "Yes"]
    keka_appr   = [r for r in approved if r["keka_status"].lower() in ("approve","approved","mark paid")]
    not_zoho    = [r for r in approved if r["zoho_booked"] == "No"]

    s_amt  = sum(r["amount"] for r in rows)
    ap_amt = sum(r["amount"] for r in approved)
    re_amt = sum(r["amount"] for r in rejected)
    zb_amt = sum(r["amount"] for r in zoho_booked)
    zp_amt = sum(r["amount"] for r in zoho_paid)
    nz_amt = sum(r["amount"] for r in not_zoho)

    # Title
    ws1.merge_cells("A1:F1")
    t = ws1["A1"]; t.value = "WIOM FINANCE — EXPENSE REIMBURSEMENT REPORT (ALL BATCHES)"
    t.font = Font(bold=True, size=14, color=WHITE, name="Calibri")
    t.fill = hdr_fill(PINK); t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:F2")
    t2 = ws1["A2"]
    t2.value = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}   |   Total Batches Scanned"
    t2.font = body_font(color=GRAY, sz=9); t2.alignment = Alignment(horizontal="center")
    ws1.row_dimensions[2].height = 18

    headers_s = ["Metric", "Count", "Amount (₹)", "% of Total Claims", "% of Total Amount", "Note"]
    for ci, h in enumerate(headers_s, 1):
        c = ws1.cell(row=4, column=ci, value=h)
        c.fill = hdr_fill(DARK); c.font = hdr_font(); c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[4].height = 22

    summary_data = [
        ("Total Claims",         total,         s_amt,  100,                       100,                       "All batches"),
        ("Approved",             len(approved),  ap_amt, round(len(approved)/total*100,1) if total else 0,    round(ap_amt/s_amt*100,1) if s_amt else 0, "AI-validated approved"),
        ("Rejected",             len(rejected),  re_amt, round(len(rejected)/total*100,1) if total else 0,   round(re_amt/s_amt*100,1) if s_amt else 0, "AI-validated rejected"),
        ("Flagged / Review",     len(flagged),   s_amt-ap_amt-re_amt, round(len(flagged)/total*100,1) if total else 0, "—", "Pending review"),
        ("Pushed to Zoho",       len(zoho_booked), zb_amt, round(len(zoho_booked)/total*100,1) if total else 0, round(zb_amt/s_amt*100,1) if s_amt else 0, "Booked in Zoho Books"),
        ("Paid in Zoho",         len(zoho_paid),  zp_amt, round(len(zoho_paid)/total*100,1) if total else 0,  round(zp_amt/s_amt*100,1) if s_amt else 0, "Marked paid in Zoho"),
        ("Keka Approved",        len(keka_appr),  "—",   round(len(keka_appr)/total*100,1) if total else 0,  "—", "Actioned in Keka"),
        ("Pending (Not in Zoho)",len(not_zoho),  nz_amt, round(len(not_zoho)/total*100,1) if total else 0,   round(nz_amt/s_amt*100,1) if s_amt else 0, "Approved, awaiting Zoho"),
    ]
    fills_s = [GRAY_LT, GREEN_LT, RED_LT, AMBER_LT, PURPLE_LT, PINK_LT, AMBER_LT, RED_LT]
    for ri, (row_data, fill_hex) in enumerate(zip(summary_data, fills_s), 5):
        for ci, val in enumerate(row_data, 1):
            cell = ws1.cell(row=ri, column=ci, value=val)
            cell.fill = hdr_fill(fill_hex)
            cell.border = thin_border()
            cell.font = body_font(bold=(ci == 1))
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == 3 and isinstance(val, (int, float)):
                cell.number_format = '₹#,##0.00'
        ws1.row_dimensions[ri + 4].height = 18

    for ci, w in enumerate([28, 12, 18, 18, 18, 28], 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w
    ws1.freeze_panes = "A5"

    # ══════════════════════════════════════════════════════════════
    # SHEET 2 — All Claims (detail)
    # ══════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Claims")
    ws2.sheet_view.showGridLines = False

    # Title banner
    ws2.merge_cells("A1:O1")
    t = ws2["A1"]; t.value = "ALL CLAIMS — DETAILED VIEW"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = hdr_fill(PINK); t.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 28

    cols2 = ["#", "Batch No.", "Batch ID", "Batch Date", "Employee Name", "Employee ID",
             "Department", "Category", "Expense Date", "Description",
             "Amount (₹)", "Status", "Zoho Booked", "Zoho Paid",
             "Keka Status", "Claim No."]
    for ci, h in enumerate(cols2, 1):
        c = ws2.cell(row=2, column=ci, value=h)
        c.fill = hdr_fill(DARK); c.font = hdr_font(); c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws2.row_dimensions[2].height = 24
    ws2.freeze_panes = "A3"

    for i, r in enumerate(rows, 1):
        rn = i + 2
        sf, ff = status_fill(r["status"])
        row_bg = hdr_fill("FFFFFF") if i % 2 == 0 else hdr_fill("F8FAFC")
        vals = [i, f"Batch #{r['batch_no']}", r["session_id"], r["session_date"], r["employee_name"],
                r["employee_id"], r["department"], r["category"],
                r["expense_date"], r["description"], r["amount"],
                r["status"], r["zoho_booked"], r["zoho_paid"],
                r["keka_status"], r["claim_number"]]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=rn, column=ci, value=v)
            cell.border = thin_border()
            cell.alignment = Alignment(vertical="center", wrap_text=(ci == 10))
            if ci == 11:
                cell.number_format = '₹#,##0.00'
                cell.font = body_font(bold=True)
            elif ci == 12:
                cell.fill = sf; cell.font = ff; cell.alignment = Alignment(horizontal="center", vertical="center")
            elif ci in (13, 14):
                yesno = str(v)
                cell.fill = hdr_fill(GREEN_LT if yesno == "Yes" else (RED_LT if yesno == "No" else "F8FAFC"))
                cell.font = body_font(color=(GREEN if yesno == "Yes" else (RED if yesno == "No" else GRAY)))
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.fill = row_bg if ci != 11 else sf
                cell.font = body_font()
        ws2.row_dimensions[rn].height = 16

    # Totals row
    tot_row = len(rows) + 3
    write_totals_row(ws2, tot_row, len(cols2), 5, 11, len(rows), sum(r["amount"] for r in rows))

    col_widths2 = [5,10,10,12,24,12,16,18,13,28,14,11,12,10,14,12]
    for ci, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ══════════════════════════════════════════════════════════════
    # SHEET 3 — By Employee
    # ══════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("By Employee")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:I1")
    t = ws3["A1"]; t.value = "EMPLOYEE-WISE SUMMARY"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = hdr_fill(PURPLE); t.alignment = Alignment(horizontal="center", vertical="center")
    ws3.row_dimensions[1].height = 28

    cols3 = ["Employee Name", "Employee ID", "Department", "Total Claims",
             "Approved", "Rejected", "Flagged", "Total Amount (₹)",
             "Approved Amount (₹)"]
    for ci, h in enumerate(cols3, 1):
        c = ws3.cell(row=2, column=ci, value=h)
        c.fill = hdr_fill(DARK); c.font = hdr_font(); c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws3.row_dimensions[2].height = 24
    ws3.freeze_panes = "A3"

    emp_map: dict = {}
    for r in rows:
        k = (r["employee_name"], r["employee_id"], r["department"])
        if k not in emp_map:
            emp_map[k] = {"total":0,"approved":0,"rejected":0,"flagged":0,"amount":0.0,"ap_amount":0.0}
        d = emp_map[k]
        d["total"]  += 1; d["amount"] += r["amount"]
        st = (r["status"] or "").lower()
        if st == "approved":   d["approved"] += 1; d["ap_amount"] += r["amount"]
        elif st == "rejected": d["rejected"] += 1
        else:                  d["flagged"]  += 1

    sorted_emps = sorted(emp_map.items(), key=lambda x: x[1]["amount"], reverse=True)
    for i, ((name, eid, dept), d) in enumerate(sorted_emps, 1):
        rn = i + 2
        row_bg = hdr_fill("FFFFFF") if i % 2 == 0 else hdr_fill("F8FAFC")
        vals = [name, eid, dept, d["total"], d["approved"], d["rejected"], d["flagged"], d["amount"], d["ap_amount"]]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=rn, column=ci, value=v)
            cell.fill = row_bg; cell.border = thin_border()
            cell.font = body_font(bold=(ci == 1))
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci > 3 else "left")
            if ci in (8, 9): cell.number_format = '₹#,##0.00'
        ws3.row_dimensions[rn].height = 16

    tot_row3 = len(sorted_emps) + 3
    write_totals_row(ws3, tot_row3, len(cols3), 1, 8, len(rows), sum(r["amount"] for r in rows))
    ws3.cell(row=tot_row3, column=9).value = ap_amt
    ws3.cell(row=tot_row3, column=9).number_format = '₹#,##0.00'

    for ci, w in enumerate([24,12,18,12,12,12,12,18,18], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    # ══════════════════════════════════════════════════════════════
    # SHEET 4 — By Department
    # ══════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("By Department")
    ws4.sheet_view.showGridLines = False

    ws4.merge_cells("A1:H1")
    t = ws4["A1"]; t.value = "DEPARTMENT-WISE SUMMARY"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = hdr_fill("0EA5E9"); t.alignment = Alignment(horizontal="center", vertical="center")
    ws4.row_dimensions[1].height = 28

    cols4 = ["Department", "Total Claims", "Approved", "Rejected", "Flagged",
             "Total Amount (₹)", "Approved Amount (₹)", "Approval Rate (%)"]
    for ci, h in enumerate(cols4, 1):
        c = ws4.cell(row=2, column=ci, value=h)
        c.fill = hdr_fill(DARK); c.font = hdr_font(); c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws4.row_dimensions[2].height = 24
    ws4.freeze_panes = "A3"

    dept_map: dict = {}
    for r in rows:
        dept = r["department"] or "—"
        if dept not in dept_map:
            dept_map[dept] = {"total":0,"approved":0,"rejected":0,"flagged":0,"amount":0.0,"ap_amount":0.0}
        d = dept_map[dept]; d["total"] += 1; d["amount"] += r["amount"]
        st = (r["status"] or "").lower()
        if st == "approved":   d["approved"] += 1; d["ap_amount"] += r["amount"]
        elif st == "rejected": d["rejected"] += 1
        else:                  d["flagged"]  += 1

    for i, (dept, d) in enumerate(sorted(dept_map.items(), key=lambda x: x[1]["amount"], reverse=True), 1):
        rn = i + 2
        row_bg = hdr_fill("FFFFFF") if i % 2 == 0 else hdr_fill("F8FAFC")
        rate = round(d["approved"]/d["total"]*100, 1) if d["total"] else 0
        vals = [dept, d["total"], d["approved"], d["rejected"], d["flagged"], d["amount"], d["ap_amount"], rate]
        for ci, v in enumerate(vals, 1):
            cell = ws4.cell(row=rn, column=ci, value=v)
            cell.fill = row_bg; cell.border = thin_border()
            cell.font = body_font(bold=(ci == 1))
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci > 1 else "left")
            if ci in (6, 7): cell.number_format = '₹#,##0.00'
            if ci == 8:      cell.number_format = '0.0"%"'
        ws4.row_dimensions[rn].height = 16

    for ci, w in enumerate([22,14,12,12,12,18,18,16], 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w

    # ══════════════════════════════════════════════════════════════
    # SHEET 5 — By Category
    # ══════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("By Category")
    ws5.sheet_view.showGridLines = False

    ws5.merge_cells("A1:H1")
    t = ws5["A1"]; t.value = "EXPENSE CATEGORY-WISE SUMMARY"
    t.font = Font(bold=True, size=12, color=WHITE, name="Calibri")
    t.fill = hdr_fill(PINK); t.alignment = Alignment(horizontal="center", vertical="center")
    ws5.row_dimensions[1].height = 28

    cols5 = ["Category", "Total Claims", "Approved", "Rejected", "Flagged",
             "Total Amount (₹)", "Approved Amount (₹)", "Approval Rate (%)"]
    for ci, h in enumerate(cols5, 1):
        c = ws5.cell(row=2, column=ci, value=h)
        c.fill = hdr_fill(DARK); c.font = hdr_font(); c.border = thin_border()
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws5.row_dimensions[2].height = 24
    ws5.freeze_panes = "A3"

    cat_map: dict = {}
    for r in rows:
        cat = r["category"] or "Other"
        if cat not in cat_map:
            cat_map[cat] = {"total":0,"approved":0,"rejected":0,"flagged":0,"amount":0.0,"ap_amount":0.0}
        d = cat_map[cat]; d["total"] += 1; d["amount"] += r["amount"]
        st = (r["status"] or "").lower()
        if st == "approved":   d["approved"] += 1; d["ap_amount"] += r["amount"]
        elif st == "rejected": d["rejected"] += 1
        else:                  d["flagged"]  += 1

    for i, (cat, d) in enumerate(sorted(cat_map.items(), key=lambda x: x[1]["amount"], reverse=True), 1):
        rn = i + 2
        row_bg = hdr_fill("FFFFFF") if i % 2 == 0 else hdr_fill("F8FAFC")
        rate = round(d["approved"]/d["total"]*100, 1) if d["total"] else 0
        vals = [cat, d["total"], d["approved"], d["rejected"], d["flagged"], d["amount"], d["ap_amount"], rate]
        for ci, v in enumerate(vals, 1):
            cell = ws5.cell(row=rn, column=ci, value=v)
            cell.fill = row_bg; cell.border = thin_border()
            cell.font = body_font(bold=(ci == 1))
            cell.alignment = Alignment(vertical="center", horizontal="center" if ci > 1 else "left")
            if ci in (6, 7): cell.number_format = '₹#,##0.00'
            if ci == 8:      cell.number_format = '0.0"%"'
        ws5.row_dimensions[rn].height = 16

    for ci, w in enumerate([28,14,12,12,12,18,18,16], 1):
        ws5.column_dimensions[get_column_letter(ci)].width = w

    buf = _io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_pdf_html(rows: list) -> str:
    """Return a print-ready HTML string that browsers can save as PDF."""
    total   = len(rows)
    approved= [r for r in rows if (r["status"] or "").lower() == "approved"]
    rejected= [r for r in rows if (r["status"] or "").lower() == "rejected"]
    flagged = [r for r in rows if (r["status"] or "").lower() not in ("approved","rejected")]
    zoho_bk = [r for r in approved if r["zoho_booked"] == "Yes"]
    zoho_pd = [r for r in approved if r["zoho_paid"]   == "Yes"]
    keka_ap = [r for r in approved if r["keka_status"].lower() in ("approve","approved","mark paid")]
    not_zh  = [r for r in approved if r["zoho_booked"] == "No"]

    s_amt  = sum(r["amount"] for r in rows)
    ap_amt = sum(r["amount"] for r in approved)
    zb_amt = sum(r["amount"] for r in zoho_bk)
    zp_amt = sum(r["amount"] for r in zoho_pd)
    nz_amt = sum(r["amount"] for r in not_zh)

    def fmt(n):
        if n >= 1_000_000: return f"₹{n/1_000_000:.2f}L"
        if n >= 1_000:     return f"₹{n/1_000:.1f}K"
        return f"₹{n:,.0f}"

    def pct(a, b): return f"{round(a/b*100,1)}%" if b else "0%"

    # ── Category breakdown
    cat_map: dict = {}
    for r in rows:
        k = r["category"] or "Other"
        cat_map.setdefault(k, {"t":0,"a":0,"amt":0.0})
        cat_map[k]["t"] += 1; cat_map[k]["amt"] += r["amount"]
        if (r["status"] or "").lower() == "approved": cat_map[k]["a"] += 1
    cat_rows = "".join(
        f"<tr><td>{cat}</td><td>{d['t']}</td><td>{d['a']}</td>"
        f"<td>{fmt(d['amt'])}</td><td>{pct(d['a'],d['t'])}</td></tr>"
        for cat, d in sorted(cat_map.items(), key=lambda x: x[1]["amt"], reverse=True)
    )

    # ── Dept breakdown
    dept_map: dict = {}
    for r in rows:
        k = r["department"] or "—"
        dept_map.setdefault(k, {"t":0,"a":0,"amt":0.0})
        dept_map[k]["t"] += 1; dept_map[k]["amt"] += r["amount"]
        if (r["status"] or "").lower() == "approved": dept_map[k]["a"] += 1
    dept_rows = "".join(
        f"<tr><td>{dept}</td><td>{d['t']}</td><td>{d['a']}</td>"
        f"<td>{fmt(d['amt'])}</td><td>{pct(d['a'],d['t'])}</td></tr>"
        for dept, d in sorted(dept_map.items(), key=lambda x: x[1]["amt"], reverse=True)
    )

    # ── Employee breakdown
    emp_map: dict = {}
    for r in rows:
        k = (r["employee_name"], r["employee_id"])
        emp_map.setdefault(k, {"t":0,"a":0,"amt":0.0,"ap_amt":0.0})
        emp_map[k]["t"] += 1; emp_map[k]["amt"] += r["amount"]
        if (r["status"] or "").lower() == "approved":
            emp_map[k]["a"] += 1; emp_map[k]["ap_amt"] += r["amount"]
    emp_rows_html = "".join(
        f"<tr><td>{name}</td><td>{eid}</td><td>{d['t']}</td><td>{d['a']}</td>"
        f"<td>{fmt(d['amt'])}</td><td>{fmt(d['ap_amt'])}</td></tr>"
        for (name, eid), d in sorted(emp_map.items(), key=lambda x: x[1]["amt"], reverse=True)
    )

    # ── Detail table (first 500 rows to keep PDF manageable)
    detail_limit = 500
    detail_html = ""
    for i, r in enumerate(rows[:detail_limit], 1):
        st = (r["status"] or "").lower()
        sc = "approved" if st == "approved" else ("rejected" if st == "rejected" else "flagged")
        detail_html += (
            f"<tr>"
            f"<td>{i}</td><td>{r['session_date']}</td><td>{r['employee_name']}</td>"
            f"<td>{r['department'] or '—'}</td><td>{r['category']}</td>"
            f"<td>{r['expense_date']}</td><td class='amt'>₹{r['amount']:,.0f}</td>"
            f"<td><span class='badge {sc}'>{r['status']}</span></td>"
            f"<td class='center'>{r['zoho_booked']}</td>"
            f"<td class='center'>{r['zoho_paid']}</td>"
            f"<td>{r['keka_status']}</td>"
            f"</tr>"
        )
    if len(rows) > detail_limit:
        detail_html += f"<tr><td colspan='11' style='text-align:center;color:#888;font-style:italic;'>… and {len(rows)-detail_limit} more claims (see Excel for complete list)</td></tr>"

    gen_date = datetime.now().strftime("%d %b %Y, %H:%M")
    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<title>Expense Report — All Batches</title>
<style>
  @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;900&display=swap');
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Inter',sans-serif;font-size:11px;color:#0f1629;background:#fff;padding:24px}}
  h1{{font-size:20px;font-weight:900;color:#e5007d;margin-bottom:4px}}
  h2{{font-size:13px;font-weight:700;color:#0f1629;margin:24px 0 10px;padding-bottom:6px;border-bottom:2px solid #e5007d}}
  .meta{{font-size:10px;color:#64748b;margin-bottom:20px}}
  /* KPI grid */
  .kpi-grid{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:24px}}
  .kpi{{border-radius:10px;padding:12px 14px;border:1px solid #e2e8f0}}
  .kpi .val{{font-size:22px;font-weight:900;line-height:1}}
  .kpi .lbl{{font-size:10px;color:#64748b;margin-top:3px}}
  .kpi .sub{{font-size:10px;font-weight:600;margin-top:2px}}
  .kpi-green{{background:#f0fdf4}}.kpi-green .val{{color:#059669}}.kpi-green .sub{{color:#059669}}
  .kpi-red  {{background:#fef2f2}}.kpi-red   .val{{color:#dc2626}}.kpi-red   .sub{{color:#dc2626}}
  .kpi-amber{{background:#fffbeb}}.kpi-amber .val{{color:#d97706}}.kpi-amber .sub{{color:#d97706}}
  .kpi-purple{{background:#f5f3ff}}.kpi-purple .val{{color:#7c3aed}}.kpi-purple .sub{{color:#7c3aed}}
  .kpi-blue{{background:#f0f9ff}}.kpi-blue .val{{color:#0ea5e9}}.kpi-blue .sub{{color:#0ea5e9}}
  .kpi-pink{{background:#fff0f8}}.kpi-pink .val{{color:#e5007d}}.kpi-pink .sub{{color:#e5007d}}
  .kpi-gray{{background:#f8fafc}}.kpi-gray .val{{color:#0f1629}}.kpi-gray .sub{{color:#64748b}}
  /* Tables */
  table{{width:100%;border-collapse:collapse;margin-bottom:24px;font-size:10px}}
  th{{background:#0f1629;color:#fff;padding:7px 8px;text-align:left;font-weight:600;font-size:10px}}
  td{{padding:5px 8px;border-bottom:1px solid #f1f5f9;vertical-align:middle}}
  tr:nth-child(even){{background:#f8fafc}}
  .amt{{font-weight:600;text-align:right}}
  .center{{text-align:center}}
  /* Badges */
  .badge{{display:inline-block;padding:2px 8px;border-radius:99px;font-size:9px;font-weight:700}}
  .badge.approved{{background:#dcfce7;color:#059669}}
  .badge.rejected{{background:#fee2e2;color:#dc2626}}
  .badge.flagged {{background:#fef3c7;color:#d97706}}
  /* Pipeline */
  .pipeline{{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:24px}}
  .pipe-tile{{border-radius:8px;padding:10px 12px;border:1px solid #e2e8f0}}
  .pipe-tile .p-val{{font-size:18px;font-weight:900}}
  .pipe-tile .p-lbl{{font-size:9px;color:#64748b;margin-top:2px}}
  .pipe-tile .p-sub{{font-size:9px;font-weight:600;margin-top:1px}}
  @media print{{
    body{{padding:12px;font-size:10px}}
    h2{{page-break-before:auto}}
    table{{page-break-inside:auto}}
    tr{{page-break-inside:avoid}}
    .no-break{{page-break-inside:avoid}}
  }}
</style>
</head>
<body>

<!-- Header -->
<h1>Wiom Finance — Expense Reimbursement Report</h1>
<div class="meta">All batches · Generated: {gen_date} · Total claims: {total}</div>

<!-- KPI Overview -->
<h2>Overall Summary</h2>
<div class="kpi-grid">
  <div class="kpi kpi-gray"><div class="val">{total}</div><div class="lbl">Total Claims</div><div class="sub">{fmt(s_amt)}</div></div>
  <div class="kpi kpi-green"><div class="val">{len(approved)}</div><div class="lbl">Approved</div><div class="sub">{fmt(ap_amt)}</div></div>
  <div class="kpi kpi-red"><div class="val">{len(rejected)}</div><div class="lbl">Rejected</div><div class="sub">{pct(len(rejected),total)}</div></div>
  <div class="kpi kpi-amber"><div class="val">{len(flagged)}</div><div class="lbl">Flagged/Review</div><div class="sub">{pct(len(flagged),total)}</div></div>
</div>

<!-- Approved Pipeline -->
<h2>Approved Claims Pipeline</h2>
<div class="pipeline">
  <div class="pipe-tile" style="background:#f5f3ff"><div class="p-val" style="color:#7c3aed">{len(zoho_bk)}</div><div class="p-lbl">Zoho Booked</div><div class="p-sub" style="color:#7c3aed">{fmt(zb_amt)}</div></div>
  <div class="pipe-tile" style="background:#f0f9ff"><div class="p-val" style="color:#0ea5e9">{len(zoho_pd)}</div><div class="p-lbl">Zoho Paid</div><div class="p-sub" style="color:#0ea5e9">{fmt(zp_amt)}</div></div>
  <div class="pipe-tile" style="background:#fffbeb"><div class="p-val" style="color:#d97706">{len(keka_ap)}</div><div class="p-lbl">Keka Approved</div><div class="p-sub" style="color:#d97706">{pct(len(keka_ap),len(approved))} of approved</div></div>
  <div class="pipe-tile" style="background:#fef2f2"><div class="p-val" style="color:#dc2626">{len(not_zh)}</div><div class="p-lbl">Pending Zoho</div><div class="p-sub" style="color:#dc2626">{fmt(nz_amt)}</div></div>
</div>

<!-- Category -->
<h2>Category-wise Breakdown</h2>
<table>
  <tr><th>Category</th><th>Total</th><th>Approved</th><th>Amount</th><th>Approval Rate</th></tr>
  {cat_rows}
</table>

<!-- Department -->
<h2>Department-wise Breakdown</h2>
<table>
  <tr><th>Department</th><th>Total</th><th>Approved</th><th>Amount</th><th>Approval Rate</th></tr>
  {dept_rows}
</table>

<!-- Employee -->
<h2>Employee-wise Breakdown</h2>
<table>
  <tr><th>Employee</th><th>ID</th><th>Total</th><th>Approved</th><th>Total Amount</th><th>Approved Amount</th></tr>
  {emp_rows_html}
</table>

<!-- Detail -->
<h2>All Claims — Detail</h2>
<table>
  <tr><th>#</th><th>Batch Date</th><th>Employee</th><th>Dept</th><th>Category</th>
      <th>Expense Date</th><th>Amount</th><th>Status</th><th>Zoho Booked</th><th>Zoho Paid</th><th>Keka</th></tr>
  {detail_html}
</table>

<script>window.addEventListener('load',()=>setTimeout(()=>window.print(),400));</script>
</body></html>"""


@app.get("/report/export-all")
def report_export_all(
    fmt: str = Query("excel", description="excel or pdf"),
    authorization: Optional[str] = _FastAPIHeader(None),
):
    """Export all batches — multi-sheet Excel or print-ready PDF HTML."""
    from services.auth import verify_token
    if not authorization or not authorization.lower().startswith("bearer "):
        raise HTTPException(401, "No token")
    if not verify_token(authorization[7:].strip()):
        raise HTTPException(401, "Invalid or expired token")

    rows = _collect_all_rows()
    date_str = datetime.now().strftime("%Y%m%d_%H%M")

    if fmt == "pdf":
        html = _build_pdf_html(rows)
        return StreamingResponse(
            iter([html.encode("utf-8")]),
            media_type="text/html; charset=utf-8",
            headers={"Content-Disposition": f"inline; filename=expense_report_{date_str}.html"},
        )

    # Default: Excel
    xlsx_bytes = _build_excel(rows)
    return StreamingResponse(
        iter([xlsx_bytes]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=expense_report_{date_str}.xlsx"},
    )


# ---------------------------------------------------------------------------
# Overview stats endpoint
# ---------------------------------------------------------------------------

@app.get("/overview/stats")
def overview_stats():
    """Aggregate KPI data across all sessions from result.json files."""
    now = datetime.now()
    this_month = (now.year, now.month)
    if now.month == 1:
        last_month = (now.year - 1, 12)
    else:
        last_month = (now.year, now.month - 1)

    this_month_total = 0.0
    last_month_total = 0.0
    total_approved_amount_all_time = 0.0
    total_pending_reimbursement = 0.0
    total_zoho_booked = 0.0
    total_approved = 0
    total_rejected = 0
    total_flagged = 0
    total_count = 0
    total_sessions = 0
    total_claims_all_time = 0
    total_zoho_booked_count = 0    # count of approved rows pushed to Zoho
    total_zoho_paid_count   = 0    # count where zoho_bill_status == "paid"
    total_zoho_paid_amount  = 0.0
    total_keka_approved_count     = 0    # keka_actioned in approve/approved/mark_paid
    total_approved_not_zoho_count = 0   # approved but NOT in Zoho yet
    category_breakdown: dict = {}
    claimants: dict = {}
    policy_violations = 0
    anomaly_count = 0

    # monthly_trend: last 6 months
    months_order = []
    for i in range(5, -1, -1):
        m = now.month - i
        y = now.year
        while m <= 0:
            m += 12
            y -= 1
        months_order.append((y, m))
    monthly_data: dict = {ym: {"total": 0.0, "approved": 0.0} for ym in months_order}

    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        entries = []

    for session_id in entries:
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            mtime = os.path.getmtime(result_path)
            session_dt = datetime.fromtimestamp(mtime)
            session_ym = (session_dt.year, session_dt.month)

            for enc in ("utf-8", "utf-8-sig", "cp1252", "latin-1"):
                try:
                    with open(result_path, encoding=enc) as f:
                        data = json.load(f)
                    break
                except (UnicodeDecodeError, json.JSONDecodeError):
                    continue
            else:
                continue

            rows = data.get("rows", [])
            total_sessions += 1
            total_claims_all_time += len(rows)

            # Keka postman batch? (postman_status.json present = downloaded from Keka)
            is_keka_batch = os.path.isfile(
                os.path.join(UPLOAD_BASE, session_id, "postman_status.json")
            )

            # Build map: row_index → {bill_id, zoho_status} from zoho_push.json
            zoho_pushed_indices: set = set()
            zoho_push_bill_map: dict = {}   # row_index → {"bill_id": str, "status": str}
            zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            if os.path.isfile(zoho_push_path):
                try:
                    with open(zoho_push_path, "r", encoding="utf-8") as _zf:
                        _zpd = json.load(_zf)
                    for _p in _zpd.get("pushed", []):
                        _ri = int(_p.get("row_index", -1))
                        zoho_pushed_indices.add(_ri)
                        zoho_push_bill_map[_ri] = {
                            "bill_id": _p.get("bill_id", ""),
                            "status":  (_p.get("status") or "").strip().lower(),
                        }
                except Exception:
                    pass

            for i, row in enumerate(rows):
                amount = float(row.get("claimed_amount") or 0)
                status = (row.get("status") or "").lower()
                nature = row.get("expense_nature") or row.get("expense_category") or "Other"
                employee = row.get("employee_name") or "Unknown"
                # A row is "in Zoho" if its row_index is in zoho_push.json OR has a zoho_bill_id set
                row_idx = row.get("row_index", -1)
                is_zoho_pushed = (row_idx in zoho_pushed_indices) or bool(row.get("zoho_bill_id"))
                remarks = row.get("remarks") or []

                total_count += 1

                if status == "approved":
                    total_approved += 1
                    total_approved_amount_all_time += amount
                    if is_zoho_pushed:
                        total_zoho_booked += amount
                        total_zoho_booked_count += 1
                        # zoho_bill_status: prefer result.json (written by sync), fall back to
                        # zoho_push.json status field (rarely populated, but handle it)
                        zoho_status = (row.get("zoho_bill_status") or "").strip().lower()
                        if not zoho_status:
                            push_entry = zoho_push_bill_map.get(row_idx, {})
                            zoho_status = push_entry.get("status", "")
                        if zoho_status == "paid":
                            total_zoho_paid_count  += 1
                            total_zoho_paid_amount += amount
                    else:
                        total_pending_reimbursement += amount
                        total_approved_not_zoho_count += 1
                    # Keka approval: explicit action OR row came from a Keka batch OR has keka_claim_id
                    keka_act = (row.get("keka_actioned") or "").strip().lower()
                    is_keka_row = (
                        keka_act in ("approve", "approved", "mark_paid")
                        or bool(row.get("keka_claim_id"))
                        or is_keka_batch
                    )
                    if is_keka_row:
                        total_keka_approved_count += 1
                elif status == "rejected":
                    total_rejected += 1
                else:
                    total_flagged += 1

                category_breakdown[nature] = category_breakdown.get(nature, 0.0) + amount

                if employee not in claimants:
                    claimants[employee] = {"name": employee, "amount": 0.0, "count": 0}
                claimants[employee]["amount"] += amount
                claimants[employee]["count"] += 1

                # Policy violations: remarks list containing "policy" (case-insensitive)
                if isinstance(remarks, list):
                    for remark in remarks:
                        if "policy" in str(remark).lower():
                            policy_violations += 1
                            break
                elif "policy" in str(remarks).lower():
                    policy_violations += 1

                # Anomaly flags count
                anomaly_flags = row.get("anomaly_flags") or []
                if anomaly_flags:
                    anomaly_count += 1

                # Monthly trend
                if session_ym in monthly_data:
                    monthly_data[session_ym]["total"] += amount
                    if status == "approved":
                        monthly_data[session_ym]["approved"] += amount

        except Exception:
            continue

    approval_rate = round(total_approved / total_count * 100, 1) if total_count > 0 else 0.0

    top_claimants = sorted(claimants.values(), key=lambda x: x["amount"], reverse=True)[:5]

    monthly_trend = []
    for (y, m) in months_order:
        label = datetime(y, m, 1).strftime("%b %Y")
        monthly_trend.append({
            "month": label,
            "total": round(monthly_data[(y, m)]["total"], 2),
            "approved": round(monthly_data[(y, m)]["approved"], 2),
        })

    return {
        # Claim counts
        "total_claims_all_time":  total_claims_all_time,
        "total_approved":         total_approved,
        "total_rejected":         total_rejected,
        "total_flagged":          total_flagged,
        # Amounts
        "total_approved_amount":      round(total_approved_amount_all_time, 2),
        "total_pending_reimbursement": round(total_pending_reimbursement, 2),
        "total_zoho_booked":          round(total_zoho_booked, 2),
        # Approved breakdown (Zoho + Keka)
        "zoho_booked_count":          total_zoho_booked_count,
        "zoho_paid_count":            total_zoho_paid_count,
        "zoho_paid_amount":           round(total_zoho_paid_amount, 2),
        "keka_approved_count":        total_keka_approved_count,
        "approved_not_zoho_count":    total_approved_not_zoho_count,
        # Rates & meta
        "approval_rate":   approval_rate,
        "total_sessions":  total_sessions,
        "category_breakdown": {k: round(v, 2) for k, v in category_breakdown.items()},
        "top_claimants": [
            {"name": c["name"], "amount": round(c["amount"], 2), "count": c["count"]}
            for c in top_claimants
        ],
        "policy_violations": policy_violations,
        "anomaly_count":     anomaly_count,
        "monthly_trend":     monthly_trend,
    }


# ---------------------------------------------------------------------------
# Policy rules endpoints
# ---------------------------------------------------------------------------

@app.get("/policy/rules")
def get_policy_rules():
    """Return current policy rules, or defaults if none saved."""
    return _load_policy_rules()


@app.post("/policy/rules")
async def save_policy_rules(request: Request):
    """Save policy rules to disk."""
    try:
        body = await request.json()
    except Exception as e:
        raise HTTPException(400, f"Invalid JSON: {e}")
    try:
        with open(POLICY_FILE, "w", encoding="utf-8") as f:
            json.dump(body, f, ensure_ascii=False, indent=2)
        return {"status": "ok"}
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/verify/zoho/{session_id}")
def verify_zoho(session_id: str):
    """Re-run AI verification on a completed Zoho push without re-pushing."""
    result = _load_session(session_id)
    if not result or result.processing_status != "completed":
        raise HTTPException(404, "Results not ready")
    from models.schemas import ExpenseStatus
    approved = [r for r in result.rows if r.status == ExpenseStatus.APPROVED]
    if not approved:
        return {"verification": [], "message": "No approved rows to verify"}
    rows_by_index = {r.row_index: r for r in result.rows}
    # Build minimal pushed list from approved rows (no expense_id needed for re-verify)
    pseudo_pushed = [
        {
            "row_index": r.row_index,
            "employee": r.employee_name,
            "amount": r.bill_amount or r.claimed_amount,
            "expense_id": "N/A",
            "account": "",
            "zoho_url": "",
        }
        for r in approved
    ]
    verification = verify_zoho_entries(pseudo_pushed, rows_by_index)
    return {"verification": verification}


@app.get("/zoho/status/{session_id}")
def zoho_status(session_id: str):
    """
    Return live Zoho vendor + bill statuses and session-level reimbursement metrics.
    Reads the stored push result (zoho_push.json) and queries Zoho for current statuses.
    """
    import sys as _sys
    try:
        result = _load_session(session_id)
    except Exception as _e:
        print(f"[zoho_status] _load_session failed: {_e}", file=_sys.stderr, flush=True)
        raise HTTPException(500, f"Session load error: {_e}")
    if not result:
        raise HTTPException(404, "Session not found")

    from services.zoho import CLIENT_ID, REFRESH_TOKEN, _get_access_token, _headers, ZOHO_API_BASE, ORG_ID
    import requests as _req

    # ── 1. Compute session metrics from stored rows ──────────────────────────
    by_employee: dict = {}
    by_category: dict = {}
    for row in result.rows:
        name = row.employee_name
        cat  = row.expense_category or "Uncategorized"
        amt  = row.bill_amount if row.bill_amount is not None else row.claimed_amount

        if name not in by_employee:
            by_employee[name] = {"employee_name": name, "employee_id": row.employee_id,
                                  "claimed": 0.0, "bill_amount": 0.0, "count": 0,
                                  "approved": 0, "rejected": 0, "flagged": 0}
        by_employee[name]["claimed"]     += row.claimed_amount
        by_employee[name]["bill_amount"] += amt
        by_employee[name]["count"]       += 1
        by_employee[name][row.status.value.lower()] += 1

        if cat not in by_category:
            by_category[cat] = {"category": cat, "total_claimed": 0.0, "count": 0}
        by_category[cat]["total_claimed"] += row.claimed_amount
        by_category[cat]["count"]         += 1

    from models.schemas import ExpenseStatus
    total_approved_amount = sum(
        (r.bill_amount if r.bill_amount is not None else r.claimed_amount)
        for r in result.rows if r.status == ExpenseStatus.APPROVED
    )

    metrics = {
        "total_approved_amount": round(total_approved_amount, 2),
        "by_employee": sorted(by_employee.values(), key=lambda x: x["claimed"], reverse=True),
        "by_category": sorted(by_category.values(), key=lambda x: x["total_claimed"], reverse=True),
    }

    # ── 2. Load stored push result ───────────────────────────────────────────
    push_path = os.path.join(_session_dir(session_id), "zoho_push.json")
    last_push = None
    if os.path.exists(push_path):
        with open(push_path, encoding="utf-8") as _f:
            last_push = json.load(_f)

    # ── 3. Query Zoho for live vendor + bill statuses ────────────────────────
    vendors = []
    bills   = []
    if last_push and CLIENT_ID and REFRESH_TOKEN:
        try:
            token = _get_access_token()
            seen_vendors: set = set()
            for entry in last_push.get("pushed", []):
                vendor_id = entry.get("vendor_id", "")
                bill_id   = entry.get("bill_id", "")

                # Vendor status
                if vendor_id and vendor_id not in seen_vendors:
                    seen_vendors.add(vendor_id)
                    try:
                        rv = _req.get(f"{ZOHO_API_BASE}/contacts/{vendor_id}",
                                      headers=_headers(token),
                                      params={"organization_id": ORG_ID}, timeout=10)
                        c = rv.json().get("contact", {})
                        vendors.append({
                            "vendor_id":     vendor_id,
                            "employee_name": entry.get("employee", ""),
                            "employee_id":   entry.get("employee_id", ""),
                            "status":        c.get("status", "unknown"),
                            "zoho_url":      f"https://books.zoho.in/app#/contacts/{vendor_id}",
                        })
                    except Exception:
                        vendors.append({"vendor_id": vendor_id,
                                        "employee_name": entry.get("employee", ""),
                                        "employee_id": entry.get("employee_id", ""),
                                        "status": "unknown", "zoho_url": ""})

                # Bill status
                if bill_id:
                    try:
                        rb = _req.get(f"{ZOHO_API_BASE}/bills/{bill_id}",
                                      headers=_headers(token),
                                      params={"organization_id": ORG_ID}, timeout=10)
                        b = rb.json().get("bill", {})
                        bills.append({
                            "bill_id":     bill_id,
                            "bill_number": b.get("bill_number", entry.get("bill_id", "")),
                            "employee":    entry.get("employee", ""),
                            "amount":      entry.get("amount", 0),
                            "status":      b.get("status", "unknown"),
                            "zoho_url":    entry.get("zoho_url", ""),
                        })
                    except Exception:
                        bills.append({
                            "bill_id":  bill_id,
                            "bill_number": "",
                            "employee": entry.get("employee", ""),
                            "amount":   entry.get("amount", 0),
                            "status":   "unknown",
                            "zoho_url": entry.get("zoho_url", ""),
                        })
        except Exception:
            pass  # Zoho unavailable — still return metrics

    return {
        "metrics":   metrics,
        "vendors":   vendors,
        "bills":     bills,
        "last_push": {
            "pushed_at": last_push.get("pushed_at") if last_push else None,
            "pushed_count": len(last_push.get("pushed", [])) if last_push else 0,
            "error_count":  len(last_push.get("errors", [])) if last_push else 0,
            "errors":       (last_push or {}).get("errors", []),
            "verification": (last_push or {}).get("verification", []),
        } if last_push else None,
    }


@app.post("/revalidate/{session_id}")
async def revalidate_session(session_id: str, background_tasks: BackgroundTasks,
                             request: Request,
                             authorization: Optional[str] = _FastAPIHeader(None)):
    """Re-run the full validation pipeline on existing uploaded files.
    Useful when the validation logic has been updated (e.g. vision fallback).
    """
    _require_admin(authorization)
    s_dir = _session_dir(session_id)
    excel_path = os.path.join(s_dir, "report.xlsx")
    zip_path   = os.path.join(s_dir, "bills.zip")

    if not os.path.exists(excel_path) or not os.path.exists(zip_path):
        raise HTTPException(404, "Original files not found — please re-upload")

    # Mark as processing immediately so the UI shows spinner
    result = _load_session(session_id)
    if result:
        result.processing_status = "processing"
        _save_session(session_id, result)

    try:
        from services.db import log_activity, update_session_record
        update_session_record(session_id, "processing")
        log_activity(_get_actor(authorization), "revalidate", "session", session_id, {},
                     ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    background_tasks.add_task(run_validation, session_id, excel_path, zip_path)
    return {"session_id": session_id, "status": "reprocessing"}


@app.post("/session/{session_id}/zoho-sync")
def session_zoho_sync(session_id: str, authorization: Optional[str] = _FastAPIHeader(None)):
    """
    Fetch live Zoho bill statuses for all pushed bills in this session.
    Matches by bill_number AND vendor_name (employee).
    Updates each matched row's zoho_bill_id and zoho_bill_status in result.json.
    Returns: { matched, unmatched, errors, total }
    """
    _require_admin(authorization)
    from services.zoho import fetch_bill_status, CLIENT_ID, REFRESH_TOKEN

    if not CLIENT_ID or not REFRESH_TOKEN:
        raise HTTPException(503, "Zoho credentials not configured")

    # Load zoho_push.json
    zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
    if not os.path.isfile(zoho_push_path):
        raise HTTPException(404, "No Zoho push data for this session. Push to Zoho first.")

    with open(zoho_push_path, "r", encoding="utf-8") as _f:
        push_data = json.load(_f)

    pushed = push_data.get("pushed", [])
    if not pushed:
        return {"matched": [], "unmatched": [], "errors": [], "total": 0}

    # Load session result
    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")

    matched   = []
    unmatched = []
    errors    = []

    for entry in pushed:
        row_index        = int(entry.get("row_index", -1))
        bill_id          = entry.get("bill_id", "")
        expected_bill_no = str(entry.get("bill_number", "")).strip()
        expected_vendor  = (entry.get("employee") or "").strip().lower()

        if row_index < 0 or not bill_id:
            errors.append({"row_index": row_index, "error": "Missing bill_id or row_index"})
            continue

        try:
            zoho_bill = fetch_bill_status(bill_id)
        except Exception as exc:
            errors.append({"row_index": row_index, "bill_id": bill_id,
                           "employee": entry.get("employee"), "error": str(exc)})
            continue

        zoho_bill_no = zoho_bill.get("bill_number", "")
        zoho_vendor  = (zoho_bill.get("vendor_name") or "").strip().lower()
        zoho_status  = zoho_bill.get("status", "unknown")

        # Match by vendor name only — bill_id already uniquely identifies the bill.
        # Vendor check guards against cross-session bill_id linkage.
        vendor_match = (
            zoho_vendor == expected_vendor
            or expected_vendor in zoho_vendor
            or zoho_vendor in expected_vendor
        )

        rec = {
            "row_index":   row_index,
            "employee":    entry.get("employee"),
            "bill_number": zoho_bill_no,
            "zoho_status": zoho_status,
            "amount":      zoho_bill.get("total"),
            "zoho_url":    entry.get("zoho_url", f"https://books.zoho.in/app#/bills/{bill_id}"),
        }

        # ── Correct row lookup: match by row.row_index, NOT by list position ──
        target_row = next((r for r in result.rows if r.row_index == row_index), None)

        if vendor_match and target_row is not None:
            target_row.zoho_bill_id     = bill_id
            target_row.zoho_bill_status = zoho_status
            matched.append(rec)
        else:
            rec["expected_vendor"] = expected_vendor
            rec["got_vendor"]      = zoho_vendor
            # Clear any stale zoho status on this row (was set by a previous wrong sync)
            if target_row is not None:
                target_row.zoho_bill_id     = None
                target_row.zoho_bill_status = None
            unmatched.append(rec)

    # ── Clear zoho status from rows that are NOT in zoho_push.json pushed list ──
    pushed_indices = {int(e.get("row_index", -1)) for e in pushed}
    for row in result.rows:
        if row.row_index not in pushed_indices:
            if row.zoho_bill_status is not None or row.zoho_bill_id is not None:
                row.zoho_bill_id     = None
                row.zoho_bill_status = None

    _save_session(session_id, result)

    return {
        "matched":   matched,
        "unmatched": unmatched,
        "errors":    errors,
        "total":     len(pushed),
    }


@app.post("/zoho/sync-all")
def zoho_sync_all(authorization: Optional[str] = _FastAPIHeader(None)):
    """
    Global Zoho status refresh across ALL completed sessions.
    Only processes rows that:
      - Have a zoho_bill_id (were previously pushed to Zoho)
      - Are NOT already marked 'paid' (no need to re-check those)
    Updates result.json in place and returns a summary.
    """
    _require_admin(authorization)
    from services.zoho import fetch_bill_status, CLIENT_ID, REFRESH_TOKEN

    if not CLIENT_ID or not REFRESH_TOKEN:
        raise HTTPException(503, "Zoho credentials not configured — set Zoho API keys in Config.")

    updated          = []
    already_paid     = []
    no_change        = []
    errors           = []
    sessions_touched = 0

    try:
        entries = os.listdir(UPLOAD_BASE)
    except Exception:
        entries = []

    # Sort oldest first so batch numbers are stable
    def _mtime(sid):
        p = os.path.join(UPLOAD_BASE, sid, "result.json")
        try: return os.path.getmtime(p)
        except: return 0

    for session_id in sorted(entries, key=_mtime):
        result_path = os.path.join(UPLOAD_BASE, session_id, "result.json")
        if not os.path.isfile(result_path):
            continue
        try:
            result = _load_session(session_id)
            if not result or result.processing_status != "completed":
                continue

            session_dirty = False

            # Build fallback map from zoho_push.json (row_index → bill_id)
            # This covers sessions where per-session sync was never run (zoho_bill_id is
            # only in zoho_push.json, not yet written back to result.json rows).
            _push_bill_map: dict = {}  # row_index (int) → bill_id (str)
            _zoho_push_path = os.path.join(UPLOAD_BASE, session_id, "zoho_push.json")
            if os.path.isfile(_zoho_push_path):
                try:
                    with open(_zoho_push_path, "r", encoding="utf-8") as _pf:
                        _pd = json.load(_pf)
                    for _pe in _pd.get("pushed", []):
                        _ri = int(_pe.get("row_index", -1))
                        _bid = _pe.get("bill_id", "")
                        if _ri >= 0 and _bid:
                            _push_bill_map[_ri] = _bid
                except Exception:
                    pass

            for row in result.rows:
                # Prefer zoho_bill_id from result.json; fall back to zoho_push.json
                bill_id = row.zoho_bill_id or _push_bill_map.get(row.row_index, "")
                if not bill_id:
                    continue  # never pushed to Zoho

                current_status = (row.zoho_bill_status or "").strip().lower()

                # Skip already-paid — no need to call Zoho
                if current_status == "paid":
                    already_paid.append({
                        "session_id":  session_id[:8],
                        "employee":    row.employee_name,
                        "bill_id":     bill_id,
                    })
                    continue

                # Fetch fresh status from Zoho
                try:
                    zoho_bill   = fetch_bill_status(bill_id)
                    new_status  = (zoho_bill.get("status") or "unknown").strip().lower()
                except Exception as exc:
                    errors.append({
                        "session_id": session_id[:8],
                        "employee":   row.employee_name,
                        "bill_id":    bill_id,
                        "error":      str(exc),
                    })
                    continue

                # Write bill_id back to result.json if it was missing (from push fallback)
                if not row.zoho_bill_id:
                    row.zoho_bill_id = bill_id
                    session_dirty = True

                if new_status != current_status:
                    row.zoho_bill_status = new_status
                    session_dirty = True
                    updated.append({
                        "session_id":  session_id[:8],
                        "employee":    row.employee_name,
                        "bill_id":     bill_id,
                        "old_status":  current_status or "—",
                        "new_status":  new_status,
                        "amount":      row.claimed_amount,
                    })
                else:
                    no_change.append({
                        "session_id": session_id[:8],
                        "employee":   row.employee_name,
                        "status":     current_status,
                    })

            if session_dirty:
                _save_session(session_id, result)
                sessions_touched += 1

        except Exception as exc:
            errors.append({"session_id": session_id[:8], "error": f"Session error: {exc}"})

    return {
        "updated":          len(updated),
        "already_paid":     len(already_paid),
        "no_change":        len(no_change),
        "errors":           len(errors),
        "sessions_touched": sessions_touched,
        "details": {
            "updated": updated,
            "errors":  errors,
        },
    }


@app.delete("/session/{session_id}")
def cleanup_session(session_id: str, authorization: Optional[str] = _FastAPIHeader(None)):
    """Delete uploaded files + cached results for a session."""
    _require_admin(authorization)
    s_dir = _session_dir(session_id)
    if os.path.exists(s_dir):
        shutil.rmtree(s_dir)
    _sessions.pop(session_id, None)
    return {"deleted": session_id}


# ── Keka Integration Endpoints ───────────────────────────────────────────────

class KekaSyncRequest(BaseModel):
    from_date:     str
    to_date:       str
    client_id:     Optional[str] = None
    client_secret: Optional[str] = None
    company:       Optional[str] = None
    claim_type:    Optional[str] = "reimbursement"
    waiting_on:    Optional[str] = "Prateek Kimothi"


async def run_keka_sync(
    session_id: str,
    from_date: str,
    to_date: str,
    client_id: Optional[str],
    client_secret: Optional[str],
    company: Optional[str],
    claim_type: str,
    waiting_on: Optional[str],
):
    """Background task: fetch Keka claims via OAuth2 → OCR → validate → store session."""
    from services.keka import (
        get_keka_token, fetch_expense_claims,
        claims_to_expense_rows,
        _extract_attachments_from_expense,
        download_all_attachments_to_dir,
    )
    from models.schemas import ExpenseRow, ExpenseStatus

    result = SessionResult(session_id=session_id, processing_status="processing", current_step="Connecting to Keka…")
    _save_session(session_id, result)

    def _step(msg: str):
        result.current_step = msg
        _save_session(session_id, result)

    try:
        _step("Authenticating with Keka…")
        token = get_keka_token(client_id=client_id, client_secret=client_secret)

        _step(f"Fetching expense claims from Keka ({from_date} → {to_date})…")
        claims = fetch_expense_claims(
            token, from_date=from_date, to_date=to_date,
            company=company,
        )

        # ── Optional filter by approver name ─────────────────────────────────
        if waiting_on:
            wl = waiting_on.lower()
            filtered_wo = [
                c for c in claims
                if wl in (
                    c.get("waitingOn") or c.get("currentApproverName") or
                    (c.get("approver") or {}).get("name") or
                    c.get("pendingWith") or ""
                ).lower()
            ]
            claims = filtered_wo if filtered_wo else claims

        if not claims:
            result = SessionResult(
                session_id=session_id,
                processing_status="completed",
                total_claims=0, rows=[],
                error=f"No expense claims found between {from_date} and {to_date}.",
            )
            _save_session(session_id, result)
            return

        _step(f"Found {len(claims)} claims — fetching claim details…")

        from services.keka import fetch_claim_details
        from services.keka_browser import (
            KEKA_EMAIL as _KEKA_EMAIL, KEKA_PASSWORD as _KEKA_PWD,
            is_authenticated as _is_authed,
            bulk_download_claims_attachments,
            download_claim_attachments_direct,
        )
        import logging as _klog
        _klog = _klog.getLogger(__name__)

        # The v1 list API already returns expenses[].attachments[].id — no detail fetch needed.
        # Build a claim_details_map from the claims themselves.
        claim_details_map: dict = {}
        for i, c in enumerate(claims):
            cid = str(c.get("id") or c.get("claimId") or c.get("expenseClaimId") or i)
            claim_details_map[cid] = c

        bills_dir = os.path.join(_session_dir(session_id), "bills")
        os.makedirs(bills_dir, exist_ok=True)
        all_files: dict = {}   # display_key → absolute_path

        _creds_set = bool(
            _KEKA_EMAIL and _KEKA_PWD and
            "your_keka" not in _KEKA_EMAIL and
            "your_keka" not in _KEKA_PWD
        )
        can_download = _creds_set and _is_authed()

        if can_download:
            # ── Direct download using attachment IDs from claim list API ──────
            # The v1 API returns expenses[].attachments[].{id, name} for each claim.
            # We download them in parallel using the browser session cookies.

            # Build per-claim attachment lists
            claim_attachments: list[tuple[str, list[dict]]] = []
            for c in claims:
                cid = str(c.get("id") or c.get("claimId") or "")
                if not cid:
                    continue
                expenses = (c.get("expenses") or c.get("expenseItems") or
                            c.get("lineItems") or c.get("items") or [])
                if any(exp.get("attachments") for exp in expenses):
                    claim_attachments.append((cid, expenses))

            total_att = sum(
                len(exp.get("attachments", []))
                for _, expenses in claim_attachments
                for exp in expenses
            )
            _step(f"Downloading {total_att} attachments from {len(claim_attachments)} claims…")

            from concurrent.futures import ThreadPoolExecutor as _TPE

            def _dl_claim(args):
                cid, expenses = args
                claim_dir = os.path.join(bills_dir, cid)
                try:
                    return cid, download_claim_attachments_direct(
                        cid, expenses, claim_dir, company=company
                    )
                except Exception as e:
                    _klog.warning("Direct download error claim %s: %s", cid, e)
                    return cid, []

            with _TPE(max_workers=8) as pool:
                dl_results = list(pool.map(_dl_claim, claim_attachments))

            for cid, fpaths in dl_results:
                for fp in fpaths:
                    fname = os.path.basename(fp)
                    dk = f"{cid}/{fname}"
                    all_files[dk] = fp

            _klog.info("Direct download: %d files from %d claims", len(all_files), len(dl_results))

            # ── Browser fallback if direct download got nothing ────────────────
            if not all_files and claim_attachments:
                _step("Direct download got 0 files — trying browser interception fallback…")
                claim_ids = [cid for cid, _ in claim_attachments]

                def _progress(done, total):
                    result.current_step = f"Browser fallback — {done}/{total} claims…"
                    _save_session(session_id, result)

                try:
                    bulk_results = bulk_download_claims_attachments(
                        claim_ids=claim_ids,
                        out_dir=bills_dir,
                        company=company,
                        on_progress=_progress,
                    )
                    for cid, fpaths in bulk_results.items():
                        for fp in fpaths:
                            fname = os.path.basename(fp)
                            dk = f"{cid}/{fname}"
                            all_files[dk] = fp
                except Exception as e:
                    _klog.warning("Browser fallback error: %s", e)

            _step(f"Downloaded {len(all_files)} bill files…")

        elif _creds_set and not _is_authed():
            _klog.warning(
                "Keka credentials set but browser session not active — "
                "bill download skipped. User must complete 2FA login."
            )
            _step("Login required for bill download — complete 2FA via the login button.")
        else:
            _klog.info("KEKA_EMAIL/PASSWORD not set — skipping bill download.")

        # ── 4. Build ExpenseRow-compatible dicts ──────────────────────────────
        _step("Preparing expense data…")
        row_dicts = claims_to_expense_rows(claims, claim_details_map)

        if not row_dicts:
            result = SessionResult(
                session_id=session_id,
                processing_status="completed",
                total_claims=0,
                rows=[],
                error="Claims found but no expense line items could be extracted.",
            )
            _save_session(session_id, result)
            return

        # ── Aggregate per-expense rows by Keka claim_id ─────────────────────
        # A single Keka claim can have multiple expenses (e.g. 1 conveyance +
        # 1 food). Each expense has its own bills BUT they all sit in the
        # same claim folder. Building one row per expense leads to mis-matched
        # bills (food bill assigned to conveyance row, etc.).
        # Fix: One row per claim. claimed_amount = sum of all expenses.
        # All bills in the claim folder belong to this one aggregated row.
        from collections import OrderedDict
        grouped: "OrderedDict[str, dict]" = OrderedDict()
        for rd in row_dicts:
            cid = rd.get("keka_claim_id", "")
            if not cid:
                # No claim id → keep as separate row
                cid = f"_orphan_{len(grouped)}"
            if cid not in grouped:
                grouped[cid] = {
                    "keka_claim_id":     rd.get("keka_claim_id", ""),
                    "keka_claim_number": rd.get("keka_claim_number", ""),
                    "employee_name":     rd.get("employee_name", ""),
                    "employee_id":       rd.get("employee_id", ""),
                    "employee_email":    rd.get("employee_email", ""),
                    "expense_date":      rd.get("expense_date", ""),
                    "expense_category":  rd.get("expense_category", ""),
                    "expense_nature":    rd.get("expense_nature", ""),
                    "description":       rd.get("description", ""),
                    "claimed_amount":    0.0,
                    "expense_ids":       [],
                    "categories":        set(),
                    "descriptions":      [],
                }
            g = grouped[cid]
            g["claimed_amount"] += float(rd.get("claimed_amount", 0) or 0)
            if rd.get("keka_expense_id"):
                g["expense_ids"].append(rd["keka_expense_id"])
            cat = rd.get("expense_category", "") or rd.get("expense_nature", "")
            if cat: g["categories"].add(cat)
            desc = rd.get("description", "")
            if desc and desc not in g["descriptions"]:
                g["descriptions"].append(desc)
            # Use earliest expense date as the claim's expense_date
            d_new = rd.get("expense_date", "")
            d_cur = g["expense_date"]
            if d_new and (not d_cur or d_new < d_cur):
                g["expense_date"] = d_new

        _klog.info("Aggregated %d expense rows into %d claim rows",
                   len(row_dicts), len(grouped))

        # Build ExpenseRow objects (one per claim)
        rows: list[ExpenseRow] = []
        for idx, (cid, g) in enumerate(grouped.items()):
            real_cid = g["keka_claim_id"]
            # Bills strictly from THIS claim's folder
            claim_files = {dk: fp for dk, fp in all_files.items()
                           if real_cid and dk.startswith(real_cid + "/")}

            # Compose category & description
            categories = sorted(g["categories"])
            cat_label  = " + ".join(categories[:3]) if categories else (g["expense_category"] or "")
            desc_label = " | ".join(g["descriptions"][:3]) if g["descriptions"] else g["description"]

            row = ExpenseRow(
                row_index=idx,
                employee_name=g["employee_name"],
                employee_id=g["employee_id"],
                expense_date=g["expense_date"],
                expense_category=cat_label,
                description=desc_label,
                claimed_amount=round(g["claimed_amount"], 2),
                expense_nature=g["expense_nature"],
                keka_claim_id=real_cid,
                keka_claim_number=g["keka_claim_number"],
                keka_expense_id=", ".join(g["expense_ids"][:5]),
                employee_email=g["employee_email"],
                matched_files=list(claim_files.keys()),
                matched_file=list(claim_files.keys())[0] if claim_files else None,
            )
            rows.append(row)

        # ── 5. OCR all bill files ─────────────────────────────────────────────
        # Keka API v1 does not expose attachment download endpoints.
        # If no files were downloaded, skip OCR and mark all rows Flagged for manual review.
        loop = asyncio.get_event_loop()
        ocr_map: dict = {}

        unique_paths = list(set(all_files.values()))

        if not unique_paths:
            # Data-only mode: no bills available — flag everything for manual review
            _step(f"No bill files available (Keka API v1 limitation) — flagging {len(rows)} rows for review…")
            for row in rows:
                row.status = ExpenseStatus.FLAGGED
                row.remarks.append("Bill PDF not downloadable via Keka API — manual review required")
        else:
            _step(f"Running OCR on {len(unique_paths)} bill files…")
            from concurrent.futures import ThreadPoolExecutor
            MAX_OCR_WORKERS = min(32, max(8, len(unique_paths)))
            path_hints: dict = {}
            for row in rows:
                for key in (row.matched_files or ([row.matched_file] if row.matched_file else [])):
                    fp = all_files.get(key)
                    if fp and row.claimed_amount:
                        path_hints.setdefault(fp, float(row.claimed_amount))
            with ThreadPoolExecutor(max_workers=MAX_OCR_WORKERS) as pool:
                tasks = {
                    fp: loop.run_in_executor(pool, process_bill_ocr, fp, path_hints.get(fp))
                    for fp in unique_paths
                }
                ocr_results_list = await asyncio.gather(*tasks.values(), return_exceptions=True)
                from services.ocr import OCRResult as _OCRResult
                for fp, ocr_res in zip(tasks.keys(), ocr_results_list):
                    if isinstance(ocr_res, Exception):
                        ocr_res = _OCRResult()
                    for dk, fp2 in all_files.items():
                        if fp2 == fp:
                            ocr_map[dk] = ocr_res

            from services.ocr import flush_ocr_cache
            flush_ocr_cache()

            for row in rows:
                row.ocr_results = [ocr_map[k] for k in row.matched_files if k in ocr_map]
                if row.ocr_results:
                    row.ocr_result = row.ocr_results[0]

            # ── 6. Validate ───────────────────────────────────────────────────
            from services.validator import prefetch_usd_rates
            prefetch_usd_rates(rows)
            _step(f"Validating {len(rows)} expense rows…")
            rows = validate_expenses(rows, ocr_map, file_map=all_files)

            # Vendor suggestion + anomaly detection
            try:
                from services.vendor_master import suggest_category as _suggest_cat
                for row in rows:
                    if row.ocr_result and row.ocr_result.vendor_name:
                        s = _suggest_cat(row.ocr_result.vendor_name)
                        if s:
                            row.suggested_category = s["category"]
                            row.suggested_category_confidence = s["confidence"]
                            row.suggested_vendor_type = s.get("type")
            except Exception:
                pass
            try:
                from services.anomaly import detect_anomalies as _detect
                rows = _detect(rows, session_id, UPLOAD_BASE)
            except Exception:
                pass

        approved = sum(1 for r in rows if r.status == ExpenseStatus.APPROVED)
        rejected = sum(1 for r in rows if r.status == ExpenseStatus.REJECTED)
        flagged  = sum(1 for r in rows if r.status == ExpenseStatus.FLAGGED)

        result = SessionResult(
            session_id=session_id,
            processing_status="completed",
            total_claims=len(rows),
            approved=approved,
            rejected=rejected,
            flagged=flagged,
            rows=rows,
            login_required=(_creds_set and not _is_authed()),
            bills_downloaded=len(all_files),
        )
        _save_session(session_id, result)
        try:
            from services.db import update_session_record
            update_session_record(session_id, "completed", len(rows), approved, rejected, flagged)
        except Exception:
            pass

    except Exception as exc:
        import traceback
        result = SessionResult(
            session_id=session_id,
            processing_status="error",
            error=f"{type(exc).__name__}: {str(exc)}\n{traceback.format_exc()[-500:]}",
        )
        _save_session(session_id, result)
        try:
            from services.db import update_session_record
            update_session_record(session_id, "error")
        except Exception:
            pass


# ── Postman-style direct fetch → Process → Output workflow ──────────────────

class PostmanSyncRequest(BaseModel):
    from_date:        str
    to_date:          str
    company:          Optional[str] = None
    waiting_on:       Optional[str] = None                # NO default approver filter
    in_approval_only: bool = True                          # match bulk receipt scope


@app.post("/keka/postman/fetch")
async def keka_postman_fetch(body: PostmanSyncRequest, background_tasks: BackgroundTasks):
    """
    Postman-style flow:
      1. Fetch claims via Keka OAuth API → save as Excel in input folder
      2. Download all bills (every endpoint tried) → bills/ subfolder
      3. Zip bills → bills.zip
      4. Save metadata.json
    Returns the input folder path + a session_id usable for the next step.
    """
    session_id = str(uuid.uuid4())
    s_dir = _session_dir(session_id)
    os.makedirs(s_dir, exist_ok=True)

    # Save status file so UI can poll
    status = {
        "session_id":       session_id,
        "stage":             "fetching",
        "current_step":      "Initializing…",
        "from_date":         body.from_date,
        "to_date":           body.to_date,
        "input_dir":         None,
        "output_dir":        None,
        "claims_count":      0,
        "bills_downloaded":  0,
        "bills_attempted":   0,
        "error":             None,
    }
    status_path = os.path.join(s_dir, "postman_status.json")
    with open(status_path, "w", encoding="utf-8") as f:
        json.dump(status, f, indent=2)

    async def _run():
        loop = asyncio.get_event_loop()
        from concurrent.futures import ThreadPoolExecutor as _TPE
        from services.keka_postman import run_postman_sync

        def _on_step(msg: str):
            try:
                with open(status_path, encoding="utf-8") as f:
                    s = json.load(f)
                s["current_step"] = msg
                with open(status_path, "w", encoding="utf-8") as f:
                    json.dump(s, f, indent=2)
            except Exception:
                pass

        try:
            with _TPE(max_workers=1) as pool:
                meta = await loop.run_in_executor(
                    pool,
                    lambda: run_postman_sync(
                        body.from_date, body.to_date,
                        company=body.company, waiting_on=body.waiting_on,
                        on_step=_on_step,
                    ),
                )
            status.update(meta)
            status["stage"] = "fetched"
            status["current_step"] = (
                f"Fetched {meta['claims_count']} claims, {meta['rows_count']} rows, "
                f"{meta['bills_downloaded']}/{meta['bills_attempted']} bills"
            )
            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status, f, indent=2)
        except Exception as e:
            import traceback
            status["stage"] = "error"
            status["error"] = f"{type(e).__name__}: {e}"
            status["traceback"] = traceback.format_exc()[-800:]
            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status, f, indent=2)

    background_tasks.add_task(_run)
    return {"session_id": session_id, "stage": "fetching"}


@app.get("/keka/postman/status/{session_id}")
def keka_postman_status(session_id: str):
    """Poll status of a Postman fetch session."""
    s_dir = _session_dir(session_id)
    status_path = os.path.join(s_dir, "postman_status.json")
    if not os.path.exists(status_path):
        raise HTTPException(404, "Postman session not found")
    with open(status_path, encoding="utf-8") as f:
        return json.load(f)


@app.post("/keka/postman/process/{session_id}")
async def keka_postman_process(session_id: str, background_tasks: BackgroundTasks):
    """
    Process a previously fetched Postman folder:
      - Reads claims.xlsx + bills.zip from the input folder
      - Runs OCR + matching + validation
      - Saves validation_report.xlsx into the output folder
    """
    s_dir = _session_dir(session_id)
    status_path = os.path.join(s_dir, "postman_status.json")
    if not os.path.exists(status_path):
        raise HTTPException(404, "Postman session not found")

    with open(status_path, encoding="utf-8") as f:
        status = json.load(f)

    inp_dir = status.get("input_dir")
    out_dir = status.get("output_dir")
    if not inp_dir or not os.path.isdir(inp_dir):
        raise HTTPException(400, "Input folder missing — run /keka/postman/fetch first")

    excel_path = os.path.join(inp_dir, "claims.xlsx")
    zip_path   = os.path.join(inp_dir, "bills.zip")

    if not os.path.exists(excel_path):
        raise HTTPException(400, f"claims.xlsx not found in {inp_dir}")

    # Copy the inputs into the session directory so the existing pipeline picks them up
    shutil.copy(excel_path, os.path.join(s_dir, "report.xlsx"))
    if os.path.exists(zip_path):
        shutil.copy(zip_path, os.path.join(s_dir, "bills.zip"))
    else:
        # Empty placeholder zip so the pipeline runs even with 0 bills
        empty_zip = os.path.join(s_dir, "bills.zip")
        with zipfile.ZipFile(empty_zip, "w") as _:
            pass

    status["stage"] = "processing"
    status["current_step"] = "Processing through validation pipeline…"
    with open(status_path, "w", encoding="utf-8") as f:
        json.dump(status, f, indent=2)

    # Kick off normal validation
    background_tasks.add_task(
        run_validation,
        session_id,
        os.path.join(s_dir, "report.xlsx"),
        os.path.join(s_dir, "bills.zip"),
    )

    # When validation completes, copy the report into the output folder
    async def _save_output_when_done():
        for _ in range(600):  # 5 minute max
            await asyncio.sleep(0.5)
            r = _load_session(session_id)
            if r and r.processing_status in ("completed", "error"):
                break
        if r and r.processing_status == "completed":
            try:
                from services.exporter import export_to_excel
                out_excel = os.path.join(out_dir, "validation_report.xlsx")
                export_to_excel(r.rows, out_excel)
                # Save summary JSON too
                summary = {
                    "session_id":   session_id,
                    "completed_at": __import__("datetime").datetime.now().isoformat(),
                    "total_claims": r.total_claims,
                    "approved":     r.approved,
                    "rejected":     r.rejected,
                    "flagged":      r.flagged,
                    "input_dir":    inp_dir,
                    "output_dir":   out_dir,
                }
                with open(os.path.join(out_dir, "summary.json"), "w", encoding="utf-8") as f:
                    json.dump(summary, f, indent=2)
                status["stage"] = "completed"
                status["current_step"] = f"Output saved → {out_excel}"
                status["output_excel"] = out_excel
                status["validation_session_id"] = session_id
                status.update(summary)
                with open(status_path, "w", encoding="utf-8") as f:
                    json.dump(status, f, indent=2)
            except Exception as e:
                status["error"] = f"Output save failed: {e}"
                with open(status_path, "w", encoding="utf-8") as f:
                    json.dump(status, f, indent=2)

    background_tasks.add_task(_save_output_when_done)
    return {"session_id": session_id, "stage": "processing"}


@app.get("/keka/postman/download/{session_id}/{folder_type}")
def keka_postman_download(session_id: str, folder_type: str):
    """Download input or output folder as a zip."""
    if folder_type not in ("input", "output"):
        raise HTTPException(400, "folder_type must be 'input' or 'output'")
    s_dir = _session_dir(session_id)
    status_path = os.path.join(s_dir, "postman_status.json")
    if not os.path.exists(status_path):
        raise HTTPException(404, "Postman session not found")
    with open(status_path, encoding="utf-8") as f:
        status = json.load(f)
    src = status.get(f"{folder_type}_dir")
    if not src or not os.path.isdir(src):
        raise HTTPException(404, f"{folder_type} folder missing")
    zip_path = os.path.join(s_dir, f"{folder_type}_folder.zip")
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _, files in os.walk(src):
            for f in files:
                full = os.path.join(root, f)
                arc  = os.path.relpath(full, os.path.dirname(src))
                zf.write(full, arc)
    return FileResponse(zip_path, filename=f"{folder_type}_{session_id[:8]}.zip",
                        media_type="application/zip")


@app.post("/keka/postman/interactive-download")
async def keka_postman_interactive(body: PostmanSyncRequest, background_tasks: BackgroundTasks):
    """
    Open a visible Chrome window logged in to Keka and fully automate:
    set date range, set Status→In Approval Process, click Run,
    click Excel download link, click Bulk Receipt. No manual steps needed.
    """
    # ── Auto-correct inverted date range (from > to → swap) ────────────────
    from datetime import date as _date
    try:
        _fd = _date.fromisoformat(body.from_date)
        _td = _date.fromisoformat(body.to_date)
        if _fd > _td:
            log.warning("Auto-swapping inverted date range: %s > %s", body.from_date, body.to_date)
            body.from_date, body.to_date = body.to_date, body.from_date
    except Exception:
        pass

    session_id = str(uuid.uuid4())
    s_dir = _session_dir(session_id)
    os.makedirs(s_dir, exist_ok=True)

    status = {
        "session_id":     session_id,
        "stage":          "awaiting_user_action",
        "current_step":   "Opening browser…",
        "from_date":      body.from_date,
        "to_date":        body.to_date,
        "input_dir":      None,
        "output_dir":     None,
        "downloaded_path": None,
        "error":          None,
    }
    status_path = os.path.join(s_dir, "postman_status.json")
    with open(status_path, "w", encoding="utf-8") as f:
        json.dump(status, f, indent=2)

    async def _run():
        loop = asyncio.get_event_loop()
        from concurrent.futures import ThreadPoolExecutor as _TPE
        from services.keka_postman import (
            download_bulk_receipts_fully_auto,
            _make_session_dirs, fetch_claims_to_excel, zip_bills_folder,
        )
        import shutil as _sh
        import zipfile as _zf

        def _on_step(msg: str):
            try:
                with open(status_path, encoding="utf-8") as f:
                    s = json.load(f)
                s["current_step"] = msg
                with open(status_path, "w", encoding="utf-8") as f:
                    json.dump(s, f, indent=2)
            except Exception:
                pass

        try:
            # 1. Setup folders
            postman_session_id, inp_dir, out_dir = _make_session_dirs(body.from_date, body.to_date)
            status["input_dir"]  = inp_dir
            status["output_dir"] = out_dir
            status["postman_session_id"] = postman_session_id
            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status, f, indent=2)

            # 2. Headless automation — run in a fresh subprocess so Playwright
            #    gets its own ProactorEventLoop (Uvicorn sets SelectorEventLoop
            #    on Windows which prevents Chrome from launching).
            _on_step("Backend mein automation shuru ho rahi hai…")

            def _run_via_subprocess():
                import subprocess as _sp
                worker = os.path.join(os.path.dirname(__file__), "keka_auto_worker.py")
                cmd = [sys.executable, worker,
                       body.from_date, body.to_date, inp_dir, status_path]
                if body.company:
                    cmd += ["--company", body.company]
                # CREATE_NO_WINDOW keeps it hidden but still needs a title;
                # the worker fixes that via SetConsoleTitleW before Playwright starts.
                _popen_flags = 0
                if sys.platform == "win32":
                    _popen_flags = _sp.CREATE_NO_WINDOW
                proc = _sp.Popen(
                    cmd,
                    stdout=_sp.PIPE, stderr=_sp.PIPE,
                    cwd=os.path.dirname(__file__),
                    creationflags=_popen_flags,
                )
                stdout, stderr = proc.communicate()
                if proc.returncode != 0:
                    raise RuntimeError(
                        f"Worker process failed (exit {proc.returncode}): "
                        f"{stderr.decode(errors='replace')[:800]}"
                    )
                return json.loads(stdout)

            with _TPE(max_workers=1) as pool:
                result = await loop.run_in_executor(pool, _run_via_subprocess)

            # ── Stale-session early exit ───────────────────────────────────
            # If Keka redirected us to the login page, our saved cookies are
            # invalid. Tell the frontend so it can prompt the user to re-login.
            if result.get("session_stale"):
                status["stage"]            = "session_expired"
                status["error"]            = result.get("error") or "Keka session expired"
                status["login_required"]   = True
                status["current_step"]     = (
                    "⚠ Keka session expired — click 'Login to Keka' button to "
                    "send OTP and re-authenticate, then try again."
                )
                with open(status_path, "w", encoding="utf-8") as f:
                    json.dump(status, f, indent=2)
                return        # exit _run early

            xlsx_path = result.get("xlsx_path")
            zip_path  = result.get("zip_path")
            status["downloaded_excel"] = xlsx_path
            status["downloaded_zip"]   = zip_path

            # 3. Use Keka's Excel as source of truth for claims.xlsx
            target_xlsx = os.path.join(inp_dir, "claims.xlsx")
            if xlsx_path and os.path.exists(xlsx_path):
                _on_step(f"Using Keka's Excel as source of truth → claims.xlsx")
                shutil.copy(xlsx_path, target_xlsx)

                # Count rows — Keka's Excel has 2 metadata rows before actual headers
                try:
                    import pandas as _pd
                    from services.matcher import _find_header_row
                    _hrow = _find_header_row(target_xlsx)
                    _df = _pd.read_excel(target_xlsx, header=_hrow, dtype=str)
                    _df.columns = [str(c).strip() for c in _df.columns]
                    claim_col = next((c for c in _df.columns if "claim" in c.lower() and "number" in c.lower()), None)
                    status["rows_count"]   = len(_df)
                    status["claims_count"] = _df[claim_col].nunique() if claim_col else len(_df)
                except Exception:
                    pass
            else:
                _on_step("⚠ No Excel captured — automation could not download it")

            # 4. Extract bills zip into bills/
            if zip_path and os.path.exists(zip_path):
                bills_dir = os.path.join(inp_dir, "bills")
                os.makedirs(bills_dir, exist_ok=True)
                try:
                    with _zf.ZipFile(zip_path, "r") as zf:
                        zf.extractall(bills_dir)
                    bills_zip = os.path.join(inp_dir, "bills.zip")
                    zip_bills_folder(bills_dir, bills_zip)
                    n_files = sum(len(f) for _, _, f in os.walk(bills_dir))
                    status["bills_downloaded"] = n_files
                    _on_step(f"Extracted {n_files} bill files")
                except Exception as e:
                    _on_step(f"Zip extract error: {e}")

            # 5. Final stage
            if xlsx_path and zip_path:
                status["stage"] = "fetched"
                _on_step(f"✓ Both Excel ({status.get('rows_count','?')} rows) + Bills captured")
            elif xlsx_path or zip_path:
                status["stage"] = "fetched"
                _on_step(f"⚠ Got {'Excel only' if xlsx_path else 'Bills only'} — partial capture")
            else:
                status["stage"] = "no_download"
                _on_step("⚠ No downloads captured — user may have closed browser")

            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status, f, indent=2)

        except Exception as e:
            import traceback
            status["stage"] = "error"
            status["error"] = f"{type(e).__name__}: {e}"
            status["traceback"] = traceback.format_exc()[-800:]
            with open(status_path, "w", encoding="utf-8") as f:
                json.dump(status, f, indent=2)

    background_tasks.add_task(_run)
    return {"session_id": session_id, "stage": "awaiting_user_action"}


@app.get("/keka/postman/list")
def keka_postman_list():
    """List all Postman sync sessions on disk."""
    from services.keka_postman import INPUT_ROOT
    sessions = []
    if os.path.isdir(INPUT_ROOT):
        for entry in sorted(os.listdir(INPUT_ROOT), reverse=True):
            meta_path = os.path.join(INPUT_ROOT, entry, "metadata.json")
            if os.path.exists(meta_path):
                try:
                    with open(meta_path, encoding="utf-8") as f:
                        sessions.append(json.load(f))
                except Exception:
                    pass
    return {"sessions": sessions}


@app.post("/keka/sync")
async def keka_sync(body: KekaSyncRequest, background_tasks: BackgroundTasks,
                    request: Request,
                    authorization: Optional[str] = _FastAPIHeader(None)):
    """
    Start a Keka sync: fetch pending reimbursement claims for the date range,
    run OCR + validation, return a session_id to poll.
    """
    session_id = str(uuid.uuid4())
    s_dir = _session_dir(session_id)
    os.makedirs(s_dir, exist_ok=True)

    # Save sync params for reference
    with open(os.path.join(s_dir, "keka_params.json"), "w") as f:
        json.dump(body.model_dump(), f, indent=2)

    actor     = _get_actor(authorization)
    client_ip = request.client.host if request.client else ""
    try:
        from services.db import create_session_record, log_activity
        create_session_record(session_id, actor,
                              f"Keka {body.from_date} → {body.to_date}",
                              source="keka_sync", ip_address=client_ip)
        log_activity(actor, "keka_sync", "session", session_id,
                     {"from_date": body.from_date, "to_date": body.to_date},
                     ip_address=client_ip)
    except Exception:
        pass

    background_tasks.add_task(
        run_keka_sync,
        session_id,
        body.from_date,
        body.to_date,
        body.client_id,
        body.client_secret,
        body.company,
        body.claim_type or "reimbursement",
        body.waiting_on,
    )
    return {"session_id": session_id, "status": "processing"}


class _KekaBulkSyncRequest(BaseModel):
    from_date:  str
    to_date:    str
    company:    Optional[str] = None
    waiting_on: Optional[str] = "Prateek Kimothi"




class KekaActionRequest(BaseModel):
    action:         str             # "approve" | "reject" | "mark_paid"
    claim_ids:      list[str]
    reason:         Optional[str]  = None
    company:        Optional[str]  = None
    payment_mode:   Optional[str]  = "BankTransfer"
    payment_date:   Optional[str]  = None    # YYYY-MM-DD; defaults to today
    reference_no:   Optional[str]  = ""
    # claim_number → employee_name; verified against Keka row before acting
    employee_names: Optional[dict] = None


@app.post("/keka/action/{session_id}")
def keka_action(session_id: str, body: KekaActionRequest, request: Request,
                authorization: Optional[str] = _FastAPIHeader(None)):
    """
    Approve / Reject / Mark-as-Paid Keka claims via OAuth2.
      - approve:   Approves the claim (no email)
      - reject:    Rejects + Keka emails the employee with the reason
      - mark_paid: Marks an approved claim as paid (records payment)
    """
    _require_admin(authorization)
    from services.keka import approve_claim, reject_claim, mark_claim_paid

    result = _load_session(session_id)
    if not result:
        raise HTTPException(404, "Session not found")

    if body.action not in ("approve", "reject", "mark_paid"):
        raise HTTPException(400, f"Unknown action: {body.action}")

    if body.action == "reject" and not (body.reason and body.reason.strip()):
        raise HTTPException(400, "Rejection reason is required")

    actioned = []
    errors   = []
    # Claims where REST returned REST_FAILED → need Playwright batch
    playwright_needed: list[dict] = []  # [{cid, display_number, numeric_id}]

    unique_claim_ids = list(dict.fromkeys(body.claim_ids))
    label_map = {"approve": "Approved", "reject": "Rejected", "mark_paid": "Paid"}

    def _mark_rows(cid: str):
        label = label_map[body.action]
        for row in result.rows:
            cid_str = str(cid)
            row_matches = (
                row.keka_claim_id == cid_str
                or str(getattr(row, "keka_claim_number", "") or "") == cid_str
                or str(getattr(row, "claim_number", "") or "") == cid_str
            )
            if row_matches:
                row.keka_actioned = body.action
                # Also update row.status so summary cards reflect the Keka decision
                if body.action == "approve":
                    row.status = ExpenseStatus.APPROVED
                    row.remarks.append(f"[Keka {label}] via Expense Validator App")
                elif body.action == "reject":
                    row.status = ExpenseStatus.REJECTED
                    if body.reason:
                        row.remarks.append(f"[Keka {label}] {body.reason}")
                    else:
                        row.remarks.append(f"[Keka {label}] via Expense Validator App")
                elif body.action == "mark_paid":
                    ref = f" Ref: {body.reference_no}" if body.reference_no else ""
                    row.remarks.append(f"[Keka {label}] {body.payment_mode} on {body.payment_date or 'today'}.{ref}")

    for cid in unique_claim_ids:
        try:
            if body.action == "approve":
                approve_claim("", cid, company=body.company)
            elif body.action == "reject":
                reject_claim("", cid, reason=body.reason, company=body.company)
            elif body.action == "mark_paid":
                mark_claim_paid(
                    "", cid,
                    company=body.company,
                    payment_mode=body.payment_mode or "BankTransfer",
                    payment_date=body.payment_date,
                    reference_no=body.reference_no or "",
                )
            actioned.append(cid)
            _mark_rows(cid)

        except (RuntimeError, Exception) as e:
            err_str = str(e)
            import traceback as _tb
            log.warning("keka_action REST failed for claim %s: %s", cid, err_str[:120])

            # Look up display_number + numeric_id + employee_name from session rows
            display_number = cid
            numeric_id     = cid
            employee_name  = ""
            for row in result.rows:
                if (str(getattr(row, "keka_claim_id", "") or "") == str(cid)
                        or str(getattr(row, "keka_claim_number", "") or "") == str(cid)):
                    display_number = str(getattr(row, "keka_claim_number", "") or cid)
                    employee_name  = str(getattr(row, "employee_name", "") or "")
                    break

            if err_str.startswith("REST_FAILED|"):
                # Parse extra fields from the structured error
                parts = err_str.split("|", 5)
                if len(parts) > 4:
                    display_number = parts[4] or display_number
                if len(parts) > 3:
                    numeric_id = parts[3] or numeric_id

            # ALL failures → Playwright browser fallback (not just REST_FAILED)
            log.info("Queuing claim %s (display=%s, employee=%s) for Playwright browser fallback",
                     cid, display_number, employee_name)
            playwright_needed.append({
                "cid":           cid,
                "display_number": display_number,
                "numeric_id":    numeric_id,
                "employee_name": employee_name,
                "rest_error":    err_str[:200],
            })

    # Batch Playwright fallback for all REST-failed claims (single browser session)
    if playwright_needed:
        log.info("Playwright batch fallback for %d claim(s): %s",
                 len(playwright_needed), [x["display_number"] for x in playwright_needed])
        try:
            from services.keka_browser import batch_approve_claims_via_browser_ui
            claim_numbers    = [x["display_number"] for x in playwright_needed]
            numeric_ids_map  = {x["display_number"]: x["numeric_id"]    for x in playwright_needed}
            emp_names_map    = {x["display_number"]: x["employee_name"] for x in playwright_needed
                                if x.get("employee_name")}
            batch_res = _run_in_proactor_thread(
                batch_approve_claims_via_browser_ui,
                claim_numbers=claim_numbers,
                action=body.action,
                reason=body.reason or "",
                company=body.company,
                numeric_ids=numeric_ids_map,
                employee_names=emp_names_map,
            )
            for x in playwright_needed:
                dn = x["display_number"]
                if dn in (batch_res.get("actioned") or []):
                    actioned.append(x["cid"])
                    _mark_rows(x["cid"])
                else:
                    err = (batch_res.get("errors") or {}).get(dn, "Playwright UI approval failed")
                    errors.append({"claim_id": x["cid"], "error": err})
            if batch_res.get("discovered_api"):
                log.info("Discovered Keka approve API: %s %s",
                         batch_res["discovered_api"]["method"],
                         batch_res["discovered_api"]["url"])
        except Exception as e:
            import traceback
            log.error("Playwright batch failed: %s\n%s", e, traceback.format_exc())
            for x in playwright_needed:
                errors.append({"claim_id": x["cid"], "error": f"Playwright batch error: {str(e)[:120]}"})

    try:
        _save_session(session_id, result)
    except Exception as e:
        log.error("keka_action _save_session failed: %s", e)

    # Log the action
    try:
        from services.auth import verify_token
        from services.db import log_activity
        actor = "unknown"
        if authorization and authorization.lower().startswith("bearer "):
            p = verify_token(authorization[7:].strip())
            if p: actor = p.get("u", "unknown")
        if actioned:
            log_activity(actor, f"keka_{body.action}", "keka_claim",
                         ",".join(actioned[:5]),
                         {"count": len(actioned), "session": session_id,
                          "reason": body.reason or ""},
                         ip_address=request.client.host if request.client else "")
    except Exception:
        pass

    return {
        "actioned": actioned,
        "errors":   errors,
        "action":   body.action,
    }


@app.get("/keka/config")
def keka_config():
    """Return Keka OAuth2 + browser session configuration status."""
    from services.keka import KEKA_CLIENT_ID, KEKA_CLIENT_SECRET, KEKA_API_KEY, KEKA_COMPANY_NAME
    from services.keka_browser import KEKA_EMAIL, KEKA_PASSWORD, is_authenticated
    configured = bool(KEKA_CLIENT_ID and KEKA_CLIENT_SECRET and KEKA_API_KEY)
    missing = []
    if not KEKA_CLIENT_ID:     missing.append("KEKA_CLIENT_ID")
    if not KEKA_CLIENT_SECRET: missing.append("KEKA_CLIENT_SECRET")
    if not KEKA_API_KEY:       missing.append("KEKA_API_KEY")
    return {
        "configured":          configured,
        "company":             KEKA_COMPANY_NAME,
        "auth_method":         "oauth" if configured else "none",
        "client_id_set":       bool(KEKA_CLIENT_ID),
        "client_secret_set":   bool(KEKA_CLIENT_SECRET),
        "api_key_set":         bool(KEKA_API_KEY),
        "missing_credentials": missing,
        "bill_download_enabled": bool(
            KEKA_EMAIL and KEKA_PASSWORD and
            "your_keka" not in KEKA_EMAIL and
            "your_keka" not in KEKA_PASSWORD
        ),
        "bill_download_note": (
            "Bill download enabled via browser session"
            if (KEKA_EMAIL and "your_keka" not in KEKA_EMAIL)
            else "Set KEKA_EMAIL + KEKA_PASSWORD in .env to enable bill download"
        ),
        "session_active": is_authenticated(),
        "approve_endpoint_cached": _has_approve_endpoint_cache(),
    }


def _has_approve_endpoint_cache() -> bool:
    try:
        from services.keka_browser import _endpoint_cache_path
        import os as _os
        return any(_os.path.exists(_endpoint_cache_path(a)) for a in ("approve", "reject"))
    except Exception:
        return False


# ── Keka Login / 2FA Endpoints ────────────────────────────────────────────────

class KekaLoginVerifyRequest(BaseModel):
    otp: str
    token: str


@app.post("/keka/login/start")
async def keka_login_start():
    """
    Initiate Keka browser login.
    Runs Playwright in a fresh thread with its own ProactorEventLoop
    (ThreadPoolExecutor reuses threads that may have SelectorEventLoop,
    which causes sync_playwright to fail on Windows with Playwright 1.47+).
    """
    import asyncio
    import traceback
    import logging as _log
    from services.keka_browser import initiate_login

    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None, lambda: _run_in_proactor_thread(initiate_login)
        )
        return result
    except Exception as exc:
        _log.getLogger(__name__).error("keka_login_start error: %s\n%s", exc, traceback.format_exc())
        raise HTTPException(500, f"Login failed: {exc}")


@app.post("/keka/login/verify")
async def keka_login_verify(body: KekaLoginVerifyRequest):
    """Submit the OTP code received by email to complete 2FA login."""
    import asyncio
    import traceback
    import logging as _log
    from services.keka_browser import verify_otp

    try:
        otp   = body.otp.strip()
        token = body.token
        result = await asyncio.get_event_loop().run_in_executor(
            None, lambda: _run_in_proactor_thread(verify_otp, otp, token)
        )
        return result
    except Exception as exc:
        _log.getLogger(__name__).error("keka_login_verify error: %s\n%s", exc, traceback.format_exc())
        raise HTTPException(500, f"OTP verification failed: {exc}")


class KekaLoginCaptchaRequest(BaseModel):
    token: str
    captcha_text: str


@app.post("/keka/login/captcha")
async def keka_login_captcha(body: KekaLoginCaptchaRequest):
    """User has solved the captcha manually — complete the login."""
    import asyncio
    import traceback
    import logging as _log
    from services.keka_browser import submit_captcha_and_login

    try:
        result = await asyncio.get_event_loop().run_in_executor(
            None,
            lambda: _run_in_proactor_thread(
                submit_captcha_and_login, body.token, body.captcha_text.strip()
            ),
        )
        return result
    except Exception as exc:
        _log.getLogger(__name__).error("keka_login_captcha error: %s\n%s", exc, traceback.format_exc())
        raise HTTPException(500, f"Captcha login failed: {exc}")


@app.post("/keka/login/logout")
def keka_login_logout():
    """Clear the cached session (forces re-login on next sync)."""
    from services.keka_browser import clear_session_cache
    clear_session_cache()
    return {"status": "ok", "message": "Session cleared"}


@app.post("/keka/approve-endpoint/clear")
def keka_clear_approve_endpoint():
    """Clear cached approve + reject API endpoints (forces re-discovery on next action)."""
    from services.keka_browser import _endpoint_cache_path
    import os as _os
    cleared = []
    for act in ("approve", "reject"):
        try:
            p = _endpoint_cache_path(act)
            if _os.path.exists(p):
                _os.remove(p)
                cleared.append(act)
        except Exception:
            pass
    msg = f"Cleared: {', '.join(cleared)}" if cleared else "Nothing to clear"
    return {"status": "ok", "message": msg}


@app.get("/keka/debug/capture-report-payload")
async def keka_debug_capture_report_payload(
    from_date: str = "2026-05-01",
    to_date:   str = "2026-05-07",
):
    """
    Open browser, navigate to Expense Claim Report, set date range,
    click Run, intercept the actual API payload the SPA sends.
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"captured_requests": [], "captured_responses": []}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            # Capture every report-related POST body
            def _on_request(req):
                if "/expense/reports" in req.url or "/expense/claims" in req.url:
                    if req.method in ("POST", "PUT"):
                        try:
                            body = req.post_data
                            out["captured_requests"].append({
                                "url": req.url,
                                "method": req.method,
                                "body": body[:2000] if body else None,
                                "headers": dict(req.headers),
                            })
                        except Exception:
                            pass

            def _on_response(resp):
                if "/expense/reports" in resp.url and resp.status == 200:
                    try:
                        body_text = resp.text()[:3000]
                        out["captured_responses"].append({
                            "url": resp.url,
                            "status": resp.status,
                            "body_preview": body_text,
                        })
                    except Exception:
                        pass

            ctx.on("request", _on_request)
            ctx.on("response", _on_response)

            page = ctx.new_page()
            page.set_viewport_size({"width": 1500, "height": 950})

            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(2000)

            # Navigate to expense claim report
            page.evaluate("window.location.hash = '/org/expenses/reports/expenseclaim';")
            page.wait_for_timeout(7000)

            # Try clicking Run with whatever default date the page uses
            try:
                run_btn = page.wait_for_selector('button:has-text("Run"):visible', timeout=8000)
                run_btn.click()
                page.wait_for_load_state("networkidle", timeout=20000)
                page.wait_for_timeout(3000)
            except Exception as e:
                out["run_error"] = str(e)

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/screenshot-reports")
async def keka_debug_screenshot_reports():
    """Screenshot the org expense reports page so we can see the UI layout."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))
            page = ctx.new_page()
            page.set_viewport_size({"width": 1600, "height": 1200})

            # Step 1: Init SPA
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Step 2: Try the Reports page directly
            shots_dir = os.path.join(_BASE_TEMP, "keka_screenshots")
            os.makedirs(shots_dir, exist_ok=True)

            for hash_path, label in [
                ("/org/expenses/reports", "reports_main"),
                ("/org/expenses/reports/expense-claim", "reports_expense_claim"),
                ("/org/expenses/reports/expenseclaim", "reports_expenseclaim"),
                ("/org/expenses/reports/claims", "reports_claims"),
                ("/org/expenses/reports/expense", "reports_expense"),
            ]:
                try:
                    page.evaluate(f"window.location.hash = '{hash_path}';")
                    page.wait_for_timeout(8000)   # let Angular fully render
                    shot_path = os.path.join(shots_dir, f"{label}.png")
                    page.screenshot(path=shot_path, full_page=True)

                    # Get all visible text containing "bulk", "receipt", "download", etc.
                    interesting = page.evaluate("""
                        () => {
                            const all = Array.from(document.querySelectorAll('button, a, [role="tab"], [role="button"]'));
                            return all
                                .filter(e => e.offsetParent !== null)
                                .map(e => ({
                                    text: (e.innerText || '').trim().substring(0, 60),
                                    tag: e.tagName.toLowerCase(),
                                    href: e.getAttribute('href') || '',
                                    class: (e.className || '').substring(0, 60),
                                    role: e.getAttribute('role') || '',
                                }))
                                .filter(x => x.text);
                        }
                    """)
                    out[label] = {
                        "url": page.url,
                        "screenshot": shot_path,
                        "html_len": len(page.content()),
                        "all_elements_count": len(interesting),
                        "bulk_receipt_matches": [
                            e for e in interesting
                            if any(k in e["text"].lower() for k in ("bulk", "receipt", "download", "zip", "export"))
                        ],
                        "first_30_buttons": interesting[:30],
                    }
                except Exception as e:
                    out[label] = {"error": str(e)}

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/explore-reports")
async def keka_debug_explore_reports():
    """
    Navigate to Org → Expense → Reports page and capture full structure
    so we can find the Bulk Receipt download button.
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"steps": [], "buttons": [], "all_apis": [], "downloads": []}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            ctx.on("response", lambda r: api_calls.append({"url": r.url, "status": r.status})
                   if "keka.com" in r.url and ("/api/" in r.url or "/k/" in r.url) else None)

            page = ctx.new_page()
            out["steps"].append("Navigating to dashboard…")
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            api_calls.clear()
            out["steps"].append("Navigating to /org/expenses/reports…")
            try:
                page.evaluate("window.location.hash = '/org/expenses/reports';")
                page.wait_for_load_state("networkidle", timeout=20000)
                page.wait_for_timeout(5000)
            except Exception as e:
                out["nav_error"] = str(e)

            out["page_url"]  = page.url
            out["html_len"]  = len(page.content())
            out["api_calls"] = [c for c in api_calls if "keka.com" in c["url"]][:30]

            # Capture all clickable elements (links, buttons, tabs)
            try:
                elements = page.evaluate("""
                    () => Array.from(document.querySelectorAll('a, button, [role="tab"], [routerLink], [class*="tab"]'))
                        .filter(e => {
                            const txt = (e.innerText || '').trim();
                            return txt && txt.length > 0 && txt.length < 80;
                        })
                        .map(e => ({
                            tag:   e.tagName.toLowerCase(),
                            text:  (e.innerText || '').trim().substring(0, 60),
                            href:  e.getAttribute('href') || e.getAttribute('routerLink') || '',
                            class: (e.className || '').substring(0, 80),
                            id:    e.id || '',
                        }))
                        .slice(0, 80)
                """)
                out["clickable_elements"] = elements
                out["bulk_receipt_candidates"] = [
                    e for e in elements
                    if any(x in e["text"].lower() for x in ("bulk", "receipt", "download", "zip", "export"))
                ]
            except Exception as e:
                out["element_error"] = str(e)

            # Try to find expense claim report sub-tab
            try:
                tabs = page.evaluate("""
                    () => Array.from(document.querySelectorAll('[role="tab"], a[routerLink], li'))
                        .map(e => (e.innerText || '').trim())
                        .filter(t => t && t.length < 40)
                        .slice(0, 50)
                """)
                out["visible_tabs"] = list(set(tabs))[:40]
            except Exception:
                pass

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/me-past")
async def keka_debug_me_past():
    """Navigate to /me/expenses/past and capture EVERY blob/file URL the SPA fetches."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"all_files_urls": [], "all_apis": [], "expense_responses": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            seen_urls: set = set()

            def _on_resp(resp):
                u = resp.url
                if "/files/" in u and u not in seen_urls:
                    seen_urls.add(u)
                    out["all_files_urls"].append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")[:30]})
                if "blob.core" in u:
                    out["all_files_urls"].append({"url": u, "status": resp.status})
                if ("/api/" in u or "/k/" in u) and resp.status == 200 and "json" in resp.headers.get("content-type", "") and any(x in u.lower() for x in ("expense", "receipt", "claim")):
                    try:
                        key = u.split("keka.com")[1].split("?")[0]
                        if key not in out["expense_responses"]:
                            out["expense_responses"][key] = resp.text()[:3000]
                    except Exception:
                        pass

            ctx.on("response", _on_resp)
            page = ctx.new_page()

            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Try a series of expense pages
            for hash_url in [
                "/me/expenses/past",
                "/me/expenses/pending",
                "/me/expenses/advancerequests",
                "/org/expenses/expenseclaims",
                "/org/expenses/summary",
            ]:
                try:
                    page.evaluate(f"window.location.hash = '{hash_url}';")
                    page.wait_for_load_state("networkidle", timeout=20000)
                    page.wait_for_timeout(5000)
                    out[f"after_{hash_url[-15:]}"] = {
                        "final_url": page.url,
                        "html_len": len(page.content()),
                    }
                except Exception as e:
                    out[f"err_{hash_url[-15:]}"] = str(e)

            # Try to click on the first expense item if visible
            try:
                links = page.evaluate("""
                    () => Array.from(document.querySelectorAll('a, [routerLink], button'))
                        .map(e => ({
                            text: (e.innerText || '').substring(0, 50),
                            href: e.getAttribute('href') || e.getAttribute('routerLink') || '',
                        }))
                        .filter(x => x.href && (x.href.includes('expense') || x.href.includes('claim')))
                        .slice(0, 30)
                """)
                out["clickable_expense_links"] = links
            except Exception as e:
                out["link_error"] = str(e)

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/inbox-list")
async def keka_debug_inbox_list():
    """Probe inbox APIs to find expense-claim-specific URL patterns."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"api_calls": [], "files_urls": [], "blob_urls": [], "json_responses": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            json_resps: dict = {}

            def _on_resp(resp):
                u = resp.url
                if "keka.com" in u:
                    if "/api/" in u or "/k/" in u:
                        api_calls.append({"url": u, "status": resp.status})
                    if "blob.core" in u:
                        out["blob_urls"].append({"url": u, "status": resp.status})
                    if "/files/" in u:
                        out["files_urls"].append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")[:40]})
                    if resp.status == 200:
                        try:
                            ct = resp.headers.get("content-type", "")
                            if "json" in ct:
                                key = u.split("keka.com")[1].split("?")[0]
                                json_resps[key] = resp.text()[:3000]
                        except Exception:
                            pass

            ctx.on("response", _on_resp)
            page = ctx.new_page()

            # Step 1: Init SPA
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Step 2: Navigate to inbox
            api_calls_before = len(api_calls)
            try:
                page.evaluate("window.location.hash = '/inbox';")
                page.wait_for_load_state("networkidle", timeout=20000)
                page.wait_for_timeout(5000)
                out["after_inbox_url"] = page.url
                out["after_inbox_html_len"] = len(page.content())
                inbox_apis = api_calls[api_calls_before:]
                out["inbox_apis"] = [c for c in inbox_apis if any(x in c["url"].lower() for x in ("inbox", "expense", "claim", "approval"))][:30]
            except Exception as e:
                out["inbox_error"] = str(e)

            # Step 3: Probe inbox APIs directly
            inbox_paths = [
                "/k/dashboard/api/inbox/list",
                "/k/dashboard/api/inbox/pending",
                "/k/dashboard/api/inbox/items",
                "/k/dashboard/api/inbox/all",
                "/k/dashboard/api/inbox/expense-claim",
                "/k/default/api/inbox/list",
                "/k/default/api/inbox/pending",
                "/k/default/api/inbox/expense",
                "/k/default/api/me/inbox/expense",
                "/k/default/api/expense/claims/pending?pageNumber=1&pageSize=2",
            ]
            js = """
                async (paths) => {
                    const t = localStorage.getItem('access_token');
                    const r = {};
                    for (const p of paths) {
                        try {
                            const resp = await fetch(p, {credentials:'include', headers:{'Authorization':'Bearer '+t,'Accept':'application/json'}});
                            r[p] = {status: resp.status, body: (await resp.text()).substring(0, 2000)};
                        } catch(e) { r[p] = {error: e.toString()}; }
                    }
                    return r;
                }
            """
            try:
                out["inbox_probes"] = page.evaluate(js, inbox_paths)
            except Exception as e:
                out["probe_error"] = str(e)

            # Step 4: Capture relevant JSON responses
            out["json_responses"] = {k: v[:2000] for k, v in json_resps.items() if any(x in k for x in ("expense", "claim", "inbox", "approval"))}

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/inbox-claim/{claim_id}")
async def keka_debug_inbox_claim(claim_id: str):
    """Navigate to inbox action page for an expense claim and capture all API + file URLs."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"api_calls": [], "files_urls": [], "blob_urls": [], "json_responses": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            json_resps: dict = {}

            def _on_resp(resp):
                u = resp.url
                if "keka.com" in u:
                    if "/api/" in u or "/k/" in u:
                        api_calls.append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")[:30]})
                    if "blob.core" in u:
                        out["blob_urls"].append({"url": u, "status": resp.status})
                    if "/files/" in u:
                        out["files_urls"].append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")[:30]})
                    if any(x in u.lower() for x in ("expense", "claim", "receipt")) and resp.status == 200:
                        try:
                            ct = resp.headers.get("content-type", "")
                            if "json" in ct:
                                key = u.split("keka.com")[1].split("?")[0]
                                if key not in json_resps:
                                    json_resps[key] = resp.text()[:3000]
                        except Exception:
                            pass

            ctx.on("response", _on_resp)
            page = ctx.new_page()

            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            api_calls.clear()

            # Try multiple inbox URL patterns with the UUID
            for hash_path in [
                f"/inbox/action/expense-claim/{claim_id}/details",
                f"/inbox/action/expenseclaim/{claim_id}/details",
                f"/inbox/action/expense/{claim_id}/details",
                f"/inbox/action/expense-payout-claim/{claim_id}/details",
            ]:
                api_calls_before = len(api_calls)
                try:
                    page.evaluate(f"window.location.hash = '{hash_path}';")
                    page.wait_for_load_state("networkidle", timeout=15000)
                    page.wait_for_timeout(4000)
                    new_apis = api_calls[api_calls_before:]
                    expense_apis = [c for c in new_apis if any(x in c["url"].lower() for x in ("expense", "claim", "attach", "receipt", "files"))]
                    out[f"hash_{hash_path[-30:]}"] = {
                        "final_url": page.url,
                        "html_len": len(page.content()),
                        "expense_apis_count": len(expense_apis),
                        "expense_apis": [{"url": c["url"][:200], "status": c["status"]} for c in expense_apis[:15]],
                    }
                except Exception as e:
                    out[f"hash_error_{hash_path[-30:]}"] = str(e)

            out["json_responses"] = {k: v[:1500] for k, v in json_resps.items() if "expense" in k.lower() or "claim" in k.lower()}

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/spa-claims-full")
async def keka_debug_spa_claims_full():
    """
    Use SPA API to fetch pending claims (numeric IDs) and find attachment locations.
    Tests downloading actual receipts via various URL patterns.
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))
            page = ctx.new_page()
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Step 1: Get pending claims (numeric IDs)
            js1 = """
                async () => {
                    const t = localStorage.getItem('access_token');
                    const r = await fetch('/k/default/api/expense/claims/pending', {
                        credentials:'include',
                        headers:{'Authorization':'Bearer '+t,'Accept':'application/json'}
                    });
                    return {status: r.status, body: await r.text()};
                }
            """
            try:
                claims_resp = page.evaluate(js1)
                if claims_resp["status"] == 200:
                    claims_data = json.loads(claims_resp["body"])
                    claims_list = claims_data.get("data", [])
                    if claims_list:
                        first_claim = claims_list[0]
                        out["first_claim_summary"] = {
                            "id": first_claim.get("id"),
                            "claimNumber": first_claim.get("claimNumber"),
                            "title": first_claim.get("title"),
                            "expenseIds": first_claim.get("expenseIds"),
                            "employeeName": first_claim.get("employeeName"),
                        }
                        out["total_pending"] = len(claims_list)

                        claim_num_id = first_claim.get("id")
                        expense_ids = first_claim.get("expenseIds", [])
                        first_expense_id = expense_ids[0] if expense_ids else None
                    else:
                        return {"error": "No pending claims"}
                else:
                    return {"error": f"Pending API: {claims_resp['status']}"}
            except Exception as e:
                return {"error": f"Step 1 failed: {e}"}

            # Step 2: Probe claim detail endpoints with NUMERIC ID
            paths = [
                f"/k/default/api/expense/claims/{claim_num_id}",
                f"/k/default/api/expense/claims/{claim_num_id}/details",
                f"/k/default/api/expense/claims/{claim_num_id}/expenses",
                f"/k/default/api/expense/expenseclaims/{claim_num_id}",
                f"/k/default/api/expense/expenseclaims/{claim_num_id}/expenses",
                f"/k/default/api/expense/claim/{claim_num_id}",
            ]
            if first_expense_id:
                paths += [
                    f"/k/default/api/expense/expenses/{first_expense_id}",
                    f"/k/default/api/expense/expenses/{first_expense_id}/receipts",
                    f"/k/default/api/me/expenses/expenses/{first_expense_id}",
                ]

            js2 = """
                async (paths) => {
                    const t = localStorage.getItem('access_token');
                    const r = {};
                    for (const p of paths) {
                        try {
                            const resp = await fetch(p, {
                                credentials:'include',
                                headers:{'Authorization':'Bearer '+t,'Accept':'application/json'}
                            });
                            r[p] = {status: resp.status, body: (await resp.text()).substring(0, 2500)};
                        } catch(e) { r[p] = {error: e.toString()}; }
                    }
                    return r;
                }
            """
            try:
                out["claim_detail_probes"] = page.evaluate(js2, paths)
            except Exception as e:
                out["probe_error"] = str(e)

            # Step 3: Get a real receipt location from /me/expenses/bills/pending
            js3 = """
                async () => {
                    const t = localStorage.getItem('access_token');
                    const r = await fetch('/k/default/api/me/expenses/bills/pending', {
                        credentials:'include',
                        headers:{'Authorization':'Bearer '+t,'Accept':'application/json'}
                    });
                    return {status: r.status, body: await r.text()};
                }
            """
            real_location = None
            real_filename = None
            try:
                bills_resp = page.evaluate(js3)
                if bills_resp["status"] == 200:
                    bills_data = json.loads(bills_resp["body"])
                    for bill in bills_data.get("data", []):
                        for receipt in bill.get("expenseReceipts", []):
                            loc = receipt.get("location", "")
                            if loc:
                                real_location = loc
                                real_filename = receipt.get("name", "test.jpg")
                                break
                        if real_location:
                            break
                    out["real_location"] = real_location
                    out["real_filename"] = real_filename
            except Exception as e:
                out["bills_error"] = str(e)

            # Step 4: Try downloading the real receipt with various URL patterns
            tenant_id = "48e83537-f402-4280-b26c-dd8016b09944"
            if real_location:
                test_urls = [
                    f"https://{co}.keka.com/files/{real_location}",
                    f"https://{co}.keka.com/files/{tenant_id}/{real_location}",
                    f"https://{co}.keka.com/files/{tenant_id}/original/{real_location}",
                    f"https://{co}.keka.com/files/original/{real_location}",
                    f"https://{co}.keka.com/api/v1/files/{real_location}",
                    f"https://{co}.keka.com/k/default/api/files/{real_location}",
                ]
                js4 = """
                    async (urls) => {
                        const t = localStorage.getItem('access_token');
                        const r = {};
                        for (const u of urls) {
                            try {
                                const resp = await fetch(u, {
                                    credentials:'include',
                                    headers:{'Authorization':'Bearer '+t}
                                });
                                const blob = await resp.blob();
                                r[u] = {
                                    status: resp.status,
                                    ct: resp.headers.get('content-type'),
                                    len: blob.size,
                                };
                            } catch(e) { r[u] = {error: e.toString()}; }
                        }
                        return r;
                    }
                """
                try:
                    out["real_download_tests"] = page.evaluate(js4, test_urls)
                except Exception as e:
                    out["download_error"] = str(e)

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/find-receipts/{claim_id}")
async def keka_debug_find_receipts(claim_id: str):
    """Probe internal SPA endpoints to find receipt locations for a claim."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"claim_id": claim_id, "responses": {}, "download_tests": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))
            page = ctx.new_page()
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Probe candidate endpoints
            paths = [
                f"/k/default/api/expense/claims/{claim_id}",
                f"/k/default/api/expense/claims/{claim_id}/expenses",
                f"/k/default/api/expense/expenses?claimId={claim_id}",
                f"/k/default/api/me/expenses/claims/{claim_id}",
                f"/k/default/api/me/expenses/claims/{claim_id}/expenses",
            ]

            js = """
                async (paths) => {
                    const token = localStorage.getItem('access_token');
                    const r = {};
                    for (const p of paths) {
                        try {
                            const resp = await fetch(p, {
                                credentials:'include',
                                headers: {'Authorization': 'Bearer ' + token, 'Accept': 'application/json'}
                            });
                            r[p] = {status: resp.status, body: (await resp.text()).substring(0, 2500)};
                        } catch (e) { r[p] = {error: e.toString()}; }
                    }
                    return r;
                }
            """
            try:
                out["responses"] = page.evaluate(js, paths)
            except Exception as e:
                out["probe_error"] = str(e)

            # Get fresh tokens
            try:
                fresh = page.evaluate("""
                    () => ({
                        token: localStorage.getItem('access_token') || '',
                        sas: localStorage.getItem('sasTokenDetails') || '',
                    })
                """)
                out["fresh_token"] = fresh.get("token", "")[:60]
                sas = json.loads(fresh.get("sas", "{}"))
                out["sas_url"] = sas.get("sasUrl", "")
                out["sas_expires"] = sas.get("sasExpiresOn", "")
            except Exception:
                pass

            # Test downloading using known location pattern from /me/expenses/bills/pending
            tenant_id = "48e83537-f402-4280-b26c-dd8016b09944"
            test_location = "expensereceipts/fdedbd8112824e9386fc2b7fcc20b541.jpg"
            sas_q = out.get("sas_url", "")

            test_urls = [
                f"https://{co}.keka.com/files/{tenant_id}/original/{test_location}",
                f"https://{co}.keka.com/files/{tenant_id}/{test_location}",
                f"https://{co}.keka.com/files/{tenant_id}/orig/{test_location}",
                f"https://stkekahrprodcin02.blob.core.windows.net/hrprodcin02/{test_location}{sas_q}",
                f"https://stkekahrprodcin02.blob.core.windows.net/{tenant_id}/{test_location}{sas_q}",
            ]

            js2 = """
                async (urls) => {
                    const r = {};
                    for (const u of urls) {
                        try {
                            const resp = await fetch(u, {credentials:'include'});
                            r[u.substring(0,120)] = {
                                status: resp.status,
                                ct: resp.headers.get('content-type'),
                                len: (await resp.blob()).size,
                            };
                        } catch (e) { r[u.substring(0,120)] = {error: e.toString()}; }
                    }
                    return r;
                }
            """
            try:
                out["download_tests"] = page.evaluate(js2, test_urls)
            except Exception as e:
                out["download_error"] = str(e)

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/org-claims")
async def keka_debug_org_claims():
    """Navigate to /org/expenses/expenseclaims and capture API + attachment URLs."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"api_calls": [], "blob_urls": [], "files_urls": [], "expense_responses": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            expense_responses = {}

            def _on_resp(resp):
                u = resp.url
                if "keka.com" in u:
                    if "/api/" in u or "/k/" in u:
                        api_calls.append({"url": u, "status": resp.status})
                    if "blob.core" in u:
                        out["blob_urls"].append({"url": u, "status": resp.status})
                    if "/files/" in u and resp.status == 200:
                        out["files_urls"].append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")})
                    # Capture expense API JSON responses
                    if any(x in u.lower() for x in ("/expense", "/claim")):
                        try:
                            if resp.status == 200 and "json" in resp.headers.get("content-type", ""):
                                key = u.split("keka.com")[1].split("?")[0]
                                expense_responses[key] = resp.text()[:1500]
                        except Exception:
                            pass

            ctx.on("response", _on_resp)
            page = ctx.new_page()

            # Step 1: Load home so SPA initializes
            page.goto(f"https://{co}.keka.com/#/home/dashboard", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(3000)

            # Step 2: Navigate to expense claims page (org-level)
            api_calls.clear()
            try:
                page.evaluate("window.location.hash = '/org/expenses/expenseclaims';")
                page.wait_for_load_state("networkidle", timeout=20000)
                page.wait_for_timeout(5000)
                out["org_claims_url"] = page.url
                out["org_claims_html_len"] = len(page.content())
                out["org_claims_apis"] = [c for c in api_calls if any(x in c["url"].lower() for x in ("expense", "claim"))]
            except Exception as e:
                out["org_claims_error"] = str(e)

            # Step 3: Try to navigate to a specific claim
            api_calls.clear()
            try:
                page.evaluate("window.location.hash = '/org/expenses/expenseclaims/349dfe22-81d3-4487-82b2-62b01ddb21ab';")
                page.wait_for_load_state("networkidle", timeout=15000)
                page.wait_for_timeout(5000)
                out["claim_detail_url"] = page.url
                out["claim_detail_apis"] = [c for c in api_calls if any(x in c["url"].lower() for x in ("expense", "claim", "attach", "files"))]
            except Exception as e:
                out["claim_detail_error"] = str(e)

            # Step 4: Try /me/expenses too
            api_calls.clear()
            try:
                page.evaluate("window.location.hash = '/me/expenses';")
                page.wait_for_load_state("networkidle", timeout=15000)
                page.wait_for_timeout(3000)
                out["me_expenses_apis"] = [c for c in api_calls if any(x in c["url"].lower() for x in ("expense", "claim"))]
            except Exception as e:
                out["me_expenses_error"] = str(e)

            out["expense_responses"] = expense_responses

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/inbox")
async def keka_debug_inbox():
    """Navigate to inbox/dashboard, find all expense-related routes and API calls."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            ctx.on("response", lambda r: api_calls.append({"url": r.url, "status": r.status}) if "keka.com" in r.url else None)

            page = ctx.new_page()
            page.goto(f"https://{co}.keka.com/home", wait_until="domcontentloaded", timeout=30000)
            page.wait_for_load_state("networkidle", timeout=30000)
            page.wait_for_timeout(4000)

            # Look at sidebar/menu links in the rendered HTML
            try:
                links = page.evaluate("""
                    () => {
                        const all = Array.from(document.querySelectorAll('a[href], [routerLink]'));
                        return all.map(a => a.getAttribute('href') || a.getAttribute('routerLink')).filter(Boolean);
                    }
                """)
                out["all_links"] = list(set(links))[:60]
                out["expense_links"] = [l for l in set(links) if 'expense' in l.lower() or 'claim' in l.lower() or 'inbox' in l.lower()]
            except Exception as e:
                out["link_error"] = str(e)

            # Look at the SPA's router routes by triggering them via JS
            try:
                # Find all distinct API paths under /k/ or /api/
                api_paths = set()
                for c in api_calls:
                    u = c["url"]
                    if "/k/" in u:
                        # Extract path after host
                        path = u.split("keka.com")[1].split("?")[0]
                        api_paths.add(path)
                out["all_api_paths"] = sorted(api_paths)[:80]
            except Exception as e:
                out["api_path_error"] = str(e)

            # Try inbox-related routes
            inbox_results = []
            for hash_path in [
                "/#/me/inbox/pending",
                "/#/me/inbox",
                "/#/me/myinbox",
                "/#/me/inbox/all",
                "/#/me/inbox/expense",
                "/#/myinbox",
                "/#/inbox",
                "/#/me/expense/myexpenses",
                "/#/me/expense/myclaims",
            ]:
                api_calls_count = len(api_calls)
                try:
                    page.evaluate(f"window.location.hash = '{hash_path[2:]}';")
                    page.wait_for_load_state("networkidle", timeout=10000)
                    page.wait_for_timeout(2000)
                    new_calls = api_calls[api_calls_count:]
                    expense_calls = [c for c in new_calls if any(x in c["url"].lower() for x in ("expense", "claim"))]
                    inbox_results.append({
                        "hash": hash_path,
                        "final_url": page.url,
                        "new_api_count": len(new_calls),
                        "expense_calls": [c["url"] for c in expense_calls][:5],
                    })
                except Exception as e:
                    inbox_results.append({"hash": hash_path, "error": str(e)})

            out["inbox_navigation"] = inbox_results

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception:
                pass

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/expense-page")
async def keka_debug_expense_page():
    """
    Navigate to the expense claims list page and capture ALL API calls.
    Reveals the actual internal endpoint URLs the SPA uses for expense data.
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"all_apis": [], "blob_urls": [], "files_urls": [], "expense_apis": []}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            api_calls = []
            def _on_resp(resp):
                u = resp.url
                if "keka.com" in u:
                    if "/api/" in u or "/k/" in u:
                        api_calls.append({"url": u, "status": resp.status, "ct": resp.headers.get("content-type", "")[:40]})
                    if "blob.core" in u:
                        out["blob_urls"].append({"url": u, "status": resp.status})
                    if "/files/" in u and resp.status == 200:
                        out["files_urls"].append({"url": u, "status": resp.status})

            ctx.on("response", _on_resp)
            page = ctx.new_page()

            # Try multiple expense-related URLs (Angular uses hash routing!)
            for path in [
                "/#/home/dashboard",
                "/#/home/expense",
                "/#/home/expense/myexpenses",
                "/#/home/expense/myclaims",
                "/#/home/expense/claims",
                "/#/me/expense",
                "/#/me/expense/myexpenses",
            ]:
                api_calls.clear()
                try:
                    page.goto(f"https://{co}.keka.com{path}", wait_until="domcontentloaded", timeout=30000)
                    page.wait_for_load_state("networkidle", timeout=20000)
                    page.wait_for_timeout(3000)
                    expense_only = [c for c in api_calls if any(x in c["url"].lower() for x in ("expense", "claim", "receipt"))]
                    out["expense_apis"].append({
                        "path": path,
                        "final_url": page.url,
                        "all_count": len(api_calls),
                        "expense_count": len(expense_only),
                        "expense_apis": expense_only[:30],
                        "html_len": len(page.content()),
                    })
                except Exception as e:
                    out["expense_apis"].append({"path": path, "error": str(e)})

            # Save fresh session
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
                    out["session_refreshed"] = True
            except Exception as e:
                out["session_save_error"] = str(e)

            browser.close()
        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/spa-api/{claim_id}")
async def keka_debug_spa_api(claim_id: str):
    """
    Use fresh access_token from localStorage to probe internal Keka SPA APIs.
    Tries /k/expense/api/... patterns to find the claim attachment endpoint.
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {"claim_id": claim_id, "probes": {}}

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))
            page = ctx.new_page()
            try:
                page.goto(f"https://{co}.keka.com/home", wait_until="domcontentloaded", timeout=30000)
                page.wait_for_load_state("networkidle", timeout=30000)
                page.wait_for_timeout(3000)
            except Exception as e:
                out["nav_error"] = str(e)

            # Probe all candidate endpoints from within the page
            paths = [
                f"/k/expense/api/claims/{claim_id}",
                f"/k/expense/api/v1/claims/{claim_id}",
                f"/k/expense/api/claims/{claim_id}/expenses",
                f"/k/expense/api/claims/{claim_id}/attachments",
                f"/k/expense/api/me/claims/{claim_id}",
                f"/k/expense/api/expense/claims/{claim_id}",
                f"/k/dashboard/api/expense/claims/{claim_id}",
                f"/k/dashboard/api/me/expenses/claims/{claim_id}",
                f"/k/expense/api/expenseclaims/{claim_id}",
                f"/k/expense/api/v1/expenseclaims/{claim_id}",
                f"/k/expense/api/v1/expense/claims/{claim_id}",
                # Also try without v1
                f"/k/expense/api/claim/{claim_id}",
            ]

            js = """
                async (paths) => {
                    const token = localStorage.getItem('access_token');
                    const results = {};
                    for (const p of paths) {
                        try {
                            const r = await fetch(p, {
                                credentials: 'include',
                                headers: { 'Authorization': 'Bearer ' + token, 'Accept': 'application/json' }
                            });
                            results[p] = { status: r.status, body: (await r.text()).substring(0, 800) };
                        } catch (e) {
                            results[p] = { error: e.toString() };
                        }
                    }
                    return results;
                }
            """
            try:
                results = page.evaluate(js, paths)
                out["probes"] = results
            except Exception as e:
                out["probe_error"] = str(e)

            # Also save fresh tokens
            try:
                fresh = page.evaluate("""
                    () => ({
                        token: localStorage.getItem('access_token') || '',
                        sas: localStorage.getItem('sasTokenDetails') || '',
                    })
                """)
                out["fresh_token_prefix"] = fresh.get("token", "")[:60]
                out["fresh_sas"] = json.loads(fresh.get("sas", "{}")).get("sasUrl", "")[:80]

                # Save refreshed session
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
            except Exception as e:
                out["token_error"] = str(e)

            browser.close()

        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/browser-intercept/{claim_id}")
async def keka_debug_browser_intercept(claim_id: str):
    """
    Debug: launch headless browser with saved session, navigate to claim,
    intercept ALL network traffic, identify attachment download URL pattern.
    Returns:
      - Fresh access_token + SAS from localStorage after SPA init
      - All blob.core.windows.net + /files/ URLs the SPA actually requests
      - All API responses containing attachment URLs
    """
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import (
        is_authenticated, _new_authenticated_context, _save_session_to_disk,
        _session_cache, KEKA_COMPANY_NAME, _STEALTH_SCRIPT,
    )
    if not is_authenticated():
        raise HTTPException(401, "No active Keka session")

    def _run():
        from playwright.sync_api import sync_playwright
        co = KEKA_COMPANY_NAME
        out: dict = {
            "claim_id": claim_id,
            "blob_urls":     [],
            "files_urls":    [],
            "all_responses": [],
            "fresh_token":   "",
            "fresh_sas":     "",
            "page_html_len": 0,
        }

        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)

            # Block bot-detection so Angular renders
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            def _on_resp(resp):
                url = resp.url
                if "blob.core.windows.net" in url:
                    out["blob_urls"].append({"url": url, "status": resp.status,
                                              "ct": resp.headers.get("content-type", "")})
                if "/files/" in url and "keka.com" in url:
                    out["files_urls"].append({"url": url, "status": resp.status,
                                               "ct": resp.headers.get("content-type", ""),
                                               "loc": resp.headers.get("location", "")})
                # Track all API JSON responses (might contain download URLs)
                if "/api/" in url and "keka.com" in url and resp.status == 200:
                    ct = resp.headers.get("content-type", "")
                    if "json" in ct:
                        out["all_responses"].append({"url": url[:120], "status": resp.status})

            ctx.on("response", _on_resp)

            page = ctx.new_page()
            try:
                page.goto(f"https://{co}.keka.com/home", wait_until="domcontentloaded", timeout=30000)
                page.wait_for_load_state("networkidle", timeout=30000)
                page.wait_for_timeout(4000)
            except Exception as e:
                out["home_error"] = str(e)

            # Get fresh tokens after SPA init
            try:
                fresh = page.evaluate("""
                    () => ({
                        access_token: localStorage.getItem('access_token') || '',
                        sasTokenDetails: localStorage.getItem('sasTokenDetails') || '',
                    })
                """)
                out["fresh_token"] = fresh.get("access_token", "")[:80]
                try:
                    sas = json.loads(fresh.get("sasTokenDetails", "{}"))
                    out["fresh_sas"]   = sas.get("sasUrl", "")
                    out["sas_expires"] = sas.get("sasExpiresOn", "")
                except Exception:
                    pass
            except Exception as e:
                out["token_read_error"] = str(e)

            # Navigate to claim page
            try:
                page.goto(f"https://{co}.keka.com/home/expense/claim/{claim_id}",
                          wait_until="domcontentloaded", timeout=30000)
                page.wait_for_load_state("networkidle", timeout=20000)
                page.wait_for_timeout(5000)
            except Exception as e:
                out["claim_nav_error"] = str(e)

            out["page_url"]      = page.url
            out["page_html_len"] = len(page.content())
            out["page_html_sample"] = page.content()[:1500]

            # Save the fresh session state (with refreshed tokens)
            try:
                fresh_storage = ctx.storage_state()
                cached = _session_cache.get(co)
                if cached:
                    cached["storage_state"] = fresh_storage
                    _save_session_to_disk()
                    out["session_refreshed"] = True
            except Exception as e:
                out["session_save_error"] = str(e)

            browser.close()

        return out

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        return await loop.run_in_executor(pool, _run)


@app.get("/keka/debug/download-test")
async def keka_debug_download_test():
    """
    Debug: fetch first claim, get its attachment IDs, then probe all download URL patterns.
    Shows exactly which URL/auth combination works for downloading attachments.
    """
    from services.keka import get_keka_token, fetch_expense_claims
    from services.keka_browser import (
        _get_session_cookies, _cookies_to_header, get_spa_access_token,
        get_spa_sas_details, KEKA_COMPANY_NAME, _UA, is_authenticated,
    )
    import requests as _req

    result: dict = {}

    # 1. Get a real attachment ID via OAuth list API
    try:
        token = get_keka_token()
        co = KEKA_COMPANY_NAME
        base = f"https://{co}.keka.com"

        r = _req.get(
            f"{base}/api/v1/expense/claims",
            headers={"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": _UA},
            params={"pageNumber": 1, "pageSize": 5, "lastModified": "2025-01-01T00:00:00Z"},
            timeout=20,
        )
        d = r.json()
        inner = d.get("data", d) if isinstance(d, dict) else d
        claims = inner.get("data", inner) if isinstance(inner, dict) else inner

        att_samples = []
        for c in (claims if isinstance(claims, list) else []):
            for exp in (c.get("expenses") or []):
                for att in (exp.get("attachments") or []):
                    if att.get("id"):
                        att_samples.append({
                            "claim_id": c.get("id"),
                            "att_id": att["id"],
                            "att_name": att.get("name", ""),
                        })
            if att_samples:
                break

        result["att_samples"] = att_samples[:3]

        if not att_samples:
            return {"error": "No attachment IDs found in first 5 claims", "claims_sample": str(claims)[:500]}

        att = att_samples[0]
        att_id = att["att_id"]
        result["testing_att_id"] = att_id
        result["testing_att_name"] = att["att_name"]

    except Exception as e:
        return {"error": f"Could not get attachment IDs: {e}"}

    # 2. Probe download patterns
    try:
        cookies = _get_session_cookies(co)
        cookie_str = _cookies_to_header(cookies)
        session_active = True
    except Exception:
        cookie_str = ""
        session_active = False

    spa_token = get_spa_access_token(co) or ""
    sas_details = get_spa_sas_details(co) or {}
    sas_query = sas_details.get("sasUrl", "")

    result["session_active"] = session_active
    result["spa_token_available"] = bool(spa_token)
    result["sas_available"] = bool(sas_query)
    result["sas_expires"] = sas_details.get("sasExpiresOn", "")

    probes = {}
    probe_urls = [
        (f"{base}/files/{att_id}", {"Cookie": cookie_str}),
        (f"{base}/files/{att_id}", {"Authorization": f"Bearer {spa_token}", "Cookie": cookie_str}),
        (f"{base}/files/{att_id}", {"Authorization": f"Bearer {token}"}),
        (f"{base}/api/v1/expense/attachments/{att_id}", {"Authorization": f"Bearer {token}"}),
        (f"{base}/api/v1/expense/attachments/{att_id}/download", {"Authorization": f"Bearer {token}"}),
    ]
    if sas_query:
        for blob_prefix in [
            f"https://stkekahrprodcin02.blob.core.windows.net/{att_id}",
            f"https://stkekahrprodcin02.blob.core.windows.net/hr/{att_id}",
            f"https://stkekahrprodcin02.blob.core.windows.net/hrprodcin02/{att_id}",
        ]:
            probe_urls.append((f"{blob_prefix}{sas_query}", {}))

    for url, extra_hdrs in probe_urls:
        key = url.split("?")[0].replace(base + "/", "")[-60:]
        hdrs = {"User-Agent": _UA, "Accept": "*/*", "Referer": f"{base}/"}
        hdrs.update(extra_hdrs)
        try:
            r = _req.get(url, headers=hdrs, timeout=10, allow_redirects=False)
            probes[key] = {
                "status": r.status_code,
                "content_type": r.headers.get("Content-Type", ""),
                "content_length": len(r.content),
                "location": r.headers.get("Location", "")[:150],
            }
            if r.status_code in (301, 302, 307, 308):
                loc = r.headers.get("Location", "")
                probes[key]["redirect_host"] = loc.split("/")[2] if loc.startswith("http") else ""
        except Exception as e:
            probes[key] = {"error": str(e)}

    result["probes"] = probes
    return result


@app.get("/keka/debug/attachment/{att_id}")
async def keka_debug_attachment(att_id: str):
    """Debug: probe every plausible download URL for a Keka attachment ID."""
    from services.keka import get_keka_token
    from services.keka_browser import _get_session_cookies, _cookies_to_header, KEKA_COMPANY_NAME, _UA, is_authenticated
    import requests as _req

    token = get_keka_token()
    co = KEKA_COMPANY_NAME
    base = f"https://{co}.keka.com"

    oauth_hdrs = {"Authorization": f"Bearer {token}", "Accept": "*/*", "User-Agent": _UA}

    cookie_hdrs = {}
    if is_authenticated():
        cookie_str = _cookies_to_header(_get_session_cookies(co))
        cookie_hdrs = {"Cookie": cookie_str, "Accept": "*/*", "User-Agent": _UA}

    result = {"att_id": att_id, "oauth": {}, "cookie": {}}

    urls = [
        f"{base}/api/v1/expense/attachments/{att_id}",
        f"{base}/api/v1/expense/attachments/{att_id}/download",
        f"{base}/api/v1/payloads/filedownload/{att_id}",
        f"{base}/api/v1/payloads/filedownload?identifier={att_id}",
        f"{base}/payloads/filedownload/{att_id}",
        f"{base}/filemanager/download/{att_id}",
        f"{base}/files/{att_id}",
        f"https://app.keka.com/api/v1/expense/attachments/{att_id}/download",
        f"https://files.keka.com/{att_id}",
    ]

    for url in urls:
        key = url.replace(f"{base}/", "").replace("https://", "")
        for label, hdrs in [("oauth", oauth_hdrs), ("cookie", cookie_hdrs)]:
            if not hdrs:
                continue
            try:
                r = _req.get(url, headers=hdrs, timeout=10, allow_redirects=False)
                result[label][key] = {
                    "status": r.status_code,
                    "location": r.headers.get("Location", ""),
                    "content_type": r.headers.get("Content-Type", ""),
                    "content_length": len(r.content),
                    "body_sample": r.text[:300] if r.status_code not in (200,) else f"<{len(r.content)} bytes>",
                }
            except Exception as e:
                result[label][key] = {"error": str(e)}

    return result


@app.get("/keka/debug/rawclaim")
async def keka_debug_rawclaim():
    """Debug: show raw structure of first claim from list API."""
    from services.keka import get_keka_token
    import requests as _req

    token = get_keka_token()
    co = os.environ.get("KEKA_COMPANY_NAME", "omniainformation")
    base = f"https://{co}.keka.com/api/v1"
    hdrs = {"Authorization": f"Bearer {token}", "Accept": "application/json", "User-Agent": "Mozilla/5.0"}

    r = _req.get(f"{base}/expense/claims", headers=hdrs, params={"pageNumber": 1, "pageSize": 1, "lastModified": "2025-04-01T00:00:00Z"}, timeout=20)
    return {"status": r.status_code, "body": r.json()}


@app.get("/keka/debug/oauth/{claim_id}")
async def keka_debug_oauth(claim_id: str):
    """Debug: check what fetch_claim_details returns via OAuth token."""
    from services.keka import get_keka_token, fetch_claim_details, fetch_claim_attachments_info
    import requests as _req

    try:
        token = get_keka_token()
    except Exception as e:
        raise HTTPException(500, f"Token error: {e}")

    co = os.environ.get("KEKA_COMPANY_NAME", "omniainformation")
    base = f"https://{co}.keka.com/api/v1"
    hdrs = {
        "Authorization": f"Bearer {token}",
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json",
    }

    result = {"claim_id": claim_id, "endpoints": {}}

    # Try every plausible endpoint
    for path in [
        f"{base}/expense/claims/{claim_id}",
        f"{base}/expense/claims/{claim_id}/attachments",
        f"{base}/expense/claims/{claim_id}/receipts",
        f"{base}/expense/expenses?claimId={claim_id}",
    ]:
        key = path.split("/v1/")[1]
        try:
            r = _req.get(path, headers=hdrs, timeout=15)
            result["endpoints"][key] = {
                "status": r.status_code,
                "body": r.text[:1000],
            }
        except Exception as e:
            result["endpoints"][key] = {"error": str(e)}

    # Also try fetch_claim_details helper
    try:
        detail = fetch_claim_details(token, claim_id)
        result["fetch_claim_details"] = str(detail)[:2000]
    except Exception as e:
        result["fetch_claim_details_error"] = str(e)

    # Attachments helper
    try:
        atts = fetch_claim_attachments_info(token, claim_id)
        result["fetch_claim_attachments_info"] = atts[:5]
    except Exception as e:
        result["fetch_claim_attachments_info_error"] = str(e)

    return result


@app.get("/keka/debug/claim/{claim_id}")
async def keka_debug_claim(claim_id: str):
    """Debug: navigate to claim page, capture ALL API calls the SPA makes, find attachment URLs."""
    from concurrent.futures import ThreadPoolExecutor
    from services.keka_browser import is_authenticated, _get_session_cookies, KEKA_COMPANY_NAME, _UA

    if not is_authenticated():
        raise HTTPException(401, "Not logged in to Keka")

    def _debug_claim():
        from playwright.sync_api import sync_playwright

        co = KEKA_COMPANY_NAME
        cookies = _get_session_cookies(co)
        result = {"claim_id": claim_id, "all_api_calls": [], "file_urls": [], "page_url": "", "page_html_sample": ""}

        from services.keka_browser import _new_authenticated_context
        with sync_playwright() as p:
            browser, ctx = _new_authenticated_context(p, co)

            # Block bot-detection script
            ctx.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            all_reqs = []
            all_resps = []

            ctx.on("request",  lambda r: all_reqs.append({"url": r.url, "method": r.method}))
            ctx.on("response", lambda r: all_resps.append({"url": r.url, "status": r.status,
                               "ct": r.headers.get("content-type", "")}))

            page = ctx.new_page()

            # Block bot-detection script so Angular initializes properly
            page.route("**/urlaccessvalidator.js", lambda route: route.fulfill(
                status=200, content_type="application/javascript",
                body="window.urlAccessValidator={validate:()=>Promise.resolve(true)};"
            ))

            # Step 1: Load dashboard and wait for Angular to fully init
            try:
                page.goto(f"https://{co}.keka.com/home", wait_until="domcontentloaded", timeout=25000)
                page.wait_for_load_state("networkidle", timeout=30000)
                page.wait_for_timeout(3000)
            except Exception as e:
                result["home_error"] = str(e)

            result["after_home_html_len"] = len(page.content())
            result["after_home_url"] = page.url

            # Step 2: Navigate to the claim page
            try:
                page.goto(f"https://{co}.keka.com/home/expense/claim/{claim_id}",
                          wait_until="domcontentloaded", timeout=25000)
                page.wait_for_load_state("networkidle", timeout=15000)
                page.wait_for_timeout(3000)
            except Exception as e:
                result["nav_error"] = str(e)

            result["page_url"]         = page.url
            result["page_html_sample"] = page.content()[:2000]

            # Step 3: Try JS fetches from inside the page (uses SPA's auth context)
            js_paths = [
                f"/api/v1/expense/claims/{claim_id}",
                f"/api/v2/expense/claims/{claim_id}",
                f"/api/v2/expense/attachments?claimId={claim_id}",
                f"/fis/api/v2/expense/claims/{claim_id}",
                f"/api/v1/expense/claims/{claim_id}/attachments",
            ]
            js_results = {}
            for path in js_paths:
                try:
                    r = page.evaluate(f"""
                        async () => {{
                            const resp = await fetch('{path}', {{credentials:'include'}});
                            return {{status: resp.status, body: (await resp.text()).substring(0, 600)}};
                        }}
                    """)
                    js_results[path] = r
                except Exception as e:
                    js_results[path] = {"error": str(e)}
            result["js_api_calls"] = js_results

            # All captured API calls
            result["all_api_calls"] = [r for r in all_reqs if "/api/" in r["url"] or "keka" in r["url"]][:40]
            result["file_urls"] = [r["url"] for r in all_resps
                                   if "blob.core.windows.net" in r["url"] or "/files/" in r["url"]]

            browser.close()

        return result

    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor(max_workers=1) as pool:
        res = await loop.run_in_executor(pool, _debug_claim)
    return res


# ── Logs Endpoints ────────────────────────────────────────────────────────────

_LOG_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "logs"))
_LOG_FILES = {
    "out":    os.path.join(_LOG_DIR, "app-out.log"),
    "err":    os.path.join(_LOG_DIR, "app-err.log"),
}


def _read_log_tail(path: str, lines: int = 500) -> list[str]:
    """Read the last `lines` lines from a log file. Returns list of strings."""
    if not os.path.exists(path):
        return []
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as f:
            all_lines = f.readlines()
        return [l.rstrip("\n") for l in all_lines[-lines:]]
    except Exception as e:
        return [f"[Error reading log: {e}]"]


@app.get("/logs")
def get_logs(
    lines: int = Query(500, ge=1, le=10000),
    file: str = Query("out", regex="^(out|err|both)$"),
    search: str = Query("", max_length=200),
):
    """
    Return last N lines from log file(s).
    file: "out" | "err" | "both"
    search: optional filter string (case-insensitive)
    """
    result = {}

    files_to_read = ["out", "err"] if file == "both" else [file]

    for f in files_to_read:
        path = _LOG_FILES.get(f, "")
        raw = _read_log_tail(path, lines)
        if search:
            sl = search.lower()
            raw = [l for l in raw if sl in l.lower()]
        result[f] = raw

    return {
        "logs": result,
        "lines_requested": lines,
        "search": search,
        "log_dir": _LOG_DIR,
    }


@app.get("/logs/stream")
async def stream_logs(
    file: str = Query("both", regex="^(out|err|both)$"),
    lines: int = Query(100, ge=1, le=2000),
):
    """
    SSE stream — sends new log lines as they appear.
    Polls both log files every second and pushes new content.
    """
    files_to_watch = ["out", "err"] if file == "both" else [file]

    # Track current file sizes
    offsets: dict = {}
    for f in files_to_watch:
        path = _LOG_FILES.get(f, "")
        if os.path.exists(path):
            offsets[f] = os.path.getsize(path)
        else:
            offsets[f] = 0

    # Send tail of existing content first
    async def event_generator():
        # Initial tail
        for f in files_to_watch:
            path = _LOG_FILES.get(f, "")
            initial = _read_log_tail(path, lines)
            for line in initial:
                tagged = f"[{f.upper()}] {line}"
                yield f"data: {json.dumps({'line': tagged, 'file': f, 'type': 'init'})}\n\n"

        yield f"data: {json.dumps({'line': '--- Live tail started ---', 'file': 'sys', 'type': 'separator'})}\n\n"

        # Poll for new content
        while True:
            await asyncio.sleep(1)
            for f in files_to_watch:
                path = _LOG_FILES.get(f, "")
                if not os.path.exists(path):
                    continue
                try:
                    size = os.path.getsize(path)
                    if size > offsets.get(f, 0):
                        with open(path, "r", encoding="utf-8", errors="replace") as fh:
                            fh.seek(offsets[f])
                            new_content = fh.read()
                        offsets[f] = size
                        for line in new_content.splitlines():
                            if line.strip():
                                tagged = f"[{f.upper()}] {line}"
                                yield f"data: {json.dumps({'line': tagged, 'file': f, 'type': 'new'})}\n\n"
                    elif size < offsets.get(f, 0):
                        # File was rotated/truncated
                        offsets[f] = 0
                except Exception:
                    pass

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


@app.get("/logs/info")
def logs_info():
    """Return log file sizes and last-modified timestamps."""
    info = {}
    for name, path in _LOG_FILES.items():
        if os.path.exists(path):
            stat = os.stat(path)
            info[name] = {
                "path": path,
                "size_bytes": stat.st_size,
                "size_kb": round(stat.st_size / 1024, 1),
                "last_modified": stat.st_mtime,
            }
        else:
            info[name] = {"path": path, "size_bytes": 0, "size_kb": 0, "last_modified": None}
    return info


# ── Serve React frontend (desktop + Railway production) ───────────────────────
# Checks two locations in priority order:
#   1. /app/static      — Railway production (Dockerfile copies dist here)
#   2. ../frontend/dist — local dev shortcut
# Mount only /assets statically so all API routes keep priority;
# catch-all serves index.html for SPA client-side routes.
import os as _os
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse as _FR

def _find_static_dir() -> str:
    candidates = [
        _os.path.join(_os.path.dirname(__file__), "static"),           # Railway
        _os.path.abspath(_os.path.join(_os.path.dirname(__file__), "..", "frontend", "dist")),  # local
    ]
    for p in candidates:
        if _os.path.isdir(p):
            return p
    return ""

_static_dir = _find_static_dir()
if _static_dir:
    _assets_dir = _os.path.join(_static_dir, "assets")
    if _os.path.isdir(_assets_dir):
        app.mount("/assets", StaticFiles(directory=_assets_dir, html=False), name="assets")

    @app.get("/", include_in_schema=False)
    async def _spa_root():
        return _FR(_os.path.join(_static_dir, "index.html"))

    @app.get("/{full_path:path}", include_in_schema=False)
    async def spa_fallback(full_path: str):
        fp = _os.path.join(_static_dir, full_path)
        if _os.path.isfile(fp):
            return _FR(fp)
        return _FR(_os.path.join(_static_dir, "index.html"))
